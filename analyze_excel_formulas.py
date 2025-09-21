#!/usr/bin/env python3
"""
Excel Formula Analysis Script
Analyzes WIZEMICE Excel file to identify astronomical value sources
"""

import openpyxl
import sys
import re
from collections import defaultdict

def analyze_excel_formulas(file_path):
    """Analyze Excel file formulas to identify issues"""
    print(f"🔍 ANALYZING EXCEL FILE: {file_path}")
    print("=" * 80)
    
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        print(f"📊 Loaded workbook with sheets: {workbook.sheetnames}")
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return
    
    # Focus on the main scenario sheet
    main_sheet = None
    for sheet_name in ['WIZEMICE Likest', 'Likest', 'Main']:
        if sheet_name in workbook.sheetnames:
            main_sheet = workbook[sheet_name]
            break
    
    if not main_sheet:
        main_sheet = workbook.active
        
    print(f"🎯 Analyzing main sheet: {main_sheet.title}")
    print()
    
    # Step 1: Examine input variables (F4, F5, F6)
    print("📋 STEP 1: INPUT VARIABLES ANALYSIS")
    print("-" * 50)
    input_vars = {}
    for cell_ref in ['F4', 'F5', 'F6', 'F7']:
        cell = main_sheet[cell_ref]
        value = cell.value
        formula = cell.formula if hasattr(cell, 'formula') else None
        input_vars[cell_ref] = {'value': value, 'formula': formula}
        print(f"   {cell_ref}: Value={value}, Formula={formula}")
    
    print()
    
    # Step 2: Examine Row 107 (Customer Growth)
    print("📋 STEP 2: ROW 107 ANALYSIS (Customer Growth)")
    print("-" * 50)
    row_107_formulas = {}
    sample_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    
    for col in sample_columns:
        cell_ref = f"{col}107"
        cell = main_sheet[cell_ref]
        value = cell.value
        formula = str(cell.value) if cell.value and str(cell.value).startswith('=') else None
        if hasattr(cell, 'formula') and cell.formula:
            formula = cell.formula
        row_107_formulas[cell_ref] = {'value': value, 'formula': formula}
        print(f"   {cell_ref}: Value={value}")
        if formula:
            print(f"       Formula: {formula}")
    
    print()
    
    # Step 3: Examine Cash Flow Row (161)
    print("📋 STEP 3: CASH FLOW ANALYSIS (Row 161)")
    print("-" * 50)
    cash_flow_formulas = {}
    cash_flow_columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB']
    
    for col in cash_flow_columns:
        cell_ref = f"{col}161"
        try:
            cell = main_sheet[cell_ref]
            value = cell.value
            formula = str(cell.value) if cell.value and str(cell.value).startswith('=') else None
            if hasattr(cell, 'formula') and cell.formula:
                formula = cell.formula
            cash_flow_formulas[cell_ref] = {'value': value, 'formula': formula}
            print(f"   {cell_ref}: Value={value}")
            if formula:
                print(f"       Formula: {formula}")
                # Check for potential amplification patterns
                if formula and any(pattern in formula.upper() for pattern in ['*', '**', 'POWER', 'EXP']):
                    print(f"       ⚠️  MULTIPLICATION/POWER DETECTED")
        except Exception as e:
            print(f"   {cell_ref}: Error reading cell: {e}")
    
    print()
    
    # Step 4: Look for intermediate calculation cells that might amplify
    print("📋 STEP 4: INTERMEDIATE CALCULATIONS ANALYSIS")
    print("-" * 50)
    
    # Check some key calculation areas
    intermediate_areas = [
        ('B120:B140', 'Revenue calculations'),
        ('AB120:AB140', 'Final calculations'),
        ('F48:F60', 'Price calculations'),
        ('C120:P140', 'Core business calculations')
    ]
    
    for area, description in intermediate_areas:
        print(f"\n   🔍 {description} ({area}):")
        try:
            start_col, start_row = openpyxl.utils.coordinate_from_string(area.split(':')[0])
            end_col, end_row = openpyxl.utils.coordinate_from_string(area.split(':')[1])
            
            start_col_idx = openpyxl.utils.column_index_from_string(start_col)
            end_col_idx = openpyxl.utils.column_index_from_string(end_col)
            
            suspicious_formulas = []
            
            for row in range(start_row, min(end_row + 1, start_row + 5)):  # Limit to first 5 rows
                for col_idx in range(start_col_idx, min(end_col_idx + 1, start_col_idx + 5)):  # Limit to first 5 cols
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    cell_ref = f"{col_letter}{row}"
                    cell = main_sheet[cell_ref]
                    
                    if cell.value is not None:
                        formula = str(cell.value) if str(cell.value).startswith('=') else None
                        if hasattr(cell, 'formula') and cell.formula:
                            formula = cell.formula
                            
                        if formula:
                            # Look for multiplication chains, power functions, or F variable references
                            if any(pattern in formula.upper() for pattern in ['F4', 'F5', 'F6', '*', 'POWER', 'EXP']):
                                suspicious_formulas.append((cell_ref, formula, cell.value))
            
            for cell_ref, formula, value in suspicious_formulas[:3]:  # Show first 3
                print(f"       {cell_ref}: {formula} = {value}")
                if any(var in formula.upper() for var in ['F4', 'F5', 'F6']):
                    print(f"           ⚠️  REFERENCES F VARIABLE")
                if formula.count('*') > 2:
                    print(f"           ⚠️  MULTIPLE MULTIPLICATIONS ({formula.count('*')} found)")
                    
        except Exception as e:
            print(f"       Error analyzing {area}: {e}")
    
    print()
    
    # Step 5: Check NPV formula
    print("📋 STEP 5: NPV FORMULA ANALYSIS")
    print("-" * 50)
    npv_cells = ['B12', 'B13']
    for cell_ref in npv_cells:
        try:
            cell = main_sheet[cell_ref]
            value = cell.value
            formula = str(cell.value) if cell.value and str(cell.value).startswith('=') else None
            if hasattr(cell, 'formula') and cell.formula:
                formula = cell.formula
            print(f"   {cell_ref}: Value={value}")
            if formula:
                print(f"       Formula: {formula}")
                # Check what range NPV is operating on
                if 'NPV' in formula.upper():
                    print(f"       ✅ NPV function found")
                if 'C161:' in formula.upper() or 'AL161' in formula.upper():
                    print(f"       ✅ References cash flow range")
        except Exception as e:
            print(f"   {cell_ref}: Error: {e}")
    
    print()
    print("🎯 ANALYSIS COMPLETE")
    print("=" * 80)

if __name__ == "__main__":
    # Use the most recent WIZEMICE file
    excel_file = "/home/paperspace/PROJECT/uploads/2a07dd9d-2048-4ff0-82e1-b6c7d74a6f68_WIZEMICE_FINAL_WORKING_MODEL.xlsx"
    analyze_excel_formulas(excel_file)
