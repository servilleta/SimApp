#!/usr/bin/env python3
"""
Monte Carlo Simulation API Test Client
======================================

This script tests the complete Monte Carlo simulation API workflow from your Mac.
It simulates how actual customers would use the API to upload Excel files,
configure simulations, and retrieve results.

Requirements:
- Python 3.6+
- requests library: pip install requests

Usage:
    python api_test_client.py --server http://209.51.170.185:8000 --api-key ak_4e968d72ca45909d97624140f9ba5d4a_sk_a73a4b3849a834e6bae20c23cccb074ef0ad2af04cc4cd155841484f52120323
"""

import requests
import json
import time
import argparse
import sys
import os
from pathlib import Path
import io
import openpyxl
from openpyxl import Workbook

class MonteCarloAPITester:
    def __init__(self, server_url, api_key):
        self.server_url = server_url.rstrip('/')
        self.api_key = api_key
        self.base_url = f"{self.server_url}/simapp-api"
        self.session = requests.Session()
        self.session.headers.update({
            'Authorization': f'Bearer {self.api_key}',
            'User-Agent': 'MonteCarloAPITester/1.0'
        })
        
        # Test data will be stored here
        self.model_id = None
        self.simulation_id = None
        
    def print_header(self, title):
        """Print a formatted header"""
        print(f"\n{'='*60}")
        print(f"🧪 {title}")
        print(f"{'='*60}")
    
    def print_step(self, step, description):
        """Print a test step"""
        print(f"\n🔸 Step {step}: {description}")
        print("-" * 40)
    
    def test_health_check(self):
        """Test API health endpoint"""
        self.print_step(1, "API Health Check")
        
        try:
            response = self.session.get(f"{self.base_url}/health", timeout=10)
            print(f"Status Code: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                print(f"✅ API is healthy")
                print(f"Version: {data.get('version', 'Unknown')}")
                print(f"GPU Available: {data.get('gpu_available', False)}")
                print(f"Uptime: {data.get('system_metrics', {}).get('uptime', 'Unknown')}")
                return True
            else:
                print(f"❌ Health check failed: {response.text}")
                return False
                
        except requests.exceptions.RequestException as e:
            print(f"❌ Connection error: {e}")
            return False
    
    def create_sample_excel_file(self):
        """Create a sample Excel file for testing"""
        self.print_step(2, "Creating Sample Excel File")
        
        # Create a simple portfolio risk model
        wb = Workbook()
        ws = wb.active
        ws.title = "Portfolio Model"
        
        # Header
        ws['A1'] = "Monte Carlo Portfolio Risk Model"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        
        # Input Variables (these will be Monte Carlo variables)
        ws['A3'] = "INPUT VARIABLES"
        ws['A3'].font = openpyxl.styles.Font(bold=True)
        
        ws['A4'] = "Market Volatility"
        ws['B4'] = 0.15  # 15% base volatility
        
        ws['A5'] = "Expected Return"
        ws['B5'] = 0.08  # 8% expected return
        
        ws['A6'] = "Risk-Free Rate"
        ws['B6'] = 0.03  # 3% risk-free rate
        
        ws['A7'] = "Initial Investment"
        ws['B7'] = 1000000  # $1M initial investment
        
        # Calculations
        ws['A9'] = "CALCULATIONS"
        ws['A9'].font = openpyxl.styles.Font(bold=True)
        
        ws['A10'] = "Adjusted Return"
        ws['B10'] = "=B5-B6"  # Expected return - risk-free rate
        
        ws['A11'] = "Risk-Adjusted Return"
        ws['B11'] = "=B10/B4"  # Sharpe-like ratio
        
        ws['A12'] = "Portfolio Value"
        ws['B12'] = "=B7*(1+B10)"  # Simple portfolio calculation
        
        ws['A13'] = "Profit/Loss"
        ws['B13'] = "=B12-B7"  # P&L calculation
        
        # Target cells (these will be our simulation outputs)
        ws['A15'] = "TARGET OUTPUTS"
        ws['A15'].font = openpyxl.styles.Font(bold=True)
        
        ws['A16'] = "Final Portfolio Value"
        ws['B16'] = "=B12"
        
        ws['A17'] = "Total Return %"
        ws['B17'] = "=(B12-B7)/B7*100"
        
        ws['A18'] = "Risk-Adjusted Performance"
        ws['B18'] = "=B11"
        
        # Save to bytes for upload
        file_buffer = io.BytesIO()
        wb.save(file_buffer)
        file_buffer.seek(0)
        
        print(f"✅ Created sample Excel file with:")
        print(f"   - Input variables: B4 (volatility), B5 (return), B6 (risk-free), B7 (investment)")
        print(f"   - Target outputs: B16 (portfolio value), B17 (return %), B18 (risk-adjusted)")
        
        return file_buffer.getvalue(), "portfolio_risk_model.xlsx"
    
    def test_file_upload(self):
        """Test file upload endpoint"""
        self.print_step(3, "File Upload Test")
        
        try:
            # Create sample file
            file_content, filename = self.create_sample_excel_file()
            
            # Prepare file upload
            files = {
                'file': (filename, file_content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            }
            
            print(f"Uploading file: {filename} ({len(file_content)} bytes)")
            
            response = self.session.post(
                f"{self.base_url}/models",
                files=files,
                timeout=30
            )
            
            print(f"Status Code: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.model_id = data.get('model_id')
                print(f"✅ File uploaded successfully")
                print(f"Model ID: {self.model_id}")
                print(f"Filename: {data.get('filename')}")
                print(f"Formulas Count: {data.get('formulas_count')}")
                print(f"Processing Status: {data.get('status')}")
                
                # Show detected variables if any
                if 'variables_detected' in data:
                    print(f"Variables detected: {len(data['variables_detected'])}")
                    for var in data['variables_detected'][:3]:  # Show first 3
                        print(f"  - {var.get('cell')}: {var.get('current_value')}")
                
                return True
            else:
                print(f"❌ Upload failed: {response.text}")
                return False
                
        except requests.exceptions.RequestException as e:
            print(f"❌ Upload error: {e}")
            return False
    
    def test_simulation_request(self):
        """Test simulation request with proper variable configuration"""
        self.print_step(4, "Simulation Request Test")
        
        if not self.model_id:
            print("❌ No model ID available. Upload a file first.")
            return False
        
        # Define Monte Carlo request with correct nested structure
        simulation_request = {
            "model_id": self.model_id,
            "simulation_config": {
                "iterations": 10000,
                "confidence_levels": [0.95, 0.99],
                "variables": [
                    {
                        "cell": "B4",
                        "name": "Market_Volatility",
                        "distribution": {
                            "type": "triangular",
                            "min": 0.05,    # 5% minimum volatility
                            "mode": 0.15,   # 15% most likely (current value)
                            "max": 0.35     # 35% maximum volatility
                        }
                    },
                    {
                        "cell": "B5", 
                        "name": "Expected_Return",
                        "distribution": {
                            "type": "normal",
                            "mean": 0.08,   # 8% expected return
                            "std": 0.02     # 2% standard deviation
                        }
                    },
                    {
                        "cell": "B6",
                        "name": "Risk_Free_Rate", 
                        "distribution": {
                            "type": "triangular",
                            "min": 0.01,    # 1% minimum
                            "mode": 0.03,   # 3% most likely (current value)
                            "max": 0.05     # 5% maximum
                        }
                    }
                ],
                "output_cells": ["B16", "B17", "B18"],  # Portfolio value, return %, risk-adjusted perf
                "webhook_url": None  # Optional for testing
            }
        }
        
        try:
            print(f"Sending simulation request for model: {self.model_id}")
            print(f"Variables: {len(simulation_request['simulation_config']['variables'])}")
            print(f"Output cells: {simulation_request['simulation_config']['output_cells']}")
            print(f"Iterations: {simulation_request['simulation_config']['iterations']:,}")
            
            response = self.session.post(
                f"{self.base_url}/simulations",
                json=simulation_request,
                timeout=30
            )
            
            print(f"Status Code: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                self.simulation_id = data.get('simulation_id')
                print(f"✅ Simulation started successfully")
                print(f"Simulation ID: {self.simulation_id}")
                print(f"Status: {data.get('status')}")
                print(f"Estimated completion: {data.get('estimated_completion')}")
                print(f"Credits consumed: {data.get('credits_consumed')}")
                
                if 'progress_url' in data:
                    print(f"Progress URL: {data['progress_url']}")
                
                return True
            else:
                print(f"❌ Simulation request failed: {response.text}")
                return False
                
        except requests.exceptions.RequestException as e:
            print(f"❌ Simulation request error: {e}")
            return False
    
    def test_simulation_progress(self):
        """Test simulation progress monitoring"""
        self.print_step(5, "Simulation Progress Monitoring")
        
        if not self.simulation_id:
            print("❌ No simulation ID available. Start a simulation first.")
            return False
        
        max_checks = 30  # Maximum number of progress checks
        check_interval = 2  # Seconds between checks
        
        for i in range(max_checks):
            try:
                response = self.session.get(
                    f"{self.base_url}/simulations/{self.simulation_id}/progress",
                    timeout=10
                )
                
                if response.status_code == 200:
                    data = response.json()
                    status = data.get('status', 'unknown')
                    progress = data.get('progress', 0)
                    
                    print(f"Check {i+1}/{max_checks}: Status={status}, Progress={progress}%")
                    
                    if status == 'completed':
                        print(f"✅ Simulation completed successfully!")
                        print(f"Execution time: {data.get('execution_time', 'Unknown')}")
                        return True
                    elif status == 'failed':
                        print(f"❌ Simulation failed: {data.get('error', 'Unknown error')}")
                        return False
                    elif status in ['running', 'queued']:
                        if i < max_checks - 1:  # Don't sleep on last iteration
                            time.sleep(check_interval)
                        continue
                    else:
                        print(f"⚠️ Unknown status: {status}")
                        
                else:
                    print(f"❌ Progress check failed: {response.status_code} - {response.text}")
                    
            except requests.exceptions.RequestException as e:
                print(f"❌ Progress check error: {e}")
                
        print(f"⏰ Timeout: Simulation did not complete within {max_checks * check_interval} seconds")
        return False
    
    def test_results_retrieval(self):
        """Test results retrieval and analysis"""
        self.print_step(6, "Results Retrieval Test")
        
        if not self.simulation_id:
            print("❌ No simulation ID available. Complete a simulation first.")
            return False
        
        try:
            response = self.session.get(
                f"{self.base_url}/simulations/{self.simulation_id}/results",
                timeout=15
            )
            
            print(f"Status Code: {response.status_code}")
            
            if response.status_code == 200:
                data = response.json()
                print(f"✅ Results retrieved successfully")
                
                # Display summary statistics
                results = data.get('results', {})
                print(f"\n📊 SIMULATION RESULTS SUMMARY:")
                print(f"Simulation ID: {data.get('simulation_id')}")
                print(f"Iterations: {data.get('iterations', 'Unknown'):,}")
                print(f"Execution time: {data.get('execution_time', 'Unknown')}")
                
                # Show statistics for each output cell
                for i, cell in enumerate(['B16', 'B17', 'B18']):
                    cell_names = ['Portfolio Value', 'Return %', 'Risk-Adjusted Performance']
                    print(f"\n{cell_names[i]} ({cell}):")
                    
                    if cell in results:
                        cell_data = results[cell]
                        print(f"  Mean: {cell_data.get('mean', 'N/A'):.4f}")
                        print(f"  Median: {cell_data.get('median', 'N/A'):.4f}")
                        print(f"  Std Dev: {cell_data.get('std_dev', 'N/A'):.4f}")
                        print(f"  Min: {cell_data.get('min_value', 'N/A'):.4f}")
                        print(f"  Max: {cell_data.get('max_value', 'N/A'):.4f}")
                        
                        # Show confidence intervals if available
                        if 'percentiles' in cell_data:
                            percentiles = cell_data['percentiles']
                            print(f"  95% CI: [{percentiles.get('5', 'N/A'):.4f}, {percentiles.get('95', 'N/A'):.4f}]")
                            print(f"  99% CI: [{percentiles.get('1', 'N/A'):.4f}, {percentiles.get('99', 'N/A'):.4f}]")
                
                # Show download options if available
                if 'download_urls' in data:
                    print(f"\n📥 DOWNLOAD OPTIONS:")
                    downloads = data['download_urls']
                    for format_type, url in downloads.items():
                        print(f"  {format_type.upper()}: {url}")
                
                return True
            else:
                print(f"❌ Results retrieval failed: {response.text}")
                return False
                
        except requests.exceptions.RequestException as e:
            print(f"❌ Results retrieval error: {e}")
            return False
    
    def test_download_reports(self):
        """Test report downloads"""
        self.print_step(7, "Report Downloads Test")
        
        if not self.simulation_id:
            print("❌ No simulation ID available. Complete a simulation first.")
            return False
        
        # Test downloading different formats
        formats_to_test = ['json', 'xlsx', 'pdf']
        
        for format_type in formats_to_test:
            try:
                print(f"\nTesting {format_type.upper()} download...")
                
                # First get download token
                response = self.session.post(
                    f"{self.base_url}/simulations/{self.simulation_id}/generate-download-token",
                    timeout=10
                )
                
                if response.status_code == 200:
                    token_data = response.json()
                    download_token = token_data.get('download_token')
                    
                    # Test the download
                    download_response = self.session.get(
                        f"{self.base_url}/download/{download_token}/{format_type}",
                        timeout=30
                    )
                    
                    if download_response.status_code == 200:
                        content_length = len(download_response.content)
                        print(f"✅ {format_type.upper()} download successful ({content_length:,} bytes)")
                        
                        # Save file for inspection
                        filename = f"test_results_{self.simulation_id}.{format_type}"
                        with open(filename, 'wb') as f:
                            f.write(download_response.content)
                        print(f"   Saved as: {filename}")
                        
                    else:
                        print(f"❌ {format_type.upper()} download failed: {download_response.status_code}")
                        
                else:
                    print(f"❌ Failed to get download token: {response.text}")
                    
            except requests.exceptions.RequestException as e:
                print(f"❌ {format_type.upper()} download error: {e}")
        
        return True
    
    def run_full_test_suite(self):
        """Run the complete API test suite"""
        self.print_header("MONTE CARLO API COMPREHENSIVE TEST SUITE")
        
        print(f"🌐 Server: {self.server_url}")
        print(f"🔑 API Key: {self.api_key[:20]}...")
        print(f"📡 Base URL: {self.base_url}")
        
        # Track test results
        tests = [
            ("Health Check", self.test_health_check),
            ("File Upload", self.test_file_upload), 
            ("Simulation Request", self.test_simulation_request),
            ("Progress Monitoring", self.test_simulation_progress),
            ("Results Retrieval", self.test_results_retrieval),
            ("Report Downloads", self.test_download_reports)
        ]
        
        results = {}
        
        for test_name, test_func in tests:
            try:
                results[test_name] = test_func()
            except Exception as e:
                print(f"❌ {test_name} crashed: {e}")
                results[test_name] = False
        
        # Print final summary
        self.print_header("TEST RESULTS SUMMARY")
        
        passed = sum(1 for result in results.values() if result)
        total = len(results)
        
        for test_name, result in results.items():
            status = "✅ PASS" if result else "❌ FAIL"
            print(f"{status} {test_name}")
        
        print(f"\n🎯 OVERALL RESULT: {passed}/{total} tests passed")
        
        if passed == total:
            print("🎉 ALL TESTS PASSED! The API is working correctly.")
        else:
            print("⚠️ Some tests failed. Check the logs above for details.")
        
        # Save test results
        results_file = f"api_test_results_{int(time.time())}.json"
        test_data = {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "server": self.server_url,
            "model_id": self.model_id,
            "simulation_id": self.simulation_id,
            "results": results,
            "summary": {
                "passed": passed,
                "total": total,
                "success_rate": passed / total if total > 0 else 0
            }
        }
        
        with open(results_file, 'w') as f:
            json.dump(test_data, f, indent=2)
        
        print(f"📄 Test results saved to: {results_file}")
        
        return passed == total

def main():
    parser = argparse.ArgumentParser(description="Monte Carlo API Test Client")
    parser.add_argument(
        '--server', 
        default='http://209.51.170.185:8000',
        help='Server URL (default: http://209.51.170.185:8000)'
    )
    parser.add_argument(
        '--api-key',
        default='ak_0f345df6f8af9ea80140bf434fdba478_sk_03ddc015b97aca737a1cb690d4f41fa86d3877c957eef8362713290055a7964a',
        help='API key for authentication (default: demo key)'
    )
    parser.add_argument(
        '--test',
        choices=['health', 'upload', 'simulation', 'progress', 'results', 'downloads', 'all'],
        default='all',
        help='Specific test to run (default: all)'
    )
    
    args = parser.parse_args()
    
    print("🚀 Monte Carlo API Test Client")
    print("=" * 50)
    
    # Check dependencies
    try:
        import openpyxl
    except ImportError:
        print("❌ Missing dependency: openpyxl")
        print("Install with: pip install openpyxl")
        sys.exit(1)
    
    # Initialize tester
    tester = MonteCarloAPITester(args.server, args.api_key)
    
    # Run requested test(s)
    if args.test == 'all':
        success = tester.run_full_test_suite()
        sys.exit(0 if success else 1)
    else:
        # Run individual test
        test_methods = {
            'health': tester.test_health_check,
            'upload': tester.test_file_upload,
            'simulation': tester.test_simulation_request,
            'progress': tester.test_simulation_progress,
            'results': tester.test_results_retrieval,
            'downloads': tester.test_download_reports
        }
        
        if args.test in test_methods:
            success = test_methods[args.test]()
            sys.exit(0 if success else 1)
        else:
            print(f"❌ Unknown test: {args.test}")
            sys.exit(1)

if __name__ == "__main__":
    main()
