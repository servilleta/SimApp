#!/usr/bin/env python3
"""Quick ad-hoc comparison between enhanced (GPU) and Arrow engines using a GPU-friendly workbook.

Run with:
    PYTHONPATH=backend BENCH_ITERATIONS=2000000 python3 backend/benchmarks/run_gpu_arrow_compare.py
"""
import asyncio, json, os, time, uuid, pathlib
from openpyxl import Workbook
from simulation.schemas import VariableConfig, ConstantConfig
from simulation.service import run_simulation_with_engine
from config import settings

ITER = int(os.environ.get("BENCH_ITERATIONS", "2000000"))

uploads = pathlib.Path(settings.UPLOAD_DIR)
uploads.mkdir(exist_ok=True)
file_id = str(uuid.uuid4())
wb_path = uploads / f"{file_id}_gpu.xlsx"

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = "Calc"
ws["A1"] = 3.14
ws["B1"] = 10
ws["C1"] = 20
ws["D1"] = 30
ws["E1"] = "=B1+C1-D1"
ws["F1"] = "=SQRT(ABS(E1))*2"
ws["G1"] = "=POWER(F1,2)+A1"
wb.save(wb_path)

# parser artefacts
formulas = {"Calc": {"E1": "=B1+C1-D1", "F1": "=SQRT(ABS(E1))*2", "G1": "=POWER(F1,2)+A1"}}
json.dump(formulas, open(uploads / f"{file_id}_formulas.json", "w"), indent=2)

sheet_data = [{
    "sheet_name": "Calc",
    "grid_data": [[
        {"value": 3.14, "formula": None, "is_formula_cell": False, "coordinate": "A1"},
        {"value": 10, "formula": None, "is_formula_cell": False, "coordinate": "B1"},
        {"value": 20, "formula": None, "is_formula_cell": False, "coordinate": "C1"},
        {"value": 30, "formula": None, "is_formula_cell": False, "coordinate": "D1"},
        {"value": None, "formula": "=B1+C1-D1", "is_formula_cell": True, "coordinate": "E1"},
        {"value": None, "formula": "=SQRT(ABS(E1))*2", "is_formula_cell": True, "coordinate": "F1"},
        {"value": None, "formula": "=POWER(F1,2)+A1", "is_formula_cell": True, "coordinate": "G1"},
    ]]
}]
json.dump(sheet_data, open(uploads / f"{file_id}_sheet_data.json", "w"))

print(f"Workbook created: {wb_path}\nIterations: {ITER}")

async def bench(engine: str):
    sim_id = str(uuid.uuid4())
    vars_cfg = [
        VariableConfig(sheet_name="Calc", name="B1", min_value=8, most_likely=10, max_value=12),
        VariableConfig(sheet_name="Calc", name="C1", min_value=18, most_likely=20, max_value=22),
        VariableConfig(sheet_name="Calc", name="D1", min_value=28, most_likely=30, max_value=32),
    ]
    consts = [ConstantConfig(sheet_name="Calc", name="A1", value=3.14)]
    t0 = time.perf_counter()
    await run_simulation_with_engine(sim_id, str(wb_path), vars_cfg, consts, "Calc!G1", ITER, engine_type=engine)
    return time.perf_counter() - t0

async def main():
    for eng in ("enhanced", "arrow"):
        try:
            dur = await bench(eng)
            print(f"{eng} finished in {dur:.2f} seconds")
        except Exception as e:
            print(f"{eng} failed: {e}")

if __name__ == "__main__":
    asyncio.run(main()) 