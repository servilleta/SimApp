#!/usr/bin/env python3
"""
Benchmark different Monte Carlo simulation engines under various workloads.

The script creates a small synthetic Excel workbook with a single variable
and a simple formula, then measures the execution time of each engine when
running that file.  Optionally, it can execute multiple simulations in
parallel to emulate several users working at the same time.

Results are printed to stdout and also saved to JSON in backend/results.
"""

import asyncio
import json
import os
import time
import uuid
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook

# Local imports – assume script is executed inside backend container
from simulation.schemas import VariableConfig, ConstantConfig  # type: ignore
from simulation.service import run_simulation_with_engine  # type: ignore
from config import settings  # type: ignore

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def _make_uploads_dir() -> Path:
    uploads_dir = Path(settings.UPLOAD_DIR)
    uploads_dir.mkdir(parents=True, exist_ok=True)
    return uploads_dir


def _build_synthetic_workbook(file_id: str) -> Path:
    """Create a tiny workbook with one stochastic variable and one formula."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # A1 constant, B1 stochastic variable, C1 formula A1+B1, D1 target (C1*2)
    ws["A1"] = 5
    ws["B1"] = 10
    ws["C1"] = "=A1+B1"
    ws["D1"] = "=C1*2"

    uploads_dir = _make_uploads_dir()
    excel_path = uploads_dir / f"{file_id}_benchmark.xlsx"
    wb.save(excel_path)
    return excel_path


def _save_supporting_json(file_id: str):
    """Mimic the parser output so that the simulation service can work."""
    uploads_dir = Path(settings.UPLOAD_DIR)

    # formulas.json
    formulas = {
        "TestSheet": {
            "C1": "=A1+B1",
            "D1": "=C1*2",
        }
    }
    with open(uploads_dir / f"{file_id}_formulas.json", "w") as fp:
        json.dump(formulas, fp)

    # sheet_data.json (minimal grid containing four cells)
    sheet_data = [
        {
            "sheet_name": "TestSheet",
            "grid_data": [
                [
                    {"value": 5, "formula": None, "is_formula_cell": False, "coordinate": "A1"},
                    {"value": 10, "formula": None, "is_formula_cell": False, "coordinate": "B1"},
                    {"value": 15, "formula": "=A1+B1", "is_formula_cell": True, "coordinate": "C1"},
                    {"value": 30, "formula": "=C1*2", "is_formula_cell": True, "coordinate": "D1"},
                ]
            ],
        }
    ]
    with open(uploads_dir / f"{file_id}_sheet_data.json", "w") as fp:
        json.dump(sheet_data, fp)


async def _run_single_simulation(engine: str, file_path: str, iterations: int) -> float:
    """Run one simulation and return wall-clock seconds."""
    sim_id = str(uuid.uuid4())

    variable_cfg = [
        VariableConfig(
            sheet_name="TestSheet",
            name="B1",
            min_value=8.0,
            most_likely=10.0,
            max_value=12.0,
        )
    ]

    start = time.perf_counter()
    try:
        await run_simulation_with_engine(
            sim_id=sim_id,
            file_path=file_path,
            mc_inputs=variable_cfg,
            constants=[],
            target_cell="TestSheet!D1",
            iterations=iterations,
            engine_type=engine,
        )
    except Exception as exc:
        # Propagate exception after time capture so caller can log failure
        duration = time.perf_counter() - start
        raise RuntimeError(f"Engine '{engine}' failed after {duration:.2f}s: {exc}") from exc

    duration = time.perf_counter() - start
    return duration


async def _benchmark_engine(engine: str, file_path: str, iterations: int, concurrency: int) -> Dict[str, float]:
    """Return timing statistics for a single engine with given parallelism."""
    tasks = [_run_single_simulation(engine, file_path, iterations) for _ in range(concurrency)]
    overall_start = time.perf_counter()
    results: List[float] = await asyncio.gather(*tasks, return_exceptions=True)  # type: ignore
    overall_elapsed = time.perf_counter() - overall_start

    # Filter out failures and convert Exceptions to None for reporting
    successful = [r for r in results if isinstance(r, (int, float))]
    failed = len(results) - len(successful)

    stats = {
        "engine": engine,
        "concurrency": concurrency,
        "iterations": iterations,
        "runs": len(results),
        "failures": failed,
        "wall_time_total": overall_elapsed,
    }

    if successful:
        stats.update(
            {
                "wall_time_min": min(successful),
                "wall_time_max": max(successful),
                "wall_time_avg": sum(successful) / len(successful),
            }
        )
    return stats


async def main():
    iterations = int(os.environ.get("BENCH_ITERATIONS", "1000"))
    concurrency_levels = [1, 4]
    engines = ["standard", "enhanced", "arrow"]

    file_id = str(uuid.uuid4())
    excel_path = _build_synthetic_workbook(file_id)
    _save_supporting_json(file_id)

    print(f"📁 Test workbook created at {excel_path}")
    results = []

    for engine in engines:
        for conc in concurrency_levels:
            print(f"\n🔬 Benchmarking {engine} engine | iterations={iterations} | concurrency={conc} …")
            try:
                stats = await _benchmark_engine(engine, str(excel_path), iterations, conc)
                results.append(stats)
                ok = "✅" if stats["failures"] == 0 else "⚠️"
                print(f"{ok} {engine} | conc={conc}: avg={stats.get('wall_time_avg', 'n/a'):.3f}s total={stats['wall_time_total']:.3f}s failures={stats['failures']}")
            except Exception as exc:
                print(f"❌ {engine} failed: {exc}")
                results.append({
                    "engine": engine,
                    "concurrency": conc,
                    "iterations": iterations,
                    "error": str(exc),
                })

    # Persist results
    results_dir = Path("backend/results")
    results_dir.mkdir(parents=True, exist_ok=True)
    out_path = results_dir / f"benchmark_{file_id}.json"
    with open(out_path, "w") as fp:
        json.dump(results, fp, indent=2)
    print(f"\n💾 Results saved to {out_path}")


if __name__ == "__main__":
    asyncio.run(main()) 