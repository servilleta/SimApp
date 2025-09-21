#!/usr/bin/env python3
"""
High-load benchmark that is biased toward GPU strength.

It runs the enhanced GPU engine and the Arrow engine with a *large* number of
iterations so that kernel-launch overhead is amortised and the GPU can shine.

Usage (inside backend container or with `PYTHONPATH=backend`):

    BENCH_ITERATIONS=500000 python3 backend/benchmarks/benchmark_gpu_vs_arrow.py

It writes a JSON results file into backend/results/ and prints a summary.
"""

import asyncio
import json
import os
import time
import uuid
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook

# Local imports – executed via PYTHONPATH=backend
from simulation.schemas import VariableConfig, ConstantConfig  # type: ignore
from simulation.service import run_simulation_with_engine  # type: ignore
from config import settings  # type: ignore

# ---------------------------------------------------------------------------
# Helpers to create a workbook & stub parser artefacts ― same as previous bench
# ---------------------------------------------------------------------------

UPLOADS = Path(settings.UPLOAD_DIR)
UPLOADS.mkdir(parents=True, exist_ok=True)


def _mk_workbook(file_id: str) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"

    # A handful of non-trivial formulas to stress floating-point math
    ws["A1"] = 5.0  # constant
    ws["B1"] = 10.0  # variable
    ws["C1"] = "=LN(B1) + SQRT(ABS(B1))"  # non-linear
    ws["D1"] = "=POWER(C1,2) + SIN(B1)"   # trig + power
    ws["E1"] = "=D1 * 2"                  # target cell

    path = UPLOADS / f"{file_id}_gpu_bench.xlsx"
    wb.save(path)
    return path


def _write_json_meta(file_id: str):
    formulas = {
        "Model": {
            "C1": "=LN(B1) + SQRT(ABS(B1))",
            "D1": "=POWER(C1,2) + SIN(B1)",
            "E1": "=D1 * 2",
        }
    }
    with open(UPLOADS / f"{file_id}_formulas.json", "w") as fp:
        json.dump(formulas, fp)

    # Minimal sheet_data.json (row with 5 cells)
    sheet_data = [
        {
            "sheet_name": "Model",
            "grid_data": [
                [
                    {"value": 5.0, "formula": None, "is_formula_cell": False, "coordinate": "A1"},
                    {"value": 10.0, "formula": None, "is_formula_cell": False, "coordinate": "B1"},
                    {"value": None, "formula": "=LN(B1) + SQRT(ABS(B1))", "is_formula_cell": True, "coordinate": "C1"},
                    {"value": None, "formula": "=POWER(C1,2) + SIN(B1)", "is_formula_cell": True, "coordinate": "D1"},
                    {"value": None, "formula": "=D1 * 2", "is_formula_cell": True, "coordinate": "E1"},
                ]
            ],
        }
    ]
    with open(UPLOADS / f"{file_id}_sheet_data.json", "w") as fp:
        json.dump(sheet_data, fp)


async def _run_once(engine: str, file_path: str, iterations: int) -> float:
    sim_id = str(uuid.uuid4())

    variables = [
        VariableConfig(
            sheet_name="Model",
            name="B1",
            min_value=8.0,
            most_likely=10.0,
            max_value=12.0,
        )
    ]
    constants = [
        ConstantConfig(sheet_name="Model", name="A1", value=5.0)
    ]

    t0 = time.perf_counter()
    await run_simulation_with_engine(
        sim_id=sim_id,
        file_path=file_path,
        mc_inputs=variables,
        constants=constants,
        target_cell="Model!E1",
        iterations=iterations,
        engine_type=engine,
    )
    return time.perf_counter() - t0


async def main():
    iterations = int(os.environ.get("BENCH_ITERATIONS", "200000"))
    engines = ["enhanced", "arrow"]

    file_id = str(uuid.uuid4())
    workbook_path = _mk_workbook(file_id)
    _write_json_meta(file_id)
    print(f"Workbook for benchmark: {workbook_path} (iterations={iterations})")

    results: List[Dict[str, float]] = []

    for eng in engines:
        print(f"\n⏱  Running {eng} …")
        try:
            wall = await _run_once(eng, str(workbook_path), iterations)
            results.append({"engine": eng, "iterations": iterations, "elapsed_sec": wall})
            print(f"   → {wall:.2f} s")
        except Exception as exc:
            results.append({"engine": eng, "iterations": iterations, "error": str(exc)})
            print(f"   ✗ failed: {exc}")

    # Persist JSON
    res_dir = Path("backend/results")
    res_dir.mkdir(parents=True, exist_ok=True)
    out = res_dir / f"gpu_vs_arrow_{file_id}.json"
    with open(out, "w") as fp:
        json.dump(results, fp, indent=2)
    print(f"\nResults written to {out}")

    # Quick winner statement
    ok = [r for r in results if "elapsed_sec" in r]
    if len(ok) == 2:
        winner = min(ok, key=lambda d: d["elapsed_sec"])
        print(f"🏆  Faster engine: {winner['engine']} ({winner['elapsed_sec']:.2f} s)")


if __name__ == "__main__":
    asyncio.run(main()) 