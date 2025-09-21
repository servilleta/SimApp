#!/usr/bin/env python3
"""
🔍 DIRECT SIMULATION SERVICE TEST
Test the simulation service directly with manually created data structures.
"""

import asyncio
import tempfile
import uuid
import json
import os
from openpyxl import Workbook
from simulation.schemas import VariableConfig, SimulationRequest
from simulation.service import run_monte_carlo_simulation_task, SIMULATION_RESULTS_STORE
from config import settings

async def test_simulation_service_directly():
    """Test simulation service directly by creating required files manually"""
    
    print("🚀 STARTING DIRECT SIMULATION SERVICE TEST")
    print("=" * 60)
    
    # Create test file ID
    file_id = str(uuid.uuid4())
    
    # Create uploads directory if it doesn't exist
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    
    # Create simple Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    ws['A1'] = 5        # Constant value
    ws['B1'] = 10       # This will be our Monte Carlo variable  
    ws['C1'] = "=A1+B1" # Simple formula
    ws['D1'] = "=C1*2"  # Target formula
    
    # Save Excel file
    excel_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_test_simulation.xlsx")
    wb.save(excel_file_path)
    
    print(f"📁 Created Excel file: {excel_file_path}")
    print(f"📊 Excel content:")
    print(f"   A1 = 5 (constant)")
    print(f"   B1 = 10 (Monte Carlo variable)")
    print(f"   C1 = =A1+B1")
    print(f"   D1 = =C1*2 (target)")
    
    try:
        # Create formulas data manually
        formulas_data = {
            "TestSheet": {
                "C1": "=A1+B1",
                "D1": "=C1*2"
            }
        }
        
        # Save formulas file
        formulas_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_formulas.json")
        with open(formulas_file_path, "w") as f:
            json.dump(formulas_data, f)
        
        print(f"✅ Created formulas file: {formulas_file_path}")
        
        # Create sheet data manually (simulate what the parser would create)
        sheet_data = [
            {
                "sheet_name": "TestSheet",
                "grid_data": [
                    [
                        {"value": 5, "formula": None, "is_formula_cell": False, "coordinate": "A1"},
                        {"value": 10, "formula": None, "is_formula_cell": False, "coordinate": "B1"},
                        {"value": 15, "formula": "=A1+B1", "is_formula_cell": True, "coordinate": "C1"},
                        {"value": 30, "formula": "=C1*2", "is_formula_cell": True, "coordinate": "D1"}
                    ]
                ]
            }
        ]
        
        # Save sheet data file
        sheet_data_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_sheet_data.json")
        with open(sheet_data_file_path, "w") as f:
            json.dump(sheet_data, f)
        
        print(f"✅ Created sheet data file: {sheet_data_file_path}")
        
        # Create simulation request
        simulation_id = str(uuid.uuid4())
        
        simulation_request = SimulationRequest(
            simulation_id=simulation_id,
            file_id=file_id,
            variables=[
                VariableConfig(
                    sheet_name="TestSheet",
                    name="B1",
                    min_value=8.0,
                    most_likely=10.0,
                    max_value=12.0
                )
            ],
            result_cell_sheet_name="TestSheet",
            result_cell_coordinate="D1",
            iterations=5  # Small number for testing
        )
        
        print(f"\n🔍 Created simulation request:")
        print(f"   Simulation ID: {simulation_id}")
        print(f"   File ID: {file_id}")
        print(f"   Target: TestSheet!D1")
        print(f"   Variables: B1 (8-12 range)")
        print(f"   Iterations: 5")
        
        # Run simulation
        print(f"\n🔍 Running simulation...")
        
        try:
            await run_monte_carlo_simulation_task(simulation_request)
            print(f"✅ Simulation task completed!")
            
        except Exception as e:
            print(f"❌ Simulation task failed: {e}")
            import traceback
            traceback.print_exc()
            return
        
        # Check results
        print(f"\n🔍 Checking simulation results...")
        
        if simulation_id in SIMULATION_RESULTS_STORE:
            stored_result = SIMULATION_RESULTS_STORE[simulation_id]
            print(f"✅ Found stored result:")
            print(f"   Status: {stored_result.status}")
            print(f"   Message: {stored_result.message}")
            
            if hasattr(stored_result, 'results') and stored_result.results:
                results = stored_result.results
                print(f"\n📊 Simulation Results:")
                print(f"   Mean: {results.mean}")
                print(f"   Std Dev: {results.std_dev}")
                print(f"   Min: {results.min_value}")
                print(f"   Max: {results.max_value}")
                print(f"   Iterations: {results.iterations_run}")
                print(f"   Errors: {results.errors}")
                
                # Analyze results
                if results.mean == 0.0 and results.std_dev == 0.0:
                    print(f"\n❌ CRITICAL ISSUE CONFIRMED: All results are zero!")
                    print(f"❌ This is the zero results bug we need to fix!")
                    
                    # Additional debugging
                    if hasattr(results, 'histogram') and results.histogram:
                        hist = results.histogram
                        print(f"\n🔍 Histogram data:")
                        print(f"   Counts: {hist.get('counts', 'N/A')}")
                        print(f"   Bin edges: {hist.get('bin_edges', 'N/A')}")
                    
                elif results.mean > 0 and results.std_dev >= 0:
                    print(f"\n✅ SUCCESS: Results look correct!")
                    
                    # Expected analysis
                    print(f"\n📊 Expected vs Actual:")
                    print(f"   Expected range: 26-34 (for D1 = (5+B1)*2 with B1=8-12)")
                    print(f"   Actual range: {results.min_value} - {results.max_value}")
                    print(f"   Expected mean: ~30")
                    print(f"   Actual mean: {results.mean}")
                    
                    if 26 <= results.mean <= 34:
                        print(f"✅ Mean is in expected range!")
                    else:
                        print(f"❌ Mean is outside expected range!")
                        
                else:
                    print(f"\n⚠️ UNEXPECTED: Results have unusual values")
                    
            else:
                print(f"❌ No results data found in stored result")
                
        else:
            print(f"❌ Simulation ID not found in results store")
            print(f"Available IDs: {list(SIMULATION_RESULTS_STORE.keys())}")
        
    finally:
        # Cleanup
        cleanup_files = [excel_file_path, formulas_file_path, sheet_data_file_path]
        for file_path in cleanup_files:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"🧹 Cleaned up: {file_path}")
            except:
                pass
    
    print("\n" + "=" * 60)
    print("🏁 DIRECT SIMULATION SERVICE TEST COMPLETED")

if __name__ == "__main__":
    asyncio.run(test_simulation_service_directly()) 