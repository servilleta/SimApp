#!/usr/bin/env python3
"""
Generic Monte Carlo Dependency Tracker

This module provides platform-level functionality to identify which cells in any Excel file
should remain as formulas (not be loaded as constants) based on Monte Carlo variable dependencies.

This replaces hardcoded file-specific fixes with a generic solution that works with any Excel structure.
"""

import logging
import re
from typing import Dict, Set, Tuple, List, Any
from openpyxl import load_workbook
from pathlib import Path

logger = logging.getLogger(__name__)

class MonteCarloDependencyTracker:
    """Tracks dependencies between Monte Carlo variables and Excel formulas"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.formulas = {
        }
        self.dependency_graph = {
        }
        
    def analyze_dependencies(self, monte_carlo_variables: List[Tuple[str, str]]) -> Set[Tuple[str, str]]:
        """Analyze which cells should be excluded from constants based on MC variable dependencies
        
        Args:
            monte_carlo_variables: List of (sheet_name, cell_address) tuples for MC variables
            
        Returns:
            Set of (sheet_name, cell_address) tuples that should remain as formulas
        """
        logger.info(f"🔍 [DEPENDENCY_TRACKER] Analyzing dependencies for {len(monte_carlo_variables)} MC variables")
        
        try:
            # Load workbook and extract all formulas
            self.workbook = load_workbook(self.file_path, data_only=False)
            self._extract_all_formulas()
            
            # Build dependency graph
            self._build_dependency_graph()
            
            # Find all cells that depend on Monte Carlo variables
            dependent_cells = self._find_dependent_cells(monte_carlo_variables)
            
            logger.info(f"✅ [DEPENDENCY_TRACKER] Found {len(dependent_cells)} cells dependent on MC variables")
            return dependent_cells
            
        except Exception as e:
            logger.error(f"❌ [DEPENDENCY_TRACKER] Error analyzing dependencies: {e}")
            return set()
    
    def _extract_all_formulas(self):
        """Extract all formulas from all sheets"""
        self.formulas = {
        }
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        cell_key = (sheet_name, cell.coordinate)
                        self.formulas[cell_key] = cell.value
    
    def _build_dependency_graph(self):
        """Build graph of which cells reference which other cells"""
        self.dependency_graph = {
        }
        
        for (sheet_name, cell_addr), formula in self.formulas.items():
            # Extract cell references from formula
            referenced_cells = self._extract_cell_references(formula, sheet_name)
            self.dependency_graph[(sheet_name, cell_addr)] = referenced_cells
    
    def _extract_cell_references(self, formula: str, current_sheet: str) -> Set[Tuple[str, str]]:
        """Extract all cell references from a formula"""
        references = set()
        
        # Pattern to match cell references (e.g., A1, $A$1, Sheet!A1, 'Sheet Name'!A1)
        cell_pattern = r"(?:(?:'([^']+)'|([^!\\s]+))!)?(\$?[A-Z]+\$?\d+)"
        
        matches = re.findall(cell_pattern, formula)
        
        for match in matches:
            quoted_sheet, unquoted_sheet, cell_ref = match
            
            # Determine sheet name
            if quoted_sheet:
                sheet_name = quoted_sheet
            elif unquoted_sheet:
                sheet_name = unquoted_sheet
            else:
                sheet_name = current_sheet  # Same sheet reference
            
            # Clean up cell reference (remove $ signs)
            clean_cell_ref = cell_ref.replace('$', '')
            references.add((sheet_name, clean_cell_ref))
        return references
    
    def _find_dependent_cells(self, monte_carlo_variables: List[Tuple[str, str]]) -> Set[Tuple[str, str]]:
        """Find all cells that directly or indirectly depend on Monte Carlo variables"""
        mc_variable_set = set(monte_carlo_variables)
        dependent_cells = set()
        visited = set()
        
        # Use breadth-first search to find all dependent cells
        queue = list(mc_variable_set)
        
        while queue:
            current_cell = queue.pop(0)
            
            if current_cell in visited:
                continue
                
            visited.add(current_cell)
            
            # Find all cells that reference this cell
            for cell, references in self.dependency_graph.items():
                if current_cell in references:
                    dependent_cells.add(cell)
                    queue.append(cell)
        
        # Also include any cells that directly reference MC variables
        for cell, references in self.dependency_graph.items():
            if any(mc_var in references for mc_var in mc_variable_set):
                dependent_cells.add(cell)
        
        return dependent_cells

def get_monte_carlo_dependent_cells(file_path: str, monte_carlo_variables: List[Tuple[str, str]]) -> Set[Tuple[str, str]]:
    """Convenience function to get cells that should be excluded from constants
    
    Args:
        file_path: Path to Excel file
        monte_carlo_variables: List of (sheet_name, cell_address) for MC variables
        
    Returns:
        Set of (sheet_name, cell_address) that should remain as formulas
    """
    tracker = MonteCarloDependencyTracker(file_path)
    return tracker.analyze_dependencies(monte_carlo_variables) 

def get_monte_carlo_direct_dependents(
    file_path: str, 
    monte_carlo_variables: List[Tuple[str, str]]
) -> Set[Tuple[str, str]]:
    """
    Get cells that DIRECTLY reference Monte Carlo variables (not entire dependency chain).
    
    This function identifies only cells that have formulas directly referencing F4, F5, F6
    rather than the entire transitive dependency chain. This prevents re-evaluation of
    complex financial calculations that should use Excel's pre-calculated values.
    
    Args:
        file_path: Path to Excel file
        monte_carlo_variables: List of (sheet, cell) tuples for MC variables
        
    Returns:
        Set of (sheet, cell) tuples that directly reference MC variables
    """
    logger.info(f"🎯 [DIRECT_DEPENDENTS] Finding direct dependents of {len(monte_carlo_variables)} MC variables")
    
    import re  # Import regex for pattern matching
    
    try:
        # Get all formulas from the file
        workbook = load_workbook(file_path, data_only=False)
        direct_dependents = set()
        
        # Create EXACT reference patterns for ANY Monte Carlo variables
        exact_mc_references = set()
        for sheet_name, cell_coord in monte_carlo_variables:
            # Create all possible Excel reference formats for this specific cell
            col = cell_coord[0]  # e.g., 'F' from 'F4'
            row = cell_coord[1:]  # e.g., '4' from 'F4'
            
            exact_mc_references.add(cell_coord)                # F4
            exact_mc_references.add(f"${cell_coord}")          # $F4  
            exact_mc_references.add(f"{col}${row}")            # F$4
            exact_mc_references.add(f"${col}${row}")           # $F$4
            
        logger.info(f"🔍 [DIRECT_DEPENDENTS] Looking for EXACT references to Monte Carlo variables: {exact_mc_references}")
        
        total_formulas_checked = 0
        direct_dependent_count = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        total_formulas_checked += 1
                        formula = cell.value
                        
                        # Check if formula contains EXACT reference to F4, F5, or F6 only
                        contains_exact_mc_ref = False
                        for exact_ref in exact_mc_references:
                            # Use word boundary matching to avoid partial matches
                            if re.search(rf'\b{re.escape(exact_ref)}\b', formula):
                                contains_exact_mc_ref = True
                                break
                        
                        if contains_exact_mc_ref:
                            cell_tuple = (sheet_name, cell.coordinate)
                            direct_dependents.add(cell_tuple)
                            direct_dependent_count += 1
                            logger.info(f"✅ [DIRECT_DEPENDENTS] EXACT MATCH: {sheet_name}!{cell.coordinate} = {formula}")
        
        workbook.close()
        
        logger.info(f"📊 [DIRECT_DEPENDENTS] Analysis complete:")
        logger.info(f"   Total formulas checked: {total_formulas_checked}")
        logger.info(f"   Direct dependents found: {direct_dependent_count}")
        logger.info(f"   MC variables: {len(monte_carlo_variables)}")
        
        return direct_dependents
        
    except Exception as e:
        logger.error(f"❌ [DIRECT_DEPENDENTS] Failed to analyze dependencies: {e}")
        return set() 