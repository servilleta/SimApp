"""
Range Analyzer for Excel Formulas

This module analyzes Excel formulas to extract all range references and determines
which empty cells need to be loaded to prevent missing cell issues during simulation.

Long-term solution to the missing cells problem:
1. Parse all formulas to find range references like C161:AL161
2. Expand ranges to individual cell coordinates
3. Ensure all cells within referenced ranges are loaded, even if empty
"""

import re
import logging
import openpyxl
from typing import Dict, List, Set, Tuple, Any, Optional
from dataclasses import dataclass
from pathlib import Path

logger = logging.getLogger(__name__)

@dataclass
class RangeInfo:
    """Information about a range reference in a formula"""
    sheet_name: str
    start_col: str
    start_row: int
    end_col: str
    end_row: int
    formula_cell: str
    formula: str
    
    def get_all_cells(self) -> Set[Tuple[str, str]]:
        """Get all cell coordinates in this range"""
        cells = set()
        start_col_num = self._column_to_number(self.start_col)
        end_col_num = self._column_to_number(self.end_col)
        
        for row in range(self.start_row, self.end_row + 1):
            for col_num in range(start_col_num, end_col_num + 1):
                col_str = self._number_to_column(col_num)
                cells.add((self.sheet_name, f"{col_str}{row}"))
        
        return cells
    
    def _column_to_number(self, col_str: str) -> int:
        """Convert column letters to number (A=1, B=2, etc.)"""
        col_str = col_str.replace('$', '').upper()
        result = 0
        for char in col_str:
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result
    
    def _number_to_column(self, col_num: int) -> str:
        """Convert column number to letters (1=A, 26=Z, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(ord('A') + col_num % 26) + result
            col_num //= 26
        return result

class FormulaRangeAnalyzer:
    """Analyzes Excel formulas to extract all range references"""
    
    # Enhanced regex for range detection
    RANGE_PATTERN = re.compile(
        r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?"  # Optional sheet name
        r"(\$?[A-Z]+)(\$?\d+)"                   # Start cell
        r":"                                     # Range separator
        r"(\$?[A-Z]+)(\$?\d+)",                 # End cell
        re.IGNORECASE
    )
    
    # Cell reference pattern for single cells
    CELL_PATTERN = re.compile(
        r"(?:(?:'([^']+)'|([A-Za-z0-9_]+))!)?"  # Optional sheet name
        r"(\$?[A-Z]+)(\$?\d+)",                 # Cell coordinate
        re.IGNORECASE
    )
    
    def __init__(self):
        self.range_cache: Dict[str, Set[RangeInfo]] = {}
        
    def analyze_formulas(self, formulas: Dict[str, Dict[str, str]]) -> Set[Tuple[str, str]]:
        """
        Analyze all formulas to extract range references and return all cells that need to be loaded
        
        Args:
            formulas: Dict of {sheet_name: {cell_coord: formula}}
            
        Returns:
            Set of (sheet_name, cell_coord) tuples for all cells that should be loaded
        """
        all_referenced_cells = set()
        ranges_found = []
        
        logger.info("🔍 [RANGE_ANALYZER] Starting comprehensive formula analysis")
        
        for sheet_name, sheet_formulas in formulas.items():
            logger.info(f"📊 [RANGE_ANALYZER] Analyzing {len(sheet_formulas)} formulas in sheet '{sheet_name}'")
            
            for cell_coord, formula in sheet_formulas.items():
                try:
                    # Extract ranges from this formula
                    ranges = self._extract_ranges_from_formula(formula, sheet_name, cell_coord)
                    ranges_found.extend(ranges)
                    
                    # Extract individual cell references
                    cells = self._extract_cells_from_formula(formula, sheet_name)
                    all_referenced_cells.update(cells)
                    
                    # Add all cells from ranges
                    for range_info in ranges:
                        range_cells = range_info.get_all_cells()
                        all_referenced_cells.update(range_cells)
                        
                except Exception as e:
                    logger.warning(f"🚨 [RANGE_ANALYZER] Error analyzing formula in {sheet_name}!{cell_coord}: {e}")
        
        logger.info(f"✅ [RANGE_ANALYZER] Analysis complete:")
        logger.info(f"   📈 Ranges found: {len(ranges_found)}")
        logger.info(f"   📋 Total cells to load: {len(all_referenced_cells)}")
        
        # Log sample ranges for debugging
        for i, range_info in enumerate(ranges_found[:5]):  # Show first 5 ranges
            logger.info(f"   🎯 Range {i+1}: {range_info.sheet_name}!{range_info.start_col}{range_info.start_row}:{range_info.end_col}{range_info.end_row}")
        
        return all_referenced_cells
    
    def _extract_ranges_from_formula(self, formula: str, current_sheet: str, cell_coord: str) -> List[RangeInfo]:
        """Extract all range references from a single formula"""
        if not formula or not formula.startswith('='):
            return []
        
        ranges = []
        
        for match in self.RANGE_PATTERN.finditer(formula):
            quoted_sheet, unquoted_sheet, start_col, start_row, end_col, end_row = match.groups()
            
            # Determine sheet name
            sheet_name = quoted_sheet or unquoted_sheet or current_sheet
            
            # Clean up column/row references (remove $)
            start_col_clean = start_col.replace('$', '')
            start_row_int = int(start_row.replace('$', ''))
            end_col_clean = end_col.replace('$', '')
            end_row_int = int(end_row.replace('$', ''))
            
            range_info = RangeInfo(
                sheet_name=sheet_name,
                start_col=start_col_clean,
                start_row=start_row_int,
                end_col=end_col_clean,
                end_row=end_row_int,
                formula_cell=cell_coord,
                formula=formula
            )
            ranges.append(range_info)
            
        return ranges
    
    def _extract_cells_from_formula(self, formula: str, current_sheet: str) -> Set[Tuple[str, str]]:
        """Extract individual cell references from a formula"""
        if not formula or not formula.startswith('='):
            return set()
        
        cells = set()
        
        # Find all cell references that are NOT part of ranges
        formula_without_ranges = self.RANGE_PATTERN.sub('', formula)
        
        for match in self.CELL_PATTERN.finditer(formula_without_ranges):
            quoted_sheet, unquoted_sheet, col, row = match.groups()
            
            # Determine sheet name
            sheet_name = quoted_sheet or unquoted_sheet or current_sheet
            
            # Clean up cell coordinate
            cell_coord = f"{col.replace('$', '')}{row.replace('$', '')}"
            
            cells.add((sheet_name, cell_coord))
            
        return cells
    
    def analyze_file_ranges(self, file_path: str) -> Dict[str, Any]:
        """
        Analyze an Excel file to extract all range information
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with analysis results
        """
        try:
            # Load workbook to get formulas
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            all_formulas = {}
            total_formulas = 0
            
            # Extract formulas from all sheets
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_formulas = {}
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f' and cell.value:
                            sheet_formulas[cell.coordinate] = cell.value
                            total_formulas += 1
                
                if sheet_formulas:
                    all_formulas[sheet_name] = sheet_formulas
            
            # Analyze the formulas
            all_referenced_cells = self.analyze_formulas(all_formulas)
            
            return {
                'total_formulas': total_formulas,
                'sheets_analyzed': len(all_formulas),
                'total_referenced_cells': len(all_referenced_cells),
                'referenced_cells': all_referenced_cells,
                'formulas': all_formulas
            }
            
        except Exception as e:
            logger.error(f"❌ [RANGE_ANALYZER] Failed to analyze file {file_path}: {e}")
            raise

def get_referenced_cells_for_file(file_id: str, upload_dir: str) -> Set[Tuple[str, str]]:
    """
    Get all cells that are referenced by formulas in an Excel file
    
    Args:
        file_id: File ID to analyze
        upload_dir: Directory containing uploaded files
        
    Returns:
        Set of (sheet_name, cell_coord) tuples for all referenced cells
    """
    # Find the Excel file
    excel_file_path = None
    try:
        import os
        for filename in os.listdir(upload_dir):
            if filename.startswith(file_id) and (filename.endswith('.xlsx') or filename.endswith('.xls')):
                excel_file_path = os.path.join(upload_dir, filename)
                break
        
        if not excel_file_path:
            logger.warning(f"⚠️ [RANGE_ANALYZER] Excel file not found for file_id: {file_id}")
            return set()
        
        # Analyze the file
        analyzer = FormulaRangeAnalyzer()
        analysis = analyzer.analyze_file_ranges(excel_file_path)
        
        logger.info(f"📊 [RANGE_ANALYZER] File {file_id} analysis:")
        logger.info(f"   Formulas: {analysis['total_formulas']}")
        logger.info(f"   Referenced cells: {analysis['total_referenced_cells']}")
        
        return analysis['referenced_cells']
        
    except Exception as e:
        logger.error(f"❌ [RANGE_ANALYZER] Error analyzing file {file_id}: {e}")
        return set() 