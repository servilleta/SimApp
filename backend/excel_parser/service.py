import pandas as pd
import openpyxl
from fastapi import UploadFile, HTTPException
import os
from uuid import uuid4
from io import BytesIO
from typing import List, Optional, Dict, Tuple, Any, Set
from pathlib import Path
import pyarrow as pa
import pyarrow.feather as feather
import json
import asyncio
from datetime import datetime, date
import logging
import time

from config import settings
from excel_parser.schemas import ExcelFileResponse, SheetData, CellData
from shared.persistent_excel_storage import get_persistent_excel_file_path

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# 🆕 PUBLIC UTILITY: upload_excel_file
# ---------------------------------------------------------------------------
async def upload_excel_file(file_id: str, filename: str, file_content: bytes) -> Dict[str, Any]:
    """Save an uploaded Excel file to disk.

    This helper is **primarily for unit-tests** and mirrors the behaviour
    of the original FastAPI upload endpoint in a minimal way so that the
    test suite can depend on it without spinning up the full HTTP stack.

    Args:
        file_id: The deterministic ID the caller wants to associate with the file.
        filename: Original filename (e.g. "example.xlsx").
        file_content: Raw bytes of the Excel file.

    Returns:
        A dict with basic metadata about the stored file (id, name, path, size).
    """
    # Ensure the upload directory exists
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)

    # Normalise filename to avoid directory traversal issues
    safe_name = os.path.basename(filename)
    stored_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_{safe_name}")

    # Write the bytes to disk (overwrite if already present for idempotency in tests)
    with open(stored_path, "wb") as fp:
        fp.write(file_content)

    file_size = len(file_content)

    logger.info(
        f"[ExcelParser] Saved uploaded file '{filename}' as '{stored_path}' (size={file_size} bytes)"
    )

    # Return minimal response compatible with tests. Full parsing is not required.
    return {
        "file_id": file_id,
        "filename": safe_name,
        "file_size": file_size,
        "path": stored_path,
    }

# Helper function to convert column number to Excel column letter
def _get_column_letter(col_num: int) -> str:
    """Convert column number to Excel column letter (1->A, 26->Z, 27->AA, etc.)"""
    result = ""
    while col_num > 0:
        col_num -= 1  # Adjust for 0-based indexing
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result or "A"

# Helper function to extract column widths from Excel worksheet
def _extract_column_widths(worksheet) -> Dict[str, float]:
    """Extract column widths from an openpyxl worksheet"""
    column_widths = {}
    
    try:
        logger.info(f"📏 [DEBUG] Starting column width extraction for sheet: {worksheet.title}")
        logger.info(f"📏 [DEBUG] Worksheet max_column: {worksheet.max_column}")
        logger.info(f"📏 [DEBUG] Column dimensions keys: {list(worksheet.column_dimensions.keys())}")
        
        # Get the column dimensions from the worksheet
        for col_letter, dimension in worksheet.column_dimensions.items():
            logger.info(f"📏 [DEBUG] Column {col_letter}: width={dimension.width}, auto_size={getattr(dimension, 'auto_size', None)}")
            if dimension.width is not None and dimension.width > 0:
                excel_width = dimension.width
                
                # Convert Excel width units to pixels using the correct Excel formula
                # Excel uses a non-linear conversion:
                # - For widths <= 1: pixels = width * 18
                # - For widths > 1: pixels = 18 + (width - 1) * 11
                if excel_width <= 1:
                    pixels = excel_width * 18
                else:
                    pixels = 18 + (excel_width - 1) * 11
                
                column_widths[col_letter] = pixels
                logger.info(f"📏 [DEBUG] ✅ Column {col_letter}: {excel_width} units → {pixels} pixels")
        
        # Fill in missing columns with default width (8.43 Excel units = ~112 pixels)
        if worksheet.max_column:
            for col_num in range(1, worksheet.max_column + 1):
                col_letter = _get_column_letter(col_num)
                if col_letter not in column_widths:
                    # Default Excel column width is 8.43 units
                    default_width = 8.43
                    default_pixels = 18 + (default_width - 1) * 11  # ≈ 112 pixels
                    column_widths[col_letter] = default_pixels
                    logger.info(f"📏 [DEBUG] ⚙️ Column {col_letter}: default width {default_width} units → {default_pixels} pixels")
        
        logger.info(f"📏 [DEBUG] ✅ Final column widths extracted: {len(column_widths)} columns")
        logger.info(f"📏 [DEBUG] Column width summary: {dict(list(column_widths.items())[:5])}...")
        
    except Exception as e:
        logger.error(f"📏 [ERROR] Failed to extract column widths: {e}")
        logger.error(f"📏 [ERROR] Falling back to default widths")
        
        # Fallback: provide default widths for all columns
        if worksheet.max_column:
            for col_num in range(1, worksheet.max_column + 1):
                col_letter = _get_column_letter(col_num)
                column_widths[col_letter] = 112  # Default Excel width in pixels
    
    return column_widths

# Helper function to extract cell formatting information
def _extract_cell_formatting(cell) -> Dict[str, Any]:
    """Extract formatting information from an openpyxl cell"""
    formatting = {}
    
    try:
        # Number format
        if hasattr(cell, 'number_format') and cell.number_format:
            formatting['number_format'] = cell.number_format
        
        # Data type
        if hasattr(cell, 'data_type') and cell.data_type:
            formatting['data_type'] = cell.data_type
            
        # Font information
        if hasattr(cell, 'font') and cell.font:
            font = cell.font
            if font.name:
                formatting['font_name'] = font.name
            if font.size:
                formatting['font_size'] = float(font.size)
            if font.bold is not None:
                formatting['font_bold'] = font.bold
            if font.italic is not None:
                formatting['font_italic'] = font.italic
            # Improved font color extraction with better error handling
            try:
                if hasattr(font, 'color') and font.color and hasattr(font.color, 'rgb'):
                    if font.color.rgb and isinstance(font.color.rgb, str):
                        formatting['font_color'] = f"#{font.color.rgb}"
            except Exception as font_color_error:
                # Silently skip font color if extraction fails
                pass
        
        # Fill/background color
        if hasattr(cell, 'fill') and cell.fill:
            fill = cell.fill
            try:
                if hasattr(fill, 'start_color') and fill.start_color and hasattr(fill.start_color, 'rgb'):
                    if fill.start_color.rgb and fill.start_color.rgb != '00000000':  # Skip default/transparent
                        if isinstance(fill.start_color.rgb, str):
                            formatting['fill_color'] = f"#{fill.start_color.rgb}"
            except Exception as fill_color_error:
                # Silently skip fill color if extraction fails
                pass
        
        # Alignment
        if hasattr(cell, 'alignment') and cell.alignment:
            alignment = cell.alignment
            if alignment.horizontal:
                formatting['alignment'] = alignment.horizontal
        
        # Border information
        if hasattr(cell, 'border') and cell.border:
            border = cell.border
            try:
                # Extract border styles and colors
                if border.top and border.top.style:
                    top_color = ""
                    if border.top.color and hasattr(border.top.color, 'rgb'):
                        if border.top.color.rgb and isinstance(border.top.color.rgb, str):
                            top_color = f"#{border.top.color.rgb}"
                    formatting['border_top'] = f"{border.top.style}:{top_color}" if top_color else border.top.style
                
                if border.bottom and border.bottom.style:
                    bottom_color = ""
                    if border.bottom.color and hasattr(border.bottom.color, 'rgb'):
                        if border.bottom.color.rgb and isinstance(border.bottom.color.rgb, str):
                            bottom_color = f"#{border.bottom.color.rgb}"
                    formatting['border_bottom'] = f"{border.bottom.style}:{bottom_color}" if bottom_color else border.bottom.style
                
                if border.left and border.left.style:
                    left_color = ""
                    if border.left.color and hasattr(border.left.color, 'rgb'):
                        if border.left.color.rgb and isinstance(border.left.color.rgb, str):
                            left_color = f"#{border.left.color.rgb}"
                    formatting['border_left'] = f"{border.left.style}:{left_color}" if left_color else border.left.style
                
                if border.right and border.right.style:
                    right_color = ""
                    if border.right.color and hasattr(border.right.color, 'rgb'):
                        if border.right.color.rgb and isinstance(border.right.color.rgb, str):
                            right_color = f"#{border.right.color.rgb}"
                    formatting['border_right'] = f"{border.right.style}:{right_color}" if right_color else border.right.style
            except Exception as border_error:
                # Silently skip border extraction if it fails
                pass
                
    except Exception as e:
        # If formatting extraction fails, continue without it
        print(f"⚠️ Warning: Could not extract formatting for cell: {e}")
    
    return formatting

# Helper function to format cell value for display
def _format_cell_display_value(cell_value, number_format: str = None, data_type: str = None) -> str:
    """Format cell value for display based on Excel formatting"""
    if cell_value is None:
        return ""
    
    try:
        # Handle dates
        if isinstance(cell_value, (datetime, date)):
            if number_format and any(fmt in number_format.lower() for fmt in ['mm', 'dd', 'yyyy', 'yy']):
                # Use Excel date format
                if 'mm/dd/yyyy' in number_format.lower():
                    return cell_value.strftime('%m/%d/%Y')
                elif 'dd/mm/yyyy' in number_format.lower():
                    return cell_value.strftime('%d/%m/%Y')
                elif 'yyyy-mm-dd' in number_format.lower():
                    return cell_value.strftime('%Y-%m-%d')
                else:
                    return cell_value.strftime('%m/%d/%Y')  # Default
            else:
                return cell_value.strftime('%m/%d/%Y')
        
        # Handle numbers with formatting
        if isinstance(cell_value, (int, float)) and number_format:
            # Currency formats
            if '$' in number_format or 'currency' in number_format.lower():
                return f"${cell_value:,.2f}"
            
            # Percentage formats
            if '%' in number_format:
                return f"{cell_value * 100:.2f}%"
            
            # Comma-separated numbers
            if '#,##0' in number_format or '0,000' in number_format:
                if '.' in number_format:
                    # Count decimal places
                    decimal_places = len(number_format.split('.')[-1].replace('0', ''))
                    return f"{cell_value:,.{decimal_places}f}"
                else:
                    return f"{cell_value:,.0f}"
            
            # Decimal places
            if '0.00' in number_format:
                decimal_places = number_format.count('0') - number_format.find('.') - 1
                return f"{cell_value:.{decimal_places}f}"
        
        # Default formatting for numbers
        if isinstance(cell_value, float):
            # Check if it's a whole number
            if cell_value.is_integer():
                return str(int(cell_value))
            else:
                return f"{cell_value:.2f}"
        elif isinstance(cell_value, int):
            return str(cell_value)
        else:
            return str(cell_value)
            
    except Exception as e:
        print(f"⚠️ Warning: Could not format display value: {e}")
        return str(cell_value) if cell_value is not None else ""

# Helper function to convert Arrow table back to SheetData models
def arrow_table_to_models(tbl: pa.Table) -> List[SheetData]:
    """Convert Arrow table back to SheetData models for backward compatibility"""
    df = tbl.to_pandas()
    sheets_dict = {}
    sheet_column_widths = {}  # Track column widths per sheet
    
    for _, row in df.iterrows():
        sheet_name = row['sheet']
        coord = row['coord']
        value = row['value']
        formula = row['formula'] if pd.notna(row['formula']) else None
        
        # Convert string 'None' back to actual None
        if value == 'None' or pd.isna(value):
            value = None
        elif value is not None:
            # Try to convert back to appropriate type
            try:
                # Check if it's a number
                if str(value).replace('.', '').replace('-', '').isdigit():
                    value = float(value) if '.' in str(value) else int(value)
            except (ValueError, TypeError):
                # Keep as string if conversion fails
                pass
        
        if sheet_name not in sheets_dict:
            sheets_dict[sheet_name] = {}
            sheet_column_widths[sheet_name] = {}
        
        # Parse coordinate to get row/col indices
        col_letters = ''.join(filter(str.isalpha, coord)).upper()
        row_num = int(''.join(filter(str.isdigit, coord)))
        
        # Extract column width information
        col_width = row.get('column_width', 60.0)
        if col_letters not in sheet_column_widths[sheet_name]:
            sheet_column_widths[sheet_name][col_letters] = col_width
        
        # Convert column letters to number (A=1, B=2, etc.)
        col_num = 0
        for char in col_letters:
            col_num = col_num * 26 + (ord(char) - ord('A') + 1)
        
        cell_data = CellData(
            value=value,
            formula=formula,
            is_formula_cell=formula is not None,
            coordinate=coord,
            # Add formatting information from Arrow table
            display_value=row.get('display_value'),
            number_format=row.get('number_format'),
            data_type=row.get('data_type'),
            font_name=row.get('font_name'),
            font_size=row.get('font_size'),
            font_bold=row.get('font_bold'),
            font_italic=row.get('font_italic'),
            font_color=row.get('font_color'),
            fill_color=row.get('fill_color'),
            alignment=row.get('alignment'),
            border_top=row.get('border_top'),
            border_bottom=row.get('border_bottom'),
            border_left=row.get('border_left'),
            border_right=row.get('border_right')
        )
        
        sheets_dict[sheet_name][(row_num-1, col_num-1)] = cell_data
    
    # Convert to grid structure
    result = []
    for sheet_name, cells in sheets_dict.items():
        if not cells:
            result.append(SheetData(sheet_name=sheet_name, grid_data=[], column_widths=sheet_column_widths.get(sheet_name, {})))
            continue
            
        max_row = max(pos[0] for pos in cells.keys()) + 1
        max_col = max(pos[1] for pos in cells.keys()) + 1
        
        grid_data = []
        for r in range(max_row):
            row_data = []
            for c in range(max_col):
                cell = cells.get((r, c))
                row_data.append(cell)
            grid_data.append(row_data)
        
        result.append(SheetData(sheet_name=sheet_name, grid_data=grid_data, column_widths=sheet_column_widths.get(sheet_name, {})))
    
    return result 

# NEW: Helper to reliably locate an uploaded Excel file by its file_id

def _find_excel_file(file_id: str) -> Optional[str]:
    """Return the absolute path to the Excel file that begins with the given file_id.

    The uploaded files are stored in settings.UPLOAD_DIR with the pattern
        "{file_id}_{original_filename}.xlsx" (or .xls).
    This helper scans the upload directory for files starting with the file_id
    and having a valid Excel extension, returning the first match found.
    """
    try:
        for fname in os.listdir(settings.UPLOAD_DIR):
            lower = fname.lower()
            if fname.startswith(file_id) and (lower.endswith('.xlsx') or lower.endswith('.xls')):
                return os.path.join(settings.UPLOAD_DIR, fname)
    except FileNotFoundError:
        pass  # Upload dir may not exist yet
    return None

def _find_latest_excel_file_by_name(filename: str) -> Optional[str]:
    """Find the most recent Excel file with the given filename pattern.
    
    This handles cases where the frontend sends an excel_file_name (like 'complex_file.xlsx')
    instead of the actual file_id. Returns the path to the most recently uploaded file
    that matches the filename pattern.
    """
    try:
        import re
        from pathlib import Path
        
        # Remove extension and normalize filename for matching
        base_name = Path(filename).stem.lower()
        logger.info(f"🔍 [FILE_MAPPING] Looking for files matching pattern: {base_name}")
        
        # Find all files that contain the base filename
        matching_files = []
        for fname in os.listdir(settings.UPLOAD_DIR):
            lower = fname.lower()
            if (lower.endswith('.xlsx') or lower.endswith('.xls')):
                # Extract original filename from pattern: {file_id}_{original_filename}.xlsx
                if '_' in fname:
                    original_name = '_'.join(fname.split('_')[1:]).lower()
                    original_base = Path(original_name).stem
                    
                    # Check if this matches our target filename
                    if base_name in original_base or original_base in base_name:
                        file_path = os.path.join(settings.UPLOAD_DIR, fname)
                        file_stat = os.stat(file_path)
                        matching_files.append((file_path, file_stat.st_mtime, fname))
                        logger.info(f"🔍 [FILE_MAPPING] Found matching file: {fname}")
        
        if matching_files:
            # Sort by modification time (most recent first)
            matching_files.sort(key=lambda x: x[1], reverse=True)
            latest_file_path = matching_files[0][0]
            latest_file_name = matching_files[0][2]
            logger.info(f"✅ [FILE_MAPPING] Using latest file: {latest_file_name} -> {latest_file_path}")
            return latest_file_path
        else:
            logger.warning(f"⚠️  [FILE_MAPPING] No files found matching pattern: {base_name}")
            
    except Exception as e:
        logger.error(f"❌ [FILE_MAPPING] Error finding file by name {filename}: {e}")
    
    return None

def resolve_file_path(file_identifier: str) -> Optional[str]:
    """Robust file resolution that handles both file_ids and filenames.
    
    This function tries multiple strategies to find the correct Excel file:
    1. Direct file_id lookup (if it looks like a UUID)
    2. Filename pattern matching (for cases like 'complex_file.xlsx')
    3. Latest file fallback (if only one Excel file exists)
    """
    logger.info(f"🔍 [FILE_RESOLUTION] Resolving file identifier: {file_identifier}")
    
    # Strategy 1: Try direct file_id lookup (if it looks like a UUID)
    if len(file_identifier) >= 32 and '-' in file_identifier:
        direct_path = _find_excel_file(file_identifier)
        if direct_path:
            logger.info(f"✅ [FILE_RESOLUTION] Found by file_id: {direct_path}")
            return direct_path
    
    # Strategy 2: Try filename pattern matching
    pattern_path = _find_latest_excel_file_by_name(file_identifier)
    if pattern_path:
        logger.info(f"✅ [FILE_RESOLUTION] Found by filename pattern: {pattern_path}")
        return pattern_path
    
    # Strategy 3: Check persistent storage for simulation files
    try:
        persistent_path = get_persistent_excel_file_path(file_identifier)
        if persistent_path and os.path.exists(persistent_path):
            logger.info(f"✅ [FILE_RESOLUTION] Found in persistent storage: {persistent_path}")
            return persistent_path
    except Exception as e:
        logger.error(f"❌ [FILE_RESOLUTION] Error checking persistent storage: {e}")
    
    # Strategy 4: Last resort - find the most recent Excel file
    try:
        all_excel_files = []
        for fname in os.listdir(settings.UPLOAD_DIR):
            lower = fname.lower()
            if lower.endswith('.xlsx') or lower.endswith('.xls'):
                file_path = os.path.join(settings.UPLOAD_DIR, fname)
                file_stat = os.stat(file_path)
                all_excel_files.append((file_path, file_stat.st_mtime, fname))
        
        if all_excel_files:
            all_excel_files.sort(key=lambda x: x[1], reverse=True)
            latest_path = all_excel_files[0][0]
            latest_name = all_excel_files[0][2]
            logger.warning(f"⚠️  [FILE_RESOLUTION] Using latest Excel file as fallback: {latest_name} -> {latest_path}")
            return latest_path
    except Exception as e:
        logger.error(f"❌ [FILE_RESOLUTION] Error in fallback resolution: {e}")
    
    logger.error(f"❌ [FILE_RESOLUTION] Could not resolve file identifier: {file_identifier}")
    return None

async def parse_excel_file(file: UploadFile) -> ExcelFileResponse:
    """Parse an Excel file and extract full grid data, formulas, and cell info for all sheets."""
    start_time = time.time()
    file_id = str(uuid4())
    # Ensure UPLOAD_DIR exists (moved from original guide's main.py for encapsulation)
    os.makedirs(settings.UPLOAD_DIR, exist_ok=True)
    file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_{file.filename}")

    content = await file.read()
    if len(content) > settings.MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail=f"File exceeds maximum size of {settings.MAX_UPLOAD_SIZE / (1024*1024)}MB")
    
    try:
        with open(file_path, "wb") as f:
            f.write(content)
    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path) # Clean up partial save
        raise HTTPException(status_code=500, detail=f"Could not save uploaded file: {str(e)}")

    all_sheets_data: List[SheetData] = []
    all_formulas: Dict[str, Dict[str,str]] = {} # Store formulas per sheet: {sheet_name: {coord: formula}}
    all_named_ranges: Dict[str, str] = {} # Store named ranges: {name: destination}
    all_metadata: Dict[str, Any] = {} # Store workbook metadata
    all_tables: Dict[str, Any] = {} # Store table data: {table_name: {col: range}}

    try:
        # 🚀 FAST PATH: Use streaming parsing (3x speed improvement)
        # Load workbook twice for compatibility with formulas and values
        # NOTE: We need read_only=False for column width access
        workbook_formulas = openpyxl.load_workbook(BytesIO(content), data_only=False, read_only=False)
        workbook_values = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)

        # 🚀 NEW: Extract Workbook Metadata
        props = workbook_formulas.properties
        all_metadata = {
            "creator": props.creator,
            "last_modified_by": props.lastModifiedBy,
            "created": props.created.isoformat() if props.created else None,
            "modified": props.modified.isoformat() if props.modified else None,
            "title": props.title,
            "subject": props.subject,
            "description": props.description,
            "version": props.version,
            "revision": props.revision,
        }
        logger.info(f"🔍 Found workbook metadata created by {props.creator}.")

        # 🚀 NEW: Extract Named Ranges
        if workbook_formulas.defined_names:
            for name in workbook_formulas.defined_names.definedName:
                # name.name is the named range, e.g., "Sales"
                # name.attr_text is the destination, e.g., "'Sheet1'!$A$1:$A$10"
                if name.name and name.attr_text:
                    all_named_ranges[name.name] = name.attr_text
        
        logger.info(f"🔍 Found {len(all_named_ranges)} named ranges.")

        # 🚀 NEW: Extract Table Data
        # Note: Tables are not available in read_only mode, so we skip this for now
        # This would require loading the workbook in non-read-only mode which is slower
        # For now, we'll just log that table extraction is skipped
        logger.info(f"🔍 Table extraction skipped (not available in read-only mode).")

        # 🚀 FAST PATH: Collect data for Arrow table
        arrow_rows = []
        
        for sheet_name in workbook_formulas.sheetnames:
            # Check timeout before processing each sheet
            if time.time() - start_time > settings.EXCEL_PARSE_TIMEOUT_SEC:
                raise HTTPException(
                    status_code=408, 
                    detail=f"Excel parsing timeout after {settings.EXCEL_PARSE_TIMEOUT_SEC} seconds. File too large to process."
                )
            
            sheet_formulas = workbook_formulas[sheet_name]
            sheet_values = workbook_values[sheet_name]
            sheet_formulas_dict: Dict[str, str] = {}
            
            # Extract column widths from the sheet
            column_widths = _extract_column_widths(sheet_formulas)
            
            # Use iter_rows for streaming (much faster than cell-by-cell access)
            rows_data = []
            
            # Get max dimensions
            max_row = sheet_formulas.max_row or 1
            max_col = sheet_formulas.max_column or 1
            
            logger.info(f"📊 Processing sheet '{sheet_name}' with {max_row} rows and {max_col} columns")
            
            # Stream through rows using iter_rows with progress reporting
            for row_idx, (formula_row, value_row) in enumerate(zip(
                sheet_formulas.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False),
                sheet_values.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False)
            )):
                # Check timeout periodically
                if row_idx % settings.EXCEL_PARSE_PROGRESS_INTERVAL == 0:
                    if time.time() - start_time > settings.EXCEL_PARSE_TIMEOUT_SEC:
                        raise HTTPException(
                            status_code=408, 
                            detail=f"Excel parsing timeout after {settings.EXCEL_PARSE_TIMEOUT_SEC} seconds. File too large to process."
                        )
                    logger.info(f"📊 Processing row {row_idx}/{max_row} in sheet '{sheet_name}'")
                
                row_data = []
                
                for col_idx, (formula_cell, value_cell) in enumerate(zip(formula_row, value_row)):
                    try:
                        # Safe access to cell properties with type conversion
                        cell_value = getattr(value_cell, 'value', None)
                        
                        # Convert cell value to string for consistent Arrow storage
                        if cell_value is not None:
                            if isinstance(cell_value, (int, float)):
                                # Keep numeric values as is for proper typing
                                pass
                            elif not isinstance(cell_value, str):
                                cell_value = str(cell_value)
                        
                        formula_str = None
                        is_formula = False
                        
                        # Safe coordinate handling with proper fallback
                        coordinate = getattr(formula_cell, 'coordinate', None)
                        if coordinate is None:
                            # Create coordinate manually using Excel column naming
                            col_letter = _get_column_letter(col_idx + 1)
                            coordinate = f"{col_letter}{row_idx+1}"
                        
                        # Check if it's a formula cell
                        if hasattr(formula_cell, 'data_type') and formula_cell.data_type == 'f':
                            is_formula = True
                            if hasattr(formula_cell, 'value') and formula_cell.value:
                                # Debug: Log all formula objects to identify potential issues
                                logger.info(f"🔍 Formula object in {coordinate}: type={type(formula_cell.value)}, value={repr(formula_cell.value)}")
                                # Handle different types of formula objects
                                if hasattr(formula_cell.value, 'text'):
                                    # ArrayFormula object - get the actual formula text
                                    formula_str = formula_cell.value.text
                                    logger.info(f"🔧 ArrayFormula fixed in {coordinate}: {formula_str}")
                                elif hasattr(formula_cell.value, 'r1') and hasattr(formula_cell.value, 'r2'):
                                    # DataTableFormula object - convert to string representation
                                    formula_str = str(formula_cell.value)
                                    logger.info(f"🔧 DataTableFormula found in {coordinate}: {formula_str}")
                                elif hasattr(formula_cell.value, '__class__') and 'Formula' in str(formula_cell.value.__class__):
                                    # Other formula objects - log and try to convert to string
                                    logger.info(f"🔧 Formula object type {type(formula_cell.value)} in {coordinate}")
                                    if hasattr(formula_cell.value, 'text'):
                                        formula_str = formula_cell.value.text
                                    else:
                                        formula_str = str(formula_cell.value)
                                    logger.info(f"🔧 Formula object converted to: {formula_str}")
                                else:
                                    # Regular formula - convert to string
                                    formula_str = str(formula_cell.value)
                                
                                if not formula_str.startswith('='):
                                    formula_str = "=" + formula_str
                                sheet_formulas_dict[coordinate] = formula_str
                        
                        # Extract formatting information from the formula cell (has formatting)
                        formatting = _extract_cell_formatting(formula_cell)
                        
                        # Create formatted display value
                        display_value = _format_cell_display_value(
                            cell_value, 
                            formatting.get('number_format'), 
                            formatting.get('data_type')
                        )
                        
                        # Skip completely empty cells
                        if cell_value is None and formula_str is None:
                            row_data.append(None)
                        else:
                            cell_data = CellData(
                                value=cell_value,
                                formula=formula_str,
                                is_formula_cell=is_formula,
                                coordinate=coordinate,
                                # Add formatting information
                                number_format=formatting.get('number_format'),
                                display_value=display_value,
                                data_type=formatting.get('data_type'),
                                font_name=formatting.get('font_name'),
                                font_size=formatting.get('font_size'),
                                font_bold=formatting.get('font_bold'),
                                font_italic=formatting.get('font_italic'),
                                font_color=formatting.get('font_color'),
                                fill_color=formatting.get('fill_color'),
                                alignment=formatting.get('alignment'),
                                border_top=formatting.get('border_top'),
                                border_bottom=formatting.get('border_bottom'),
                                border_left=formatting.get('border_left'),
                                border_right=formatting.get('border_right')
                            )
                            row_data.append(cell_data)
                            
                            # Add to Arrow data with safe type conversion (including formatting)
                            col_letter = _get_column_letter(col_idx + 1)
                            col_width = column_widths.get(col_letter, 60.0)  # Default width if not found
                            
                            arrow_rows.append({
                                'sheet': str(sheet_name) if sheet_name else '',
                                'coord': str(coordinate) if coordinate else '',
                                'value': cell_value,  # Keep original type for proper schema inference
                                'formula': str(formula_str) if formula_str else None,
                                'display_value': display_value,
                                'number_format': formatting.get('number_format'),
                                'data_type': formatting.get('data_type'),
                                'font_name': formatting.get('font_name'),
                                'font_size': formatting.get('font_size'),
                                'font_bold': formatting.get('font_bold'),
                                'font_italic': formatting.get('font_italic'),
                                'font_color': formatting.get('font_color'),
                                'fill_color': formatting.get('fill_color'),
                                'alignment': formatting.get('alignment'),
                                'border_top': formatting.get('border_top'),
                                'border_bottom': formatting.get('border_bottom'),
                                'border_left': formatting.get('border_left'),
                                'border_right': formatting.get('border_right'),
                                'column_width': col_width
                            })
                    
                    except Exception as cell_error:
                        # Handle problematic cells gracefully
                        safe_coord = f"R{row_idx+1}C{col_idx+1}"
                        print(f"⚠️ Cell parsing error at {safe_coord}: {cell_error}")
                        row_data.append(None)
                
                rows_data.append(row_data)
            
            all_sheets_data.append(SheetData(sheet_name=sheet_name, grid_data=rows_data, column_widths=column_widths))
            all_formulas[sheet_name] = sheet_formulas_dict
            logger.info(f"✅ Completed sheet '{sheet_name}' with {len(rows_data)} rows")

        # 🚀 CREATE ARROW CACHE for ultra-fast subsequent loads
        if arrow_rows:
            cache_dir = "/app/cache"
            os.makedirs(cache_dir, exist_ok=True)
            
            try:
                # Create Arrow table with explicit schema for better type handling
                import pyarrow as pa
                schema = pa.schema([
                    ('sheet', pa.string()),
                    ('coord', pa.string()), 
                    ('value', pa.string()),  # Use string type to avoid type conflicts
                    ('formula', pa.string()),
                    ('display_value', pa.string()),
                    ('number_format', pa.string()),
                    ('data_type', pa.string()),
                    ('font_name', pa.string()),
                    ('font_size', pa.float64()),
                    ('font_bold', pa.bool_()),
                    ('font_italic', pa.bool_()),
                    ('font_color', pa.string()),
                    ('fill_color', pa.string()),
                    ('alignment', pa.string()),
                    ('border_top', pa.string()),
                    ('border_bottom', pa.string()),
                    ('border_left', pa.string()),
                    ('border_right', pa.string()),
                    ('column_width', pa.float64())
                ])
                
                # Convert all values to appropriate types for consistent storage
                safe_arrow_rows = []
                for row in arrow_rows:
                    safe_row = {
                        'sheet': str(row['sheet']) if row['sheet'] else '',
                        'coord': str(row['coord']) if row['coord'] else '',
                        'value': str(row['value']) if row['value'] is not None else None,
                        'formula': str(row['formula']) if row['formula'] else None,
                        'display_value': str(row['display_value']) if row['display_value'] else None,
                        'number_format': str(row['number_format']) if row['number_format'] else None,
                        'data_type': str(row['data_type']) if row['data_type'] else None,
                        'font_name': str(row['font_name']) if row['font_name'] else None,
                        'font_size': float(row['font_size']) if row['font_size'] is not None else None,
                        'font_bold': bool(row['font_bold']) if row['font_bold'] is not None else None,
                        'font_italic': bool(row['font_italic']) if row['font_italic'] is not None else None,
                        'font_color': str(row['font_color']) if row['font_color'] else None,
                        'fill_color': str(row['fill_color']) if row['fill_color'] else None,
                        'alignment': str(row['alignment']) if row['alignment'] else None,
                        'border_top': str(row['border_top']) if row['border_top'] else None,
                        'border_bottom': str(row['border_bottom']) if row['border_bottom'] else None,
                        'border_left': str(row['border_left']) if row['border_left'] else None,
                        'border_right': str(row['border_right']) if row['border_right'] else None,
                        'column_width': float(row['column_width']) if row['column_width'] is not None else 60.0
                    }
                    safe_arrow_rows.append(safe_row)
                
                arrow_table = pa.Table.from_pylist(safe_arrow_rows, schema=schema)
                feather_path = os.path.join(cache_dir, f"{file_id}.feather")
                
                feather.write_feather(arrow_table, feather_path, compression='lz4')
                logger.info(f"✅ Arrow cache created: {feather_path} ({len(arrow_rows)} cells)")
            except Exception as cache_error:
                logger.warning(f"⚠️ Arrow cache creation failed: {cache_error}")
                # Continue without cache - not critical

        # After parsing formulas, add logging
        logger.warning(f"[EXCEL_PARSER_DEBUG] Total formulas found: {len(all_formulas)}")
        vlookup_count = 0
        for sheet_name, sheet_formulas in all_formulas.items():
            for cell, formula in sheet_formulas.items():
                if 'VLOOKUP' in str(formula).upper():
                    vlookup_count += 1
                    logger.warning(f"[EXCEL_PARSER_DEBUG] VLOOKUP found in {sheet_name}!{cell}: {formula}")
        logger.warning(f"[EXCEL_PARSER_DEBUG] Total VLOOKUP formulas: {vlookup_count}")

    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path) # Clean up
        raise HTTPException(status_code=400, detail=f"Failed to parse Excel file with openpyxl: {str(e)}")
    
    # Store all extracted formulas (per sheet) for later retrieval by simulation service
    formulas_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_formulas.json")
    import json
    try:
        with open(formulas_file_path, "w") as ff:
            json.dump(all_formulas, ff) # Save all_formulas dict
    except Exception as e:
        # Clean up if saving formulas fails
        if os.path.exists(file_path): os.remove(file_path)
        if os.path.exists(formulas_file_path): os.remove(formulas_file_path)
        raise HTTPException(status_code=500, detail=f"Could not save parsed formulas: {str(e)}")

    # Store all named ranges
    named_ranges_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_named_ranges.json")
    try:
        with open(named_ranges_file_path, "w") as nrf:
            json.dump(all_named_ranges, nrf)
    except Exception as e:
        # Clean up if saving named ranges fails
        if os.path.exists(file_path): os.remove(file_path)
        if os.path.exists(formulas_file_path): os.remove(formulas_file_path)
        if os.path.exists(named_ranges_file_path): os.remove(named_ranges_file_path)
        raise HTTPException(status_code=500, detail=f"Could not save named ranges: {str(e)}")

    # Store all metadata
    metadata_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_metadata.json")
    try:
        with open(metadata_file_path, "w") as mf:
            json.dump(all_metadata, mf)
    except Exception as e:
        # Clean up if saving metadata fails
        if os.path.exists(file_path): os.remove(file_path)
        if os.path.exists(formulas_file_path): os.remove(formulas_file_path)
        if os.path.exists(named_ranges_file_path): os.remove(named_ranges_file_path)
        if os.path.exists(metadata_file_path): os.remove(metadata_file_path)
        raise HTTPException(status_code=500, detail=f"Could not save workbook metadata: {str(e)}")

    # Store all table data
    tables_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_tables.json")
    try:
        with open(tables_file_path, "w") as tf:
            json.dump(all_tables, tf)
    except Exception as e:
        # Clean up if saving tables fails
        if os.path.exists(file_path): os.remove(file_path)
        if os.path.exists(formulas_file_path): os.remove(formulas_file_path)
        if os.path.exists(named_ranges_file_path): os.remove(named_ranges_file_path)
        if os.path.exists(metadata_file_path): os.remove(metadata_file_path)
        if os.path.exists(tables_file_path): os.remove(tables_file_path)
        raise HTTPException(status_code=500, detail=f"Could not save table data: {str(e)}")

    # Store all parsed sheet data (including CellData objects)
    # This will be a larger file, ensure it's handled efficiently if sheets are huge.
    # Pydantic models need to be converted to dicts for JSON serialization.
    all_sheets_data_serializable = [sheet.model_dump() for sheet in all_sheets_data]
    sheet_data_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_sheet_data.json")
    try:
        with open(sheet_data_file_path, "w") as sdf:
            json.dump(all_sheets_data_serializable, sdf)
    except Exception as e:
        # Clean up if saving sheet data fails
        if os.path.exists(file_path): os.remove(file_path)
        if os.path.exists(formulas_file_path): os.remove(formulas_file_path)
        if os.path.exists(named_ranges_file_path): os.remove(named_ranges_file_path)
        if os.path.exists(sheet_data_file_path): os.remove(sheet_data_file_path)
        if os.path.exists(metadata_file_path): os.remove(metadata_file_path)
        if os.path.exists(tables_file_path): os.remove(tables_file_path)
        raise HTTPException(status_code=500, detail=f"Could not save parsed sheet data: {str(e)}")
    
    # 🚀 OPTIMIZATION: For very large files, limit the grid data size
    total_cells = 0
    for sheet in all_sheets_data:
        if sheet.grid_data and len(sheet.grid_data) > 0:
            max_cols = max(len(row) for row in sheet.grid_data) if sheet.grid_data else 0
            sheet_cells = len(sheet.grid_data) * max_cols
            total_cells += sheet_cells
    logger.info(f"📊 Total cells in response: {total_cells}")
    
    # 🚀 PERFORMANCE OPTIMIZATION: Use sparse format for large files  
    # This maintains full functionality while dramatically reducing response size
    if total_cells > 1000000:  # More than 1M cells - use sparse format (temporarily disabled for debugging)
        logger.info(f"🚀 Large file detected ({total_cells} cells), converting to sparse format for performance")
        optimized_sheets = []
        for sheet in all_sheets_data:
            # Extract all non-empty cells from grid
            sparse_cells = []
            max_row = len(sheet.grid_data)
            max_col = max(len(row) for row in sheet.grid_data) if sheet.grid_data else 0
            
            for row_idx, row in enumerate(sheet.grid_data):
                for col_idx, cell in enumerate(row):
                    if cell is not None:  # Only include non-empty cells
                        sparse_cells.append(cell)
            
            # Create sparse format sheet - maintains all data but in efficient format
            sparse_sheet = SheetData(
                sheet_name=sheet.sheet_name,
                grid_data=[],  # Empty grid for large files
                column_widths=sheet.column_widths,
                sparse_cells=sparse_cells,  # All non-empty cells with coordinates
                is_sparse_format=True,
                total_rows=max_row,
                total_cols=max_col
            )
            optimized_sheets.append(sparse_sheet)
            logger.info(f"📊 Sheet '{sheet.sheet_name}': {len(sparse_cells)} non-empty cells from {max_row}x{max_col} grid")
        
        all_sheets_data = optimized_sheets
        total_sparse_cells = sum(len(sheet.sparse_cells) for sheet in all_sheets_data if sheet.sparse_cells)
        logger.info(f"✅ Sparse optimization: {total_sparse_cells} cells vs {total_cells} full grid - {((total_cells-total_sparse_cells)/total_cells*100):.1f}% reduction")
    else:
        logger.info(f"📊 Standard response: {total_cells} cells, using full grid format")
    
    # 🧪 DEBUG: Log response creation details
    logger.info(f"🔍 Creating ExcelFileResponse with {len(all_sheets_data)} sheets")
    for i, sheet in enumerate(all_sheets_data):
        logger.info(f"  Sheet {i}: {sheet.sheet_name}, grid_data rows: {len(sheet.grid_data)}, sparse_cells: {len(sheet.sparse_cells) if sheet.sparse_cells else 0}")
    
    try:
        response = ExcelFileResponse(
            file_id=file_id,
            filename=file.filename,
            file_size=len(content),  # Include file size in bytes
            sheet_names=[sheet.sheet_name for sheet in all_sheets_data],  # Include sheet names
            sheets=all_sheets_data # This now contains the full grid_data for each sheet
        )
        logger.info(f"✅ ExcelFileResponse created successfully")
        return response
    except Exception as response_error:
        logger.error(f"❌ Failed to create ExcelFileResponse: {response_error}")
        # Log detailed information about the issue
        logger.error(f"❌ file_id: {file_id}")
        logger.error(f"❌ filename: {file.filename}")
        logger.error(f"❌ file_size: {len(content)}")
        logger.error(f"❌ sheet_names: {[sheet.sheet_name for sheet in all_sheets_data]}")
        logger.error(f"❌ sheets count: {len(all_sheets_data)}")
        raise HTTPException(status_code=500, detail=f"Failed to create response: {str(response_error)}")

# Function to get all formulas for a given file_id (now potentially per sheet)
async def get_formulas_for_file(file_identifier: str) -> Dict[str, Dict[str, str]]: # Returns Dict[sheet_name, Dict[coord, formula]]
    # 🚀 ROBUST FILE RESOLUTION: Handle both file_ids and filenames
    file_path = resolve_file_path(file_identifier)
    if not file_path:
        raise HTTPException(status_code=404, detail=f"Excel file for identifier '{file_identifier}' not found.")
    
    # Extract file_id from resolved path for formula lookup
    basename = os.path.basename(file_path)
    
    # Handle B2B API file naming pattern: b2b_xxxxx_filename.xlsx
    if basename.startswith('b2b_'):
        # For B2B files, extract file_id as everything before the original filename
        # Example: b2b_5378ada35b84472d_sim3.xlsx -> b2b_5378ada35b84472d
        parts = basename.split('_')
        if len(parts) >= 3:  # b2b, unique_id, filename
            file_id = '_'.join(parts[:2])  # b2b_unique_id
        else:
            file_id = parts[0]  # fallback to just b2b
    else:
        # For regular files, use the first part before underscore
        file_id = basename.split('_')[0]
    
    logger.info(f"🔍 [FORMULAS] Resolved '{file_identifier}' -> file_id: {file_id}")
    
    formulas_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_formulas.json")
    if not os.path.exists(formulas_file_path):
        raise HTTPException(status_code=404, detail=f"Formulas for file_id {file_id} not found.")
    import json # Already imported earlier, but good for clarity if functions are moved
    try:
        with open(formulas_file_path, "r") as ff:
            formulas = json.load(ff)
        return formulas
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not load formulas: {str(e)}")

async def get_named_ranges_for_file(file_id: str) -> Dict[str, str]:
    """Retrieves the dictionary of named ranges for a file."""
    named_ranges_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_named_ranges.json")
    if not os.path.exists(named_ranges_file_path):
        # Return empty dict if no named ranges file exists
        return {}
    
    try:
        with open(named_ranges_file_path, "r") as nrf:
            named_ranges = json.load(nrf)
        return named_ranges
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not load named ranges: {str(e)}")

async def get_metadata_for_file(file_id: str) -> Dict[str, Any]:
    """Retrieves the dictionary of workbook metadata for a file."""
    metadata_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_metadata.json")
    if not os.path.exists(metadata_file_path):
        return {}
    
    try:
        with open(metadata_file_path, "r") as mf:
            metadata = json.load(mf)
        return metadata
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not load workbook metadata: {str(e)}")

async def get_tables_for_file(file_id: str) -> Dict[str, Any]:
    """Retrieves the dictionary of tables for a file."""
    tables_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_tables.json")
    if not os.path.exists(tables_file_path):
        return {}
    
    try:
        with open(tables_file_path, "r") as tf:
            tables = json.load(tf)
        return tables
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not load table data: {str(e)}")

async def get_all_parsed_sheets_data(file_identifier: str) -> List[SheetData]:
    """Retrieves the full parsed sheet data (including values, formulas, coordinates) for a file."""
    
    # 🚀 ROBUST FILE RESOLUTION: Handle both file_ids and filenames
    file_path = resolve_file_path(file_identifier)
    if not file_path:
        raise HTTPException(status_code=404, detail=f"Excel file for identifier '{file_identifier}' not found.")
    
    # Extract file_id from resolved path
    basename = os.path.basename(file_path)
    
    # Handle B2B API file naming pattern: b2b_xxxxx_filename.xlsx
    if basename.startswith('b2b_'):
        # For B2B files, extract file_id as everything before the original filename
        # Example: b2b_5378ada35b84472d_sim3.xlsx -> b2b_5378ada35b84472d
        parts = basename.split('_')
        if len(parts) >= 3:  # b2b, unique_id, filename
            file_id = '_'.join(parts[:2])  # b2b_unique_id
        else:
            file_id = parts[0]  # fallback to just b2b
    else:
        # For regular files, use the first part before underscore
        file_id = basename.split('_')[0]
    
    logger.info(f"🔍 [SHEET_DATA] Resolved '{file_identifier}' -> file_id: {file_id}")
    
    # 🚀 FAST PATH: Try Arrow cache first (milliseconds vs seconds)
    arrow_path = Path(f"/app/cache/{file_id}.feather")
    if arrow_path.exists():
        try:
            # Memory-mapped load for ultra speed
            tbl = feather.read_table(str(arrow_path), memory_map=True)
            print(f"⚡ Loaded from Arrow cache: {arrow_path} ({len(tbl)} rows)")
            return arrow_table_to_models(tbl)
        except Exception as e:
            print(f"⚠️ Arrow cache load failed: {e}, falling back to JSON")
    
    # FALLBACK: Legacy JSON path
    sheet_data_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_sheet_data.json")
    if not os.path.exists(sheet_data_file_path):
        # 🔄 PERSISTENT STORAGE FALLBACK: Regenerate sheet data from Excel file
        logger.info(f"🔄 [PERSISTENT_FALLBACK] Sheet data JSON not found, regenerating from Excel file for: {file_id}")
        try:
            # Parse the Excel file directly to regenerate sheet data
            excel_file_path = file_path
            if not excel_file_path or not os.path.exists(excel_file_path):
                raise HTTPException(status_code=404, detail=f"Excel file for file_id {file_id} not found.")
            
            # Use pandas to read Excel file and regenerate sheet data
            df_dict = pd.read_excel(excel_file_path, sheet_name=None)  # Read all sheets
            sheets_data = []
            
            for sheet_name, df in df_dict.items():
                # Convert DataFrame to SheetData model using grid format
                max_rows = len(df)
                max_cols = len(df.columns) if max_rows > 0 else 0
                
                grid_data = []
                for row_idx in range(max_rows):
                    row_data = []
                    for col_idx in range(max_cols):
                        value = df.iloc[row_idx, col_idx] if row_idx < len(df) and col_idx < len(df.columns) else None
                        if pd.notna(value):
                            cell_address = f"{_get_column_letter(col_idx + 1)}{row_idx + 2}"  # +2 because Excel is 1-indexed and we assume row 1 is headers
                            cell_data = CellData(
                                coordinate=cell_address,
                                value=str(value),
                                data_type="string"  # Simplified for fallback
                            )
                            row_data.append(cell_data)
                        else:
                            row_data.append(None)
                    grid_data.append(row_data)
                
                sheet_data = SheetData(
                    sheet_name=sheet_name,
                    grid_data=grid_data
                )
                sheets_data.append(sheet_data)
            
            logger.info(f"✅ [PERSISTENT_FALLBACK] Successfully regenerated sheet data for {file_id} with {len(sheets_data)} sheets")
            return sheets_data
            
        except Exception as e:
            logger.error(f"❌ [PERSISTENT_FALLBACK] Failed to regenerate sheet data for {file_id}: {e}")
            raise HTTPException(status_code=500, detail=f"Could not regenerate sheet data for file_id {file_id}: {str(e)}")
    
    try:
        with open(sheet_data_file_path, "r") as sdf:
            all_sheets_data_raw = json.load(sdf)
        # Convert raw dicts back to Pydantic models
        return [SheetData(**sheet_data) for sheet_data in all_sheets_data_raw]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not load or parse sheet data: {str(e)}")

# Function to get a specific formula for a given file_id, sheet_name, and cell coordinate
async def get_formula_from_file(file_id: str, sheet_name: str, cell_coordinate: str) -> str:
    all_sheet_formulas = await get_formulas_for_file(file_id)
    sheet_formulas = all_sheet_formulas.get(sheet_name)
    if sheet_formulas is None:
        raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found in formulas for file {file_id}.")
    
    formula = sheet_formulas.get(cell_coordinate.upper())
    if formula is None:
        raise HTTPException(status_code=404, detail=f"Formula for cell {cell_coordinate} on sheet '{sheet_name}' not found in file {file_id}.")
    return formula

# get_file_variables and get_parsed_file_info might need adjustment if they relied on the simpler structure
# For now, they might not be directly compatible with the new multi-sheet formula storage.

async def get_file_variables(file_id: str, sheet_name: Optional[str] = None) -> list:
    """Get available variables (e.g., column names or cell coordinates) from a parsed Excel file's sheet."""
    # This function needs to be more sophisticated. For now, let's try to return column headers
    # of the first sheet if no sheet_name is provided, or a specified sheet.
    # A more robust version would inspect `all_sheets_data` if that was stored, or re-open the excel.
    
    # Find the original Excel file path
    found_file_path = None
    for f_name in os.listdir(settings.UPLOAD_DIR):
        if f_name.startswith(file_id) and (f_name.endswith('.xlsx') or f_name.endswith('.xls')):
            if "_formulas.json" not in f_name: # ensure it's not the formulas file
                found_file_path = os.path.join(settings.UPLOAD_DIR, f_name)
                break
    if not found_file_path:
        raise HTTPException(status_code=404, detail=f"Original Excel file for ID {file_id} not found.")

    try:
        # If no sheet_name provided, use the first sheet from the workbook.
        # Otherwise, use the specified sheet_name.
        df_cols = pd.read_excel(found_file_path, sheet_name=sheet_name if sheet_name else 0, nrows=0).columns.tolist()
        return df_cols
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not read columns for file {file_id} (sheet: {sheet_name or 'first'}). Error: {str(e)}")

async def get_parsed_file_info(file_id: str):
    """Get basic metadata about a previously uploaded and parsed Excel file."""
    formulas_file_path = os.path.join(settings.UPLOAD_DIR, f"{file_id}_formulas.json")
    original_file_name = "Unknown"
    sheet_names_from_formulas = []

    if os.path.exists(formulas_file_path):
        import json
        try:
            with open(formulas_file_path, "r") as ff:
                all_formulas = json.load(ff)
                sheet_names_from_formulas = list(all_formulas.keys())
        except Exception:
            pass # Could not load or parse formulas json

    # Find the original file name
    for f_name in os.listdir(settings.UPLOAD_DIR):
        if f_name.startswith(file_id) and (f_name.endswith('.xlsx') or f_name.endswith('.xls')):
             if "_formulas.json" not in f_name:
                original_file_name = f_name.replace(f"{file_id}_", "", 1)
                break

    if not os.path.exists(formulas_file_path) and original_file_name == "Unknown":
        raise HTTPException(status_code=404, detail=f"Parsed information for file ID {file_id} not found.")
    
    return {
        "file_id": file_id, 
        "filename": original_file_name, 
        "status": "parsed", 
        "message": "Basic info. Full grid data is returned by /upload.",
        "available_sheets": sheet_names_from_formulas # From the keys of the stored formulas
    } 

# Function to fetch specific cell values efficiently (Arrow cache first, fallback workbook)
async def get_cell_values(file_id: str, cells: Set[Tuple[str, str]]) -> Dict[Tuple[str, str], Any]:
    """Get values for specific cells from the Excel file."""
    # Locate the actual Excel file on disk
    file_path = _find_excel_file(file_id)

    if not file_path or not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail=f"File {file_id} not found")
    
    cell_values = {}
    
    try:
        # Load workbook with data_only=True to get calculated values
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        for sheet_name, cell_coord in cells:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                try:
                    cell = sheet[cell_coord]
                    if cell.value is not None:
                        cell_values[(sheet_name, cell_coord)] = cell.value
                except Exception as e:
                    logger.warning(f"Could not read cell {sheet_name}!{cell_coord}: {e}")
        
        workbook.close()
        
    except Exception as e:
        logger.error(f"Error reading cell values from {file_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error reading cell values: {str(e)}")
    
    return cell_values

async def get_constants_for_file(file_identifier: str, exclude_cells: Set[Tuple[str, str]] = None, target_sheet: str = None) -> Dict[Tuple[str, str], Any]:
    """
    Get ALL constant values AND formula-referenced empty cells from the Excel file.
    
    PRIORITY 1 FIX: Exclude Row 107 Customer Growth cells that should be formulas
    These must be loaded as formulas to maintain F4→Growth dependency chain.
    
    LONG-TERM SOLUTION: This function now analyzes formulas to find all referenced ranges
    and ensures empty cells within those ranges are loaded to prevent missing cell issues.
    
    This ensures we have all data cells available for formula evaluation.
    """
    # 🚀 ROBUST FILE RESOLUTION: Handle both file_ids and filenames
    file_path = resolve_file_path(file_identifier)
    if not file_path:
        raise HTTPException(status_code=404, detail=f"Excel file for identifier '{file_identifier}' not found.")
    
    # Extract file_id from resolved path for data lookup
    file_id = os.path.basename(file_path).split('_')[0]
    logger.info(f"🔍 [CONSTANTS] Resolved '{file_identifier}' -> file_id: {file_id}")
    
    exclude_cells = exclude_cells or set()
    
    # CRITICAL PRIORITY 1 FIX: Also exclude Row 107 Customer Growth cells
    # These must be loaded as formulas to maintain F4→Growth dependency chain
    customer_growth_cells = set()
    
    # Helper function to convert column number to Excel column letter
    def col_num_to_letter(n):
        """Convert column number to Excel column letter (1->A, 2->B, ..., 27->AA, etc.)"""
        result = ""
        while n > 0:
            n -= 1
            result = chr(n % 26 + ord('A')) + result
            n //= 26
        return result
    
    # Generate columns C through AL (columns 3 through 38)
    for col_num in range(3, 39):  # C=3, D=4, ..., AL=38
        col_letter = col_num_to_letter(col_num)
        for sheet in ['WIZEMICE Likest', 'WIZEMICE High', 'WIZEMICE Low']:
            customer_growth_cells.add((sheet, f"{col_letter}107"))
    
    logger.info(f"🔧 [PRIORITY_1_FIX] Excluding {len(customer_growth_cells)} Row 107 cells from constants to preserve dependency chain")
    exclude_cells.update(customer_growth_cells)
    
    cell_values = {}
    
    try:
        # STEP 1: Analyze formulas to find all referenced cells (including empty ones)
        from excel_parser.range_analyzer import get_referenced_cells_for_file
        referenced_cells = get_referenced_cells_for_file(file_id, settings.UPLOAD_DIR)
        logger.info(f"🔍 [CONSTANTS] Found {len(referenced_cells)} formula-referenced cells")
        
        # STEP 2: Load workbook with data_only=True to get calculated values
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # STEP 3: Collect all cells with values (with optional sheet filtering)
        sheets_to_process = [target_sheet] if target_sheet else workbook.sheetnames
        
        for sheet_name in sheets_to_process:
            # Skip if target_sheet specified but this sheet doesn't exist
            if target_sheet and sheet_name not in workbook.sheetnames:
                logger.warning(f"⚠️ [CONSTANTS] Target sheet '{target_sheet}' not found in workbook")
                continue
                
            sheet = workbook[sheet_name]
            
            # Iterate through all cells with values
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell_coord = cell.coordinate
                        cell_key = (sheet_name, cell_coord)
                        
                        # Skip excluded cells (like MC inputs)
                        if cell_key in exclude_cells:
                            continue
                        
                        # Include ALL cells with values (including formula results)
                        # The data_only=True workbook gives us calculated values for formulas
                        cell_values[cell_key] = cell.value
        
        # STEP 4: CRITICAL FIX - Include referenced cells from ANY sheet with their ACTUAL values
        # If a referenced cell has a real value in the workbook, use it; if it's truly empty, use 0.0
        empty_cells_added = 0
        for sheet_name, cell_coord in referenced_cells:
            cell_key = (sheet_name, cell_coord)
            
            # Skip if already loaded or excluded
            if cell_key in cell_values or cell_key in exclude_cells:
                continue
            
            # Check if this cell exists in the Excel structure
            if sheet_name in workbook.sheetnames:
                try:
                    sheet = workbook[sheet_name]
                    cell = sheet[cell_coord]
                    
                    # Use actual value when present; fallback to 0.0 for truly empty cells
                    if cell.value is not None:
                        cell_values[cell_key] = cell.value
                    else:
                        cell_values[cell_key] = 0.0
                        empty_cells_added += 1
                    
                except Exception as cell_err:
                    # Cell coordinate might be invalid, skip it
                    logger.debug(f"Could not access cell {sheet_name}!{cell_coord}: {cell_err}")
        
        workbook.close()
        
        logger.info(f"✅ [CONSTANTS] Loaded constants for {file_id}:")
        logger.info(f"   📋 Sheets processed: {sheets_to_process}")
        logger.info(f"   📊 Cells with values: {len(cell_values) - empty_cells_added}")
        logger.info(f"   🔧 Empty formula-referenced cells: {empty_cells_added}")
        logger.info(f"   📋 Total cells loaded: {len(cell_values)}")
        
        # Save constants to JSON for debugging
        try:
            import json
            constants_file = f"uploads/{file_id}_constants.json"
            with open(constants_file, 'w') as f:
                # Convert tuple keys to strings for JSON serialization
                json_data = {f"{sheet}!{coord}": value for (sheet, coord), value in cell_values.items()}
                json.dump(json_data, f, indent=2)
            logger.info(f"💾 Saved constants to {constants_file}")
        except Exception as save_err:
            logger.warning(f"Could not save constants file: {save_err}")
        
    except Exception as e:
        logger.error(f"Error reading constants from {file_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Error reading constants: {str(e)}")
    
    return cell_values 