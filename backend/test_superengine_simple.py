#!/usr/bin/env python3
"""
Simple SuperEngine Test
=======================
This script tests the SuperEngine integration with the current system.
"""

import asyncio
import json
import logging
import sys
import os
from typing import Dict, List, Any

# Add the backend directory to Python path
sys.path.insert(0, '/app' if os.path.exists('/app') else os.path.dirname(os.path.abspath(__file__)))

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

async def test_excel_upload():
    """Test Excel file upload through the API"""
    logger.info("🧪 Testing Excel Upload...")
    
    try:
        import httpx
        
        # Get auth token
        async with httpx.AsyncClient() as client:
            # Login
            login_response = await client.post(
                "http://localhost:8000/api/auth/token",
                data={"username": "testuser", "password": "testpass123"}
            )
            
            if login_response.status_code != 200:
                logger.error(f"❌ Login failed: {login_response.text}")
                return None
                
            token = login_response.json()["access_token"]
            logger.info("✅ Login successful")
            
            # Create test Excel file
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'SuperEngine Test'
            
            # Add test data
            ws['A1'] = 'Input A'
            ws['B1'] = 100
            ws['A2'] = 'Input B' 
            ws['B2'] = 200
            ws['A3'] = 'Input C'
            ws['B3'] = 300
            ws['A4'] = 'Result 1'
            ws['B4'] = '=B1+B2+B3'  # Sum formula
            ws['A5'] = 'Result 2'
            ws['B5'] = '=B4*2'      # Multiply by 2
            ws['A6'] = 'Result 3'
            ws['B6'] = '=IF(B5>1000,B5*1.1,B5*0.9)'  # Conditional formula
            
            # Save to temp file
            temp_file = "/tmp/superengine_test.xlsx"
            wb.save(temp_file)
            logger.info("✅ Test Excel file created")
            
            # Upload file
            with open(temp_file, 'rb') as f:
                files = {'file': ('superengine_test.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
                headers = {'Authorization': f'Bearer {token}'}
                
                upload_response = await client.post(
                    "http://localhost:8000/api/excel-parser/upload",
                    files=files,
                    headers=headers
                )
            
            if upload_response.status_code != 200:
                logger.error(f"❌ Upload failed: {upload_response.text}")
                return None
                
            upload_data = upload_response.json()
            logger.info(f"✅ Excel uploaded successfully: {upload_data['file_id']}")
            return upload_data['file_id'], token
            
    except Exception as e:
        logger.error(f"❌ Excel upload test failed: {e}")
        return None

async def test_simulation_setup(file_id: str, token: str):
    """Test simulation setup and execution"""
    logger.info("🧪 Testing Simulation Setup...")
    
    try:
        import httpx
        import uuid
        
        async with httpx.AsyncClient() as client:
            headers = {'Authorization': f'Bearer {token}'}
            
            # Create simulation request
            simulation_request = {
                "simulation_id": str(uuid.uuid4()),
                "file_id": file_id,
                "original_filename": "superengine_test.xlsx",
                "iterations": 1000,
                "engine_type": "enhanced",  # Using enhanced engine (which is now SuperEngine)
                "variables": [
                    {
                        "name": "B1",
                        "sheet_name": "SuperEngine Test",
                        "min_value": 80,
                        "most_likely": 100,
                        "max_value": 120
                    },
                    {
                        "name": "B2",
                        "sheet_name": "SuperEngine Test",
                        "min_value": 150,
                        "most_likely": 200,
                        "max_value": 250
                    },
                    {
                        "name": "B3",
                        "sheet_name": "SuperEngine Test",
                        "min_value": 250,
                        "most_likely": 300,
                        "max_value": 350
                    }
                ],
                "result_cell_coordinate": "B6",
                "result_cell_sheet_name": "SuperEngine Test"
            }
            
            # Start simulation
            sim_response = await client.post(
                "http://localhost:8000/api/simulations/run",
                json=simulation_request,
                headers=headers
            )
            
            if sim_response.status_code not in [200, 202]:
                logger.error(f"❌ Simulation start failed: {sim_response.text}")
                return None
                
            sim_data = sim_response.json()
            simulation_id = sim_data["simulation_id"]
            logger.info(f"✅ Simulation started: {simulation_id}")
            
            # Poll for results
            max_attempts = 30
            for i in range(max_attempts):
                await asyncio.sleep(2)  # Wait 2 seconds between polls
                
                status_response = await client.get(
                    f"http://localhost:8000/api/simulations/{simulation_id}",
                    headers=headers
                )
                
                if status_response.status_code != 200:
                    logger.error(f"❌ Status check failed: {status_response.text}")
                    continue
                    
                status_data = status_response.json()
                logger.info(f"📊 Simulation status: {status_data['status']}")
                
                if status_data['status'] == 'completed':
                    logger.info("✅ Simulation completed successfully!")
                    return status_data
                elif status_data['status'] == 'failed':
                    logger.error(f"❌ Simulation failed: {status_data.get('message', 'Unknown error')}")
                    return None
            
            logger.error("❌ Simulation timeout - took too long to complete")
            return None
            
    except Exception as e:
        logger.error(f"❌ Simulation test failed: {e}")
        return None

async def analyze_results(results: Dict[str, Any]):
    """Analyze simulation results"""
    logger.info("🧪 Analyzing Results...")
    
    try:
        if 'results' not in results:
            logger.error("❌ No results found in response")
            return False
            
        sim_results = results['results']
        
        # Check basic statistics
        logger.info(f"📊 Mean: {sim_results.get('mean', 'N/A')}")
        logger.info(f"📊 Median: {sim_results.get('median', 'N/A')}")
        logger.info(f"📊 Std Dev: {sim_results.get('std_dev', 'N/A')}")
        logger.info(f"📊 Min: {sim_results.get('min_value', 'N/A')}")
        logger.info(f"📊 Max: {sim_results.get('max_value', 'N/A')}")
        
        # Verify results make sense
        # Expected: B1+B2+B3 = ~600, then *2 = ~1200, then *1.1 = ~1320 (since >1000)
        mean = sim_results.get('mean', 0)
        if 1200 < mean < 1400:
            logger.info("✅ Results are in expected range!")
            return True
        else:
            logger.error(f"❌ Results out of expected range: {mean}")
            return False
            
    except Exception as e:
        logger.error(f"❌ Result analysis failed: {e}")
        return False

async def main():
    """Run all tests"""
    logger.info("=" * 60)
    logger.info("🚀 SIMPLE SUPERENGINE TEST")
    logger.info("=" * 60)
    
    # Test 1: Upload Excel file
    upload_result = await test_excel_upload()
    if not upload_result:
        logger.error("❌ Excel upload failed, cannot continue")
        return False
        
    file_id, token = upload_result
    
    # Test 2: Run simulation
    sim_results = await test_simulation_setup(file_id, token)
    if not sim_results:
        logger.error("❌ Simulation failed")
        return False
    
    # Test 3: Analyze results
    success = await analyze_results(sim_results)
    
    logger.info("\n" + "=" * 60)
    logger.info("📊 TEST SUMMARY")
    logger.info("=" * 60)
    
    if success:
        logger.info("✅ ALL TESTS PASSED!")
        logger.info("✅ SuperEngine is working correctly!")
    else:
        logger.info("❌ Some tests failed")
        
    return success

if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1) 