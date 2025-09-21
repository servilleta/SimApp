#!/usr/bin/env python3
"""
🔍 FULL SIMULATION TEST
Test complete Monte Carlo simulation flow to find where zero results occur.
"""

import asyncio
import tempfile
import uuid
from openpyxl import Workbook
from simulation.engine import MonteCarloSimulation
from simulation.schemas import VariableConfig, SimulationRequest
from simulation.service import run_monte_carlo_simulation_task
from excel_parser.service import upload_excel_file
import json

async def test_full_simulation_flow():
    """Test complete simulation flow to find zero results issue"""
    
    print("🚀 STARTING FULL SIMULATION FLOW TEST")
    print("=" * 60)
    
    # Create a simple Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Set up test data
    ws['A1'] = 5        # Constant value
    ws['B1'] = 10       # This will be our Monte Carlo variable  
    ws['C1'] = "=A1+B1" # Simple formula
    ws['D1'] = "=C1*2"  # Target formula
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    print(f"📁 Created test Excel file: {temp_file.name}")
    print(f"📊 Excel content:")
    print(f"   A1 = 5 (constant)")
    print(f"   B1 = 10 (Monte Carlo variable)")
    print(f"   C1 = =A1+B1")
    print(f"   D1 = =C1*2 (target)")
    
    try:
        # Test 1: Upload Excel file
        print("\n🔍 TEST 1: Uploading Excel file...")
        file_id = str(uuid.uuid4())
        
        # Read file content
        with open(temp_file.name, 'rb') as f:
            file_content = f.read()
        
        # Simulate file upload
        try:
            upload_result = await upload_excel_file(
                file_id=file_id,
                filename="test_simulation.xlsx",
                file_content=file_content
            )
            print(f"✅ File uploaded successfully: {upload_result}")
            
        except Exception as e:
            print(f"❌ File upload failed: {e}")
            return
        
        # Test 2: Create simulation request
        print("\n🔍 TEST 2: Creating simulation request...")
        
        simulation_request = SimulationRequest(
            simulation_id=str(uuid.uuid4()),
            file_id=file_id,
            variables=[
                VariableConfig(
                    sheet_name="TestSheet",
                    name="B1",
                    min_value=8.0,
                    most_likely=10.0,
                    max_value=12.0
                )
            ],
            result_cell_sheet_name="TestSheet",
            result_cell_coordinate="D1",
            iterations=10  # Small number for testing
        )
        
        print(f"✅ Simulation request created:")
        print(f"   Simulation ID: {simulation_request.simulation_id}")
        print(f"   File ID: {simulation_request.file_id}")
        print(f"   Target: {simulation_request.result_cell_sheet_name}!{simulation_request.result_cell_coordinate}")
        print(f"   Variables: {len(simulation_request.variables)}")
        print(f"   Iterations: {simulation_request.iterations}")
        
        # Test 3: Run simulation
        print("\n🔍 TEST 3: Running simulation...")
        
        try:
            # This will run the actual simulation task
            result = await run_monte_carlo_simulation_task(simulation_request)
            
            print(f"✅ Simulation completed!")
            print(f"✅ Result: {result}")
            
        except Exception as e:
            print(f"❌ Simulation failed: {e}")
            import traceback
            traceback.print_exc()
            
        # Test 4: Check simulation results
        print("\n🔍 TEST 4: Checking simulation results...")
        
        from simulation.service import SIMULATION_RESULTS_STORE
        
        if simulation_request.simulation_id in SIMULATION_RESULTS_STORE:
            stored_result = SIMULATION_RESULTS_STORE[simulation_request.simulation_id]
            print(f"✅ Found stored result:")
            print(f"   Status: {stored_result.status}")
            print(f"   Message: {stored_result.message}")
            
            if hasattr(stored_result, 'results') and stored_result.results:
                results = stored_result.results
                print(f"   Results:")
                print(f"     Mean: {results.mean}")
                print(f"     Std Dev: {results.std_dev}")
                print(f"     Min: {results.min_value}")
                print(f"     Max: {results.max_value}")
                print(f"     Iterations: {results.iterations_run}")
                
                # Check for zero results issue
                if results.mean == 0.0 and results.std_dev == 0.0:
                    print(f"❌ FOUND THE ISSUE: All results are zero!")
                    print(f"❌ This confirms the zero results problem exists in the full simulation flow")
                else:
                    print(f"✅ Results look good - no zero issue found")
                
                # Expected analysis
                print(f"\n📊 Expected results analysis:")
                print(f"   D1 = (A1 + B1) * 2 = (5 + B1) * 2")
                print(f"   With B1 ranging 8-12, D1 should be 26-34")
                print(f"   Expected mean: ~30")
                
            else:
                print(f"❌ No results data found")
        else:
            print(f"❌ Simulation ID not found in results store")
            print(f"Available IDs: {list(SIMULATION_RESULTS_STORE.keys())}")
        
    finally:
        # Cleanup
        import os
        try:
            os.unlink(temp_file.name)
            print(f"\n🧹 Cleaned up test file: {temp_file.name}")
        except:
            pass
    
    print("\n" + "=" * 60)
    print("🏁 FULL SIMULATION FLOW TEST COMPLETED")

if __name__ == "__main__":
    asyncio.run(test_full_simulation_flow()) 