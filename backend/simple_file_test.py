#!/usr/bin/env python3
"""
Simple File Size and Processing Test

Tests file creation, parsing, and memory usage without requiring the full FastAPI stack.
"""

import os
import time
import tempfile
import psutil
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook

def get_memory_mb():
    """Get current memory usage in MB."""
    return psutil.Process().memory_info().rss / 1024 / 1024

def create_excel_file(rows: int, cols: int, formula_percentage: float = 0.1) -> str:
    """Create an Excel file with specified dimensions and formula density."""
    print(f"🏗️ Creating Excel file: {rows} x {cols} = {rows*cols:,} cells ({formula_percentage:.1%} formulas)")
    
    start_time = time.time()
    start_memory = get_memory_mb()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Fill the sheet
    formula_count = 0
    target_formulas = int(rows * cols * formula_percentage)
    
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            
            # Add formulas based on target percentage
            if formula_count < target_formulas and (row * col) % int(1/formula_percentage) == 0:
                if col > 1:
                    cell.value = f"=A{row}*RAND()+{col}"
                else:
                    cell.value = "=RAND()*100"
                formula_count += 1
            else:
                # Data cells with varying content
                if (row + col) % 3 == 0:
                    cell.value = (row * col * 0.1) % 1000  # Numbers
                else:
                    cell.value = f"Data_{row}_{col}"  # Text
    
    # Save to temporary file
    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, f"test_{rows}x{cols}.xlsx")
    wb.save(file_path)
    
    end_time = time.time()
    end_memory = get_memory_mb()
    
    # Get file stats
    file_size = os.path.getsize(file_path)
    file_mb = file_size / (1024 * 1024)
    
    stats = {
        "file_path": file_path,
        "dimensions": f"{rows}x{cols}",
        "total_cells": rows * cols,
        "formula_count": formula_count,
        "file_size_mb": file_mb,
        "creation_time": end_time - start_time,
        "memory_used_mb": end_memory - start_memory
    }
    
    print(f"✅ Created: {file_mb:.1f} MB, {formula_count:,} formulas in {stats['creation_time']:.2f}s")
    print(f"   Memory used: {stats['memory_used_mb']:.1f} MB")
    
    return file_path, stats

def test_excel_parsing(file_path: str) -> dict:
    """Test Excel file parsing performance."""
    print(f"\n🔍 Testing Excel parsing: {os.path.basename(file_path)}")
    
    start_time = time.time()
    start_memory = get_memory_mb()
    
    try:
        # Load with openpyxl
        wb = load_workbook(file_path, data_only=False)
        ws = wb.active
        
        # Count cells and formulas
        total_cells = 0
        formula_cells = 0
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                total_cells += 1
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_cells += 1
        
        end_time = time.time()
        end_memory = get_memory_mb()
        
        stats = {
            "success": True,
            "parsing_time": end_time - start_time,
            "memory_used_mb": end_memory - start_memory,
            "total_cells": total_cells,
            "formula_cells": formula_cells,
            "formula_density": formula_cells / max(1, total_cells),
            "max_row": max_row,
            "max_col": max_col
        }
        
        print(f"✅ Parsed: {total_cells:,} cells, {formula_cells:,} formulas in {stats['parsing_time']:.2f}s")
        print(f"   Memory used: {stats['memory_used_mb']:.1f} MB")
        
        return stats
        
    except Exception as e:
        print(f"❌ Parsing failed: {e}")
        return {
            "success": False,
            "error": str(e),
            "parsing_time": time.time() - start_time,
            "memory_used_mb": get_memory_mb() - start_memory
        }

def run_size_category_test(category: str, rows: int, cols: int, formula_pct: float = 0.1) -> dict:
    """Run a complete test for a specific file size category."""
    print(f"\n🎯 TESTING {category.upper()} FILE")
    print("-" * 50)
    
    initial_memory = get_memory_mb()
    
    try:
        # Create file
        file_path, creation_stats = create_excel_file(rows, cols, formula_pct)
        
        # Parse file
        parsing_stats = test_excel_parsing(file_path)
        
        # Final memory check
        final_memory = get_memory_mb()
        
        # Cleanup
        os.remove(file_path)
        os.rmdir(os.path.dirname(file_path))
        
        # Compile results
        results = {
            "category": category,
            "success": creation_stats and parsing_stats.get("success", False),
            "creation": creation_stats,
            "parsing": parsing_stats,
            "total_memory_used_mb": final_memory - initial_memory,
            "peak_memory_mb": final_memory
        }
        
        status = "✅" if results["success"] else "❌"
        print(f"{status} {category.upper()} RESULT: {creation_stats['file_size_mb']:.1f} MB file processed")
        
        return results
        
    except Exception as e:
        print(f"❌ {category.upper()} FAILED: {e}")
        return {
            "category": category,
            "success": False,
            "error": str(e),
            "total_memory_used_mb": get_memory_mb() - initial_memory
        }

def main():
    """Run comprehensive file size tests."""
    print("🚀 BIG FILE PROCESSING SMOKE TEST")
    print("=" * 60)
    
    # System info
    memory = psutil.virtual_memory()
    print(f"💾 System Memory: {memory.total / (1024**3):.1f} GB total, {memory.available / (1024**3):.1f} GB available")
    print(f"🖥️ CPU Cores: {psutil.cpu_count()}")
    
    baseline_memory = get_memory_mb()
    print(f"📊 Baseline Memory: {baseline_memory:.1f} MB")
    
    # Test scenarios - progressively larger files
    test_scenarios = [
        ("small", 100, 50, 0.1),      # 5K cells, ~1MB
        ("medium", 200, 100, 0.1),    # 20K cells, ~3-5MB
        ("large", 300, 200, 0.1),     # 60K cells, ~10-15MB
        ("xlarge", 500, 300, 0.1),    # 150K cells, ~25-40MB
        ("xxlarge", 700, 400, 0.1),   # 280K cells, ~50-80MB
    ]
    
    results = {}
    
    for category, rows, cols, formula_pct in test_scenarios:
        estimated_cells = rows * cols
        
        # Skip if it would exceed reasonable limits for testing
        if estimated_cells > 500000:  # 500K cells limit for testing
            print(f"\n⚠️ Skipping {category} test ({estimated_cells:,} cells) - too large for smoke test")
            continue
            
        result = run_size_category_test(category, rows, cols, formula_pct)
        results[category] = result
        
        # Memory check after each test
        current_memory = get_memory_mb()
        print(f"   Current memory: {current_memory:.1f} MB (+{current_memory - baseline_memory:.1f} MB from baseline)")
        
        # Break if we see memory issues
        if current_memory - baseline_memory > 2000:  # 2GB increase
            print("⚠️ High memory usage detected, stopping further tests")
            break
    
    # Summary
    print("\n🎉 TEST SUMMARY")
    print("=" * 60)
    
    successful_tests = [r for r in results.values() if r.get("success", False)]
    total_tests = len(results)
    
    print(f"✅ Success Rate: {len(successful_tests)}/{total_tests} ({len(successful_tests)/max(1,total_tests):.1%})")
    
    if successful_tests:
        max_size = max(r["creation"]["file_size_mb"] for r in successful_tests if "creation" in r)
        max_cells = max(r["creation"]["total_cells"] for r in successful_tests if "creation" in r)
        avg_parse_time = sum(r["parsing"]["parsing_time"] for r in successful_tests if "parsing" in r) / len(successful_tests)
        
        print(f"📊 Largest file processed: {max_size:.1f} MB ({max_cells:,} cells)")
        print(f"⏱️ Average parsing time: {avg_parse_time:.2f}s")
        
        # Performance analysis
        if max_cells >= 100000:
            print("🚀 System handles large files (100K+ cells) successfully")
        elif max_cells >= 50000:
            print("✅ System handles medium-large files (50K+ cells)")
        elif max_cells >= 10000:
            print("✅ System handles medium files (10K+ cells)")
        else:
            print("⚠️ System limited to small files (<10K cells)")
    
    # Memory analysis
    final_memory = get_memory_mb()
    total_memory_increase = final_memory - baseline_memory
    
    print(f"\n💾 MEMORY ANALYSIS:")
    print(f"   Baseline: {baseline_memory:.1f} MB")
    print(f"   Final: {final_memory:.1f} MB")
    print(f"   Total increase: {total_memory_increase:.1f} MB")
    
    if total_memory_increase < 500:
        print("   ✅ Excellent memory management")
    elif total_memory_increase < 1000:
        print("   ✅ Good memory management")
    elif total_memory_increase < 2000:
        print("   ⚠️ Moderate memory usage")
    else:
        print("   ❌ High memory usage - investigate memory leaks")
    
    # Recommendations
    print(f"\n💡 RECOMMENDATIONS:")
    
    if len(successful_tests) == total_tests:
        print("   ✅ All tests passed - system ready for big file processing")
        print("   🚀 Proceed with production deployment")
    elif len(successful_tests) >= total_tests * 0.8:
        print("   ✅ Most tests passed - system mostly ready")
        print("   🔧 Monitor memory usage in production")
    else:
        print("   ⚠️ Some tests failed - investigate issues")
        print("   🔧 Consider optimizing file processing pipeline")
    
    # Save detailed results
    results_file = "big_file_test_results.json"
    with open(results_file, "w") as f:
        json.dump({
            "timestamp": time.time(),
            "system_info": {
                "memory_gb": memory.total / (1024**3),
                "cpu_cores": psutil.cpu_count(),
                "baseline_memory_mb": baseline_memory,
                "final_memory_mb": final_memory
            },
            "test_results": results,
            "summary": {
                "success_rate": len(successful_tests) / max(1, total_tests),
                "total_tests": total_tests,
                "successful_tests": len(successful_tests),
                "total_memory_increase_mb": total_memory_increase
            }
        }, f, indent=2)
    
    print(f"\n📁 Detailed results saved to: {results_file}")

if __name__ == "__main__":
    main() 