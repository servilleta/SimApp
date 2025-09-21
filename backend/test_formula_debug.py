#!/usr/bin/env python3
"""
🔍 FORMULA EVALUATION DEBUG TEST
Test script to debug why all Monte Carlo simulations return zero results.
"""

import asyncio
import tempfile
import pandas as pd
from openpyxl import Workbook
from simulation.engine import MonteCarloSimulation, _safe_excel_eval, SAFE_EVAL_NAMESPACE
from simulation.schemas import VariableConfig
from simulation.formula_utils import get_evaluation_order
from excel_parser.service import get_formulas_for_file, get_all_parsed_sheets_data

async def test_simple_formula_evaluation():
    """Test basic formula evaluation with simple formulas"""
    
    print("🚀 STARTING FORMULA EVALUATION DEBUG TEST")
    print("=" * 60)
    
    # Create a simple Excel file with basic formulas
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # Set up simple test data
    ws['A1'] = 5        # Constant value
    ws['B1'] = 10       # This will be our Monte Carlo variable
    ws['C1'] = "=A1+B1" # Simple formula: should be 5 + B1
    ws['D1'] = "=C1*2"  # Another formula: should be (5 + B1) * 2
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    temp_file.close()
    
    print(f"📁 Created test file: {temp_file.name}")
    print(f"📊 Test data:")
    print(f"   A1 = 5 (constant)")
    print(f"   B1 = [Monte Carlo variable: 8-12 range]")
    print(f"   C1 = =A1+B1 (should be 5 + B1)")
    print(f"   D1 = =C1*2 (should be (5 + B1) * 2)")
    print()
    
    try:
        # Test 1: Parse Excel file
        print("🔍 TEST 1: Parsing Excel file...")
        all_formulas = await get_formulas_for_file(temp_file.name)
        all_parsed_sheet_data = await get_all_parsed_sheets_data(temp_file.name)
        
        print(f"✅ Found formulas: {all_formulas}")
        print(f"✅ Parsed sheets: {len(all_parsed_sheet_data)} sheets")
        
        # Test 2: Set up Monte Carlo configuration
        print("\n🔍 TEST 2: Setting up Monte Carlo configuration...")
        mc_input_configs = [
            VariableConfig(
                sheet_name="TestSheet",
                name="B1",
                min_value=8.0,
                most_likely=10.0,
                max_value=12.0
            )
        ]
        
        # Test 3: Build constant values
        print("\n🔍 TEST 3: Building constant values...")
        constant_values = {}
        mc_input_cells = {("TestSheet", "B1")}
        
        for sheet_data in all_parsed_sheet_data:
            for row_idx, row in enumerate(sheet_data.grid_data):
                for col_idx, cell_data_obj in enumerate(row):
                    if cell_data_obj:
                        cell_key = (sheet_data.sheet_name, cell_data_obj.coordinate.upper())
                        if cell_key not in mc_input_cells:
                            constant_values[cell_key] = cell_data_obj.value
        
        print(f"✅ Constant values: {constant_values}")
        print(f"✅ MC input cells: {mc_input_cells}")
        
        # Test 4: Get evaluation order
        print("\n🔍 TEST 4: Getting evaluation order...")
        try:
            ordered_calc_steps = get_evaluation_order(
                target_sheet_name="TestSheet",
                target_cell_coord="D1",  # Our target is D1
                all_formulas=all_formulas,
                mc_input_cells=mc_input_cells
            )
            print(f"✅ Evaluation order: {ordered_calc_steps}")
        except Exception as e:
            print(f"❌ Error getting evaluation order: {e}")
            return
        
        # Test 5: Manual formula evaluation test
        print("\n🔍 TEST 5: Manual formula evaluation test...")
        
        # Create test iteration values
        test_iter_values = constant_values.copy()
        test_iter_values[("TestSheet", "B1")] = 10.0  # Set B1 to 10 for testing
        
        print(f"🔍 Test iteration values: {test_iter_values}")
        
        # Test each formula step by step
        for sheet, cell, formula in ordered_calc_steps:
            print(f"\n🔍 Testing formula in {sheet}!{cell}: {formula}")
            
            try:
                result = _safe_excel_eval(
                    formula_string=formula,
                    current_eval_sheet=sheet,
                    all_current_iter_values=test_iter_values,
                    safe_eval_globals=SAFE_EVAL_NAMESPACE,
                    current_calc_cell_coord=f"{sheet}!{cell}",
                    constant_values=constant_values
                )
                
                print(f"✅ Result: {result} (type: {type(result)})")
                test_iter_values[(sheet, cell)] = result
                
                # Check if result is suspicious
                if result == 0 or result == 0.0:
                    print(f"❌ SUSPICIOUS: Formula '{formula}' returned zero!")
                
            except Exception as e:
                print(f"❌ Error evaluating formula: {e}")
        
        # Test 6: Run full Monte Carlo simulation
        print("\n🔍 TEST 6: Running full Monte Carlo simulation (5 iterations)...")
        
        simulation_engine = MonteCarloSimulation(iterations=5)
        
        try:
            raw_results, iteration_errors = await simulation_engine.run_simulation(
                mc_input_configs=mc_input_configs,
                ordered_calc_steps=ordered_calc_steps,
                target_sheet_name="TestSheet",
                target_cell_coordinate="D1",
                constant_values=constant_values
            )
            
            print(f"✅ Raw results: {raw_results}")
            print(f"✅ Iteration errors: {iteration_errors}")
            
            # Analyze results
            if raw_results is not None and len(raw_results) > 0:
                unique_results = set(raw_results)
                print(f"✅ Unique results: {unique_results}")
                
                if len(unique_results) == 1 and list(unique_results)[0] == 0.0:
                    print("❌ CRITICAL ISSUE: All results are zero!")
                elif len(unique_results) == 1:
                    print("❌ ISSUE: All results are the same value (no variation)")
                else:
                    print("✅ Results show variation - this is good!")
                    
                # Expected result analysis
                # D1 = (A1 + B1) * 2 = (5 + B1) * 2
                # With B1 ranging from 8-12, D1 should range from (5+8)*2=26 to (5+12)*2=34
                print(f"📊 Expected range: 26-34 (for B1 range 8-12)")
                print(f"📊 Actual range: {min(raw_results) if len(raw_results) > 0 else 'N/A'} - {max(raw_results) if len(raw_results) > 0 else 'N/A'}")
                
            else:
                print("❌ CRITICAL: No results returned!")
                
        except Exception as e:
            print(f"❌ Error running simulation: {e}")
            import traceback
            traceback.print_exc()
        
    finally:
        # Cleanup
        import os
        try:
            os.unlink(temp_file.name)
            print(f"\n🧹 Cleaned up test file: {temp_file.name}")
        except:
            pass
    
    print("\n" + "=" * 60)
    print("🏁 FORMULA EVALUATION DEBUG TEST COMPLETED")

if __name__ == "__main__":
    asyncio.run(test_simple_formula_evaluation()) 