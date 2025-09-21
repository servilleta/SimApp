"""
XLSX Export Service - Exact same logic as frontend
"""
import logging
from datetime import datetime
from typing import Dict, Any, List
from pathlib import Path
import tempfile

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - XLSX export will not work")

class XLSXExportService:
    """Service for generating XLSX exports using exact same logic as frontend"""
    
    def __init__(self):
        self.temp_dir = Path(tempfile.gettempdir()) / "monte_carlo_exports"
        self.temp_dir.mkdir(exist_ok=True)
        
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for XLSX export")
    
    def generate_xlsx_export(
        self, 
        simulation_id: str, 
        results_data: Dict[str, Any],
        metadata: Dict[str, Any] = None
    ) -> str:
        """
        Generate XLSX export using exact same logic as frontend.
        Returns path to the generated XLSX file.
        """
        try:
            logger.info(f"Starting XLSX export for simulation {simulation_id}")
            
            # Create workbook (same as frontend)
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Summary sheet (same structure as frontend)
            summary_ws = wb.create_sheet("Summary")
            
            # Header styling
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Summary header
            summary_ws.append([
                "Variable", "Mean", "Std Dev", "Min", "Max", 
                "95% VaR", "99% VaR", "Skewness", "Kurtosis"
            ])
            
            # Style header row
            for cell in summary_ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Add simulation metadata at the top
            summary_ws.insert_rows(1, 4)
            summary_ws['A1'] = 'Monte Carlo Simulation Results'
            summary_ws['A1'].font = Font(bold=True, size=14)
            summary_ws['A2'] = f'Simulation ID: {simulation_id}'
            summary_ws['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
            if metadata:
                summary_ws['A4'] = f'Iterations: {metadata.get("iterations", "N/A")}'
            
            # Process results (same logic as frontend)
            if "results" in results_data:
                row_num = 6  # Start after metadata
                for cell_name, cell_result in results_data["results"].items():
                    target_name = cell_result.get("cell_name", cell_name)
                    stats = cell_result.get("statistics", {})
                    
                    # Add row data (same format as frontend)
                    summary_ws.append([
                        target_name,
                        stats.get("mean"),
                        stats.get("std"),
                        stats.get("min"),
                        stats.get("max"),
                        stats.get("var_95"),
                        stats.get("var_99"),
                        stats.get("skewness"),
                        stats.get("kurtosis")
                    ])
                    
                    # Style data row
                    for cell in summary_ws[row_num]:
                        cell.border = border
                        if cell.column > 1:  # Numeric columns
                            cell.alignment = Alignment(horizontal='right')
                            cell.number_format = '0.0000'
                    
                    row_num += 1
            
            # Individual sheets for each variable with raw data (same as frontend)
            if "results" in results_data:
                for cell_name, cell_result in results_data["results"].items():
                    target_name = cell_result.get("cell_name", cell_name)
                    values = cell_result.get("values", [])
                    
                    if values:
                        # Clean sheet name for Excel compatibility (same logic as frontend)
                        import re
                        clean_sheet_name = re.sub(r'[^\w\s]', '', target_name).replace(' ', '_')[:31]
                        if not clean_sheet_name:
                            clean_sheet_name = f"Variable_{len(wb.sheetnames)}"
                        
                        # Create sheet
                        raw_ws = wb.create_sheet(clean_sheet_name)
                        
                        # Headers
                        raw_ws.append(['Iteration', 'Value'])
                        for cell in raw_ws[1]:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = border
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Add raw data (same format as frontend)
                        for i, value in enumerate(values, 1):
                            raw_ws.append([i, float(value)])
                        
                        # Auto-size columns
                        raw_ws.column_dimensions['A'].width = 12
                        raw_ws.column_dimensions['B'].width = 15
            
            # Auto-size summary sheet columns
            for column in summary_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                summary_ws.column_dimensions[column_letter].width = adjusted_width
            
            # Generate file path
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            xlsx_path = self.temp_dir / f"simulation_results_{simulation_id}_{timestamp}.xlsx"
            
            # Save file
            wb.save(xlsx_path)
            
            logger.info(f"XLSX export completed: {xlsx_path}")
            return str(xlsx_path)
            
        except Exception as e:
            logger.error(f"XLSX export failed for simulation {simulation_id}: {e}")
            raise

# Global instance
xlsx_export_service = XLSXExportService() if OPENPYXL_AVAILABLE else None
