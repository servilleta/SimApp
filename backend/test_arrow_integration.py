"""
Test Arrow Integration
Verifies that the Arrow-based Monte Carlo engine works correctly
"""

import asyncio
import sys
import os
import logging
import tempfile
from openpyxl import Workbook

# Add current directory to path for imports
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

async def test_arrow_schema():
    """Test Arrow schema definitions"""
    try:
        from arrow_utils.schema_builder import PARAMETERS_SCHEMA, RESULTS_SCHEMA, STATISTICS_SCHEMA
        from arrow_utils.schema_builder import create_empty_parameters_table, create_empty_results_table
        
        logger.info("Testing Arrow schemas...")
        
        # Test schema creation
        params_table = create_empty_parameters_table()
        results_table = create_empty_results_table()
        
        logger.info(f"Parameters schema: {params_table.schema}")
        logger.info(f"Results schema: {results_table.schema}")
        
        return True
        
    except Exception as e:
        logger.error(f"Schema test failed: {e}")
        return False

async def test_memory_manager():
    """Test Arrow memory manager"""
    try:
        from arrow_utils.memory_manager import get_memory_manager, MemoryConfig
        
        logger.info("Testing Arrow memory manager...")
        
        # Create memory manager
        memory_manager = get_memory_manager()
        
        # Get memory stats
        stats = memory_manager.get_memory_stats()
        logger.info(f"Memory stats: {stats}")
        
        # Test batch size calculation
        batch_size = memory_manager.get_optimal_batch_size(1000)  # 1KB per row
        logger.info(f"Optimal batch size: {batch_size}")
        
        return True
        
    except Exception as e:
        logger.error(f"Memory manager test failed: {e}")
        return False

async def test_excel_loader():
    """Test Excel to Arrow loader"""
    try:
        from arrow_engine.arrow_loader import ArrowExcelLoader
        
        logger.info("Testing Excel to Arrow loader...")
        
        # Create test Excel file
        test_file = create_test_excel_file()
        
        # Create loader
        loader = ArrowExcelLoader()
        
        # Load Excel to Arrow
        arrow_table = await loader.load_excel_to_arrow(test_file)
        
        logger.info(f"Loaded Arrow table: {len(arrow_table)} rows, {len(arrow_table.columns)} columns")
        logger.info(f"Schema: {arrow_table.schema}")
        
        # Print first few rows
        if len(arrow_table) > 0:
            first_rows = arrow_table.slice(0, min(3, len(arrow_table)))
            logger.info(f"First rows: {first_rows.to_pydict()}")
        
        # Cleanup
        os.unlink(test_file)
        
        return True
        
    except Exception as e:
        logger.error(f"Excel loader test failed: {e}")
        return False

async def test_streaming_processor():
    """Test Arrow streaming processor"""
    try:
        from arrow_engine.arrow_streaming import ArrowStreamProcessor
        from arrow_utils.schema_builder import create_empty_parameters_table
        import pyarrow as pa
        
        logger.info("Testing Arrow streaming processor...")
        
        # Create test parameters
        test_params = pa.Table.from_pydict({
            'cell_id': ['A1', 'B1', 'C1'],
            'formula': ['=NORM.INV(RAND(),100,10)', '=UNIFORM(50,150)', '=NORM.INV(RAND(),200,20)'],
            'distribution_type': ['normal', 'uniform', 'normal'],
            'param1': [100.0, 50.0, 200.0],
            'param2': [10.0, 150.0, 20.0],
            'param3': [0.0, 0.0, 0.0],
            'correlation_group': ['', '', ''],
            'dependencies': [[], [], []]
        })
        
        # Create processor
        processor = ArrowStreamProcessor(batch_size=5)
        
        # Test streaming with small number of iterations
        iterations = 10
        batch_count = 0
        
        async for batch in processor.stream_simulation_batches(test_params, iterations):
            batch_count += 1
            logger.info(f"Batch {batch_count}: {len(batch)} results")
            
            # Show sample results
            if batch_count == 1:
                sample = batch.slice(0, min(3, len(batch)))
                logger.info(f"Sample results: {sample.to_pydict()}")
        
        logger.info(f"Processed {batch_count} batches successfully")
        
        return True
        
    except Exception as e:
        logger.error(f"Streaming processor test failed: {e}")
        return False

def create_test_excel_file() -> str:
    """Create a test Excel file with Monte Carlo formulas"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    
    # Add some test formulas
    ws['A1'] = '=NORM.INV(RAND(),100,10)'  # Normal distribution
    ws['B1'] = '=UNIFORM(50,150)'           # Uniform distribution  
    ws['C1'] = '=NORM.INV(RAND(),200,20)'   # Another normal
    ws['D1'] = '=A1+B1+C1'                  # Dependent formula
    
    # Add some regular values
    ws['A2'] = 'Mean'
    ws['B2'] = 100
    ws['C2'] = 'StdDev'
    ws['D2'] = 10
    
    # Save to temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(temp_file.name)
    wb.close()
    temp_file.close()
    
    return temp_file.name

async def main():
    """Run all Arrow integration tests"""
    logger.info("🚀 Starting Arrow Integration Tests")
    logger.info("=" * 50)
    
    tests = [
        ("Arrow Schemas", test_arrow_schema),
        ("Memory Manager", test_memory_manager),
        ("Excel Loader", test_excel_loader),
        ("Streaming Processor", test_streaming_processor),
    ]
    
    passed = 0
    failed = 0
    
    for test_name, test_func in tests:
        logger.info(f"\n🧪 Running test: {test_name}")
        try:
            result = await test_func()
            if result:
                logger.info(f"✅ {test_name} PASSED")
                passed += 1
            else:
                logger.error(f"❌ {test_name} FAILED")
                failed += 1
        except Exception as e:
            logger.error(f"❌ {test_name} FAILED with exception: {e}")
            failed += 1
    
    logger.info("\n" + "=" * 50)
    logger.info(f"📊 Test Results: {passed} passed, {failed} failed")
    
    if failed == 0:
        logger.info("🎉 All Arrow integration tests PASSED!")
        return True
    else:
        logger.error("💥 Some tests FAILED - check logs above")
        return False

if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1) 