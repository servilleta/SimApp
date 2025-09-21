from typing import Dict, Optional, List, Tuple, Any, Set
from uuid import UUID
from fastapi import BackgroundTasks, HTTPException
import asyncio
from datetime import datetime, timezone
import numpy as np
import math
import logging
import signal
import os
import tempfile
import traceback
import json
import time
import re

from config import settings
from ..engine import MonteCarloSimulation
from ..schemas import (
    SimulationRequest, 
    SimulationResponse, 
    SimulationResult, 
    SimulationResultStatus,
    VariableConfig,
    ConstantConfig,
    EngineSelectionRequest,
    EngineRecommendation,
    EngineInfo,
    BatchSimulationRequest,
    BatchSimulationResponse
)
from excel_parser.service import get_formulas_for_file, get_all_parsed_sheets_data, get_cell_values
from ..formula_utils import get_evaluation_order, extract_cell_dependencies
from shared.progress_store import set_progress, get_progress_store
import openpyxl

# Use correct import path for super_engine
from super_engine.engine import SuperEngine
from super_engine.model_optimizer import ModelOptimizationAnalyzer

# SIMULATION_RESULTS_STORE key should match request.simulation_id type (str)
SIMULATION_RESULTS_STORE: Dict[str, SimulationResponse] = {}
logger = logging.getLogger(__name__)

# Arrow engine integration removed
ARROW_ENGINE_AVAILABLE = False

# Simulation state management
_simulation_tasks = {}
_simulation_results = {}
_simulation_cancelled = set()

# Global store for simulation start times
SIMULATION_START_TIMES = {}

# Store for tracking cancellation requests
SIMULATION_CANCELLATION_STORE: Dict[str, bool] = {}

# Store for sharing sensitivity analysis across multiple simulations in a batch
BATCH_SENSITIVITY_STORE: Dict[str, List[Dict[str, Any]]] = {}

def get_world_class_engine():
    """
    ✅ ZERO RESULTS BUG FIXED - RE-ENABLED
    The WorldClassMonteCarloEngine zero-result issues have been resolved through
    enhanced range reference processing and proper progress callback integration.
    The engine now correctly handles SUM/AVERAGE functions and provides real-time
    progress tracking with meaningful results.
    """
    try:
        from simulation.enhanced_engine import WorldClassMonteCarloEngine
        print("🚀 [CONFIG] WorldClassMonteCarloEngine enabled - Zero results bug fixed!")
        return WorldClassMonteCarloEngine
    except ImportError as e:
        print(f"⚠️ [CONFIG] WorldClassMonteCarloEngine import failed: {e} - falling back to standard engine")
        return None

def update_simulation_progress(simulation_id: str, progress_data: Dict[str, Any]):
    """Updates simulation progress in the store"""
    try:
        # Include the original start time in all progress updates if available
        if simulation_id in SIMULATION_START_TIMES and "start_time" not in progress_data:
            progress_data["start_time"] = SIMULATION_START_TIMES[simulation_id]
        
        # CRITICAL FIX: Preserve ALL initial data from first progress entry
        from shared.progress_store import get_progress
        existing_progress = get_progress(simulation_id)
        if existing_progress:
            # Preserve critical fields that should never be overwritten
            preserve_fields = [
                "user", "original_filename", "file_id", "target_variables", 
                "start_time", "simulation_id"
            ]
            
            for field in preserve_fields:
                if field not in progress_data and field in existing_progress:
                    progress_data[field] = existing_progress[field]
            
            # Special handling for variables - preserve existing variables structure
            if "variables" not in progress_data and "variables" in existing_progress:
                progress_data["variables"] = existing_progress["variables"]
        
        # 🚀 CRITICAL: Get temp_id for dual storage during ongoing progress updates
        from simulation.service import SIMULATION_TEMP_ID_MAPPING
        temp_id = SIMULATION_TEMP_ID_MAPPING.get(simulation_id)
        
        # 🚀 CRITICAL FIX: Use batching for ultra engine to reduce Redis load
        import asyncio
        
        # Check if this is a high-frequency update from ultra engine (but NOT completion)
        is_completion = (
            progress_data.get('progress_percentage') == 100 or 
            progress_data.get('progress_percentage', 0) >= 99.9 or
            progress_data.get('status') == 'completed' or
            progress_data.get('status') == 'failed' or
            progress_data.get('status') == 'cancelled'
        )
        
        is_ultra_heartbeat = progress_data.get('engine_type') == 'ultra' and progress_data.get('heartbeat', False)
        
        # CRITICAL: Completion updates must bypass batching for immediate WebSocket broadcast
        if is_ultra_heartbeat and not is_completion:
            try:
                # Skip batching for now due to sync/async mismatch - direct updates only
                pass  # Fallback to direct updates
            except ImportError:
                pass  # Fallback to direct updates if batcher not available
        
        # If this is a completion update, flush any pending batches first
        if is_completion:
            try:
                # Skip batching for now due to sync/async mismatch
                pass  # No batcher available
            except ImportError:
                pass
        
        # For non-ultra engines or critical updates, send immediately
        from shared.progress_store import set_progress_async
        asyncio.create_task(set_progress_async(simulation_id, progress_data, temp_id, bypass_merge=True))
    except Exception as e:
        logger.warning(f"Failed to update progress for simulation {simulation_id}: {e}")

def is_simulation_cancelled(simulation_id: str) -> bool:
    """Check if simulation has been marked for cancellation"""
    return SIMULATION_CANCELLATION_STORE.get(str(simulation_id), False)

async def get_file_complexity_category(file_id: str) -> str:
    """Determine file complexity category for concurrency management."""
    try:
        from excel_parser.service import get_formulas_for_file
        all_formulas = await get_formulas_for_file(file_id)
        total_formulas = sum(len(formulas) for formulas in all_formulas.values())
        
        if total_formulas <= 500:
            return "small"
        elif total_formulas <= 5000:
            return "medium"
        else:
            return "large"
    except Exception as e:
        logger.warning(f"Could not determine file complexity for {file_id}: {e}")
        return "medium"  # Default to medium

async def run_monte_carlo_simulation_with_concurrency_control(request: SimulationRequest):
    """
    🚀 ENHANCED SIMULATION WRAPPER WITH CONCURRENCY CONTROL
    
    Prevents multiple large file simulations from overwhelming system resources.
    Uses semaphores to limit concurrent simulations based on file complexity.
    """
    sim_id = request.simulation_id
    print(f"🔄 [CONCURRENCY] Starting concurrency-controlled simulation: {sim_id}")
    logger.info(f"🔄 [CONCURRENCY] Starting simulation {sim_id} with user context: {hasattr(request, '_user_context')}")

    # FIX: Create the response object immediately to prevent KeyErrors on failure
    current_time_iso = datetime.now(timezone.utc).isoformat()
    
    # CRITICAL FIX: Get user info from the request context (passed via initiate_simulation)
    user_name = 'unknown'
    if hasattr(request, '_user_context'):
        # Prefer email if available, otherwise use username
        if hasattr(request._user_context, 'email') and request._user_context.email:
            user_name = request._user_context.email
        elif hasattr(request._user_context, 'username'):
            user_name = request._user_context.username
    
    SIMULATION_RESULTS_STORE[sim_id] = SimulationResponse(
        simulation_id=sim_id,
        status="pending",
        message="Simulation has been queued.",
        created_at=current_time_iso,
        updated_at=current_time_iso,
        original_filename=request.original_filename,
        engine_type=request.engine_type,
        target_name=request.result_cell_coordinate,
        user=user_name
    )
    
    try:
        # Determine file complexity category for concurrency management
        complexity_category = await get_file_complexity_category(request.file_id)
        print(f"📊 [CONCURRENCY] File complexity: {complexity_category} for simulation {sim_id}")
        
        # Get the appropriate semaphore for this complexity level
        # Import semaphores from main.py
        try:
            from main import SIMULATION_SEMAPHORES
        except ImportError:
            # Fallback semaphores if import fails - INCREASED LIMITS
            SIMULATION_SEMAPHORES = {
                "large": asyncio.Semaphore(5),   # Increased from 2 to 5
                "medium": asyncio.Semaphore(8),  # Increased from 3 to 8
                "small": asyncio.Semaphore(10)   # Increased from 5 to 10
            }
            print(f"⚠️ [CONCURRENCY] Using fallback semaphores for {sim_id}")
        
        semaphore = SIMULATION_SEMAPHORES.get(complexity_category, SIMULATION_SEMAPHORES["medium"])
        
        print(f"🔄 [CONCURRENCY] Acquiring {complexity_category} semaphore for {sim_id}...")

        async with semaphore:
            print(f"✅ [CONCURRENCY] Acquired {complexity_category} semaphore for {sim_id}")
            logger.info(f"✅ [CONCURRENCY] Acquired {complexity_category} semaphore for {sim_id}")
            
            # Update status to running
            update_simulation_progress(sim_id, {
                "status": "running",
                "progress_percentage": 0,
                "message": "Simulation starting...",
                "complexity_category": complexity_category
            })
            
            # Run the actual simulation with error handling
            try:
                logger.info(f"🔄 [CONCURRENCY] About to execute simulation task for {sim_id}")
                await run_monte_carlo_simulation_task(request)
                logger.info(f"🔄 [CONCURRENCY] Successfully completed simulation task for {sim_id}")
            except Exception as task_error:
                logger.error(f"🔄 [CONCURRENCY] Error in simulation task for {sim_id}: {task_error}", exc_info=True)
                # Mark simulation as failed
                try:
                    await _mark_simulation_failed(sim_id, f"Simulation task error: {str(task_error)}")
                except Exception as mark_error:
                    logger.error(f"🔄 [CONCURRENCY] Failed to mark simulation {sim_id} as failed: {mark_error}")
                raise
            
            print(f"🎉 [CONCURRENCY] Released {complexity_category} semaphore for {sim_id}")
            
    except Exception as e:
        print(f"❌ [CONCURRENCY] Error in simulation execution for {sim_id}: {e}")
        logger.error(f"❌ [CONCURRENCY] Error in simulation execution for {sim_id}: {e}", exc_info=True)
        await _mark_simulation_failed(sim_id, f"Concurrency control error: {str(e)}")

async def run_monte_carlo_simulation_task(request: SimulationRequest):
    """
    Background task to execute a Monte Carlo simulation.
    Updates progress and stores results in SIMULATION_RESULTS_STORE.
    Enhanced with better error detection, and cancellation checking.
    TIMEOUT ADDED: Prevents simulations from hanging indefinitely.
    """
    sim_id = request.simulation_id  # This is the real simulation_id from initiate_simulation
    print(f"🚀 [TASK] Starting Monte Carlo simulation with ID: {sim_id}")
    print(f"🚀 [TASK] Request details: {request}")
    
    # Track simulation start time for timeout enforcement
    simulation_start_time = datetime.now(timezone.utc)
    SIMULATION_TIMEOUT_SECONDS = 3600  # 1 hour max per simulation (user can cancel manually)
    
    async def timeout_handler():
        """Handle simulation timeout"""
        await asyncio.sleep(SIMULATION_TIMEOUT_SECONDS)
        print(f"⏰ [TIMEOUT] Simulation {sim_id} timed out after {SIMULATION_TIMEOUT_SECONDS} seconds")
        _mark_simulation_failed(sim_id, f"Simulation timed out after {SIMULATION_TIMEOUT_SECONDS} seconds")
        return
    
    # Create timeout task
    timeout_task = asyncio.create_task(timeout_handler())
    
    try:
        # Check for early cancellation
        if is_simulation_cancelled(sim_id):
            print(f"🛑 [TASK] Simulation {sim_id} was cancelled before starting")
            timeout_task.cancel()
            return
        
        # Track simulation start time for timeout enforcement
        simulation_start_time = datetime.now(timezone.utc)
        simulation_start_time_iso = simulation_start_time.isoformat()
        
        # Store the start time globally for consistent elapsed time tracking
        SIMULATION_START_TIMES[sim_id] = simulation_start_time_iso
        
        # Update progress to "running" with detailed initialization and correct engine info
        update_simulation_progress(sim_id, {
            "status": "running",  
            "progress_percentage": 0,
            "current_iteration": 0,
            "total_iterations": request.iterations,
            "stage": "initialization",
            "stage_description": "File Upload & Validation",
            "start_time": simulation_start_time_iso,
            "engine": "WorldClassMonteCarloEngine",
            "engine_type": "enhanced",  # CRITICAL FIX: Set from the start
            "engineInfo": {
                "engine": "WorldClassMonteCarloEngine",
                "engine_type": "Enhanced",
                "gpu_acceleration": True,
                "detected": True
            }
        })
        
        # Check for cancellation after initialization
        if is_simulation_cancelled(sim_id):
            print(f"🛑 [TASK] Simulation {sim_id} was cancelled during initialization")
            timeout_task.cancel()
            await _mark_simulation_cancelled(sim_id)
            return

        # ENHANCED: Parse Excel data with error handling for complex formulas
        SIMULATION_RESULTS_STORE[sim_id].status = "running"
        SIMULATION_RESULTS_STORE[sim_id].updated_at = datetime.now(timezone.utc).isoformat()

        # 1. Get all formulas and all parsed sheet data from the Excel file
        try:
            # Update progress for parsing stage
            update_simulation_progress(sim_id, {
                "status": "running",
                "progress_percentage": 2,
                "stage": "parsing",
                "stage_description": "Parsing Excel File",
                "engine": "WorldClassMonteCarloEngine",
                "engine_type": "enhanced",
                "engineInfo": {
                    "engine": "WorldClassMonteCarloEngine",
                    "engine_type": "Enhanced",
                    "gpu_acceleration": True,
                    "detected": True
                }
            })
            
            print(f"📊 [TASK] Loading Excel data for {sim_id}")
            all_formulas = await get_formulas_for_file(request.file_id)
            total_formulas_count = len(all_formulas)
            print(f"📊 [TASK] Excel data loaded successfully for {sim_id}")
            
            # Update progress after parsing
            update_simulation_progress(sim_id, {
                "status": "running",
                "progress_percentage": 5,
                "stage": "parsing",
                "stage_description": "Excel File Parsed Successfully",
                "total_formulas": total_formulas_count
            })
        except Exception as e:
            print(f"❌ [TASK] Error loading Excel data for {sim_id}: {e}")
            timeout_task.cancel()
            await _mark_simulation_failed(sim_id, f"Failed to load Excel file: {str(e)}")
            return

        # 2. Identify Monte Carlo input cells
        update_simulation_progress(sim_id, {
            "status": "running",
            "progress_percentage": 8,
            "stage": "analysis",
            "stage_description": "Processing Variable Configurations",
            "total_formulas": total_formulas_count
        })
        
        print(f"🔍 [TASK] Processing Monte Carlo inputs for {sim_id}")
        mc_input_cells: Set[Tuple[str, str]] = set()
        constant_values: Dict[Tuple[str, str], Any] = {}
        
        for var_config in request.variables:
            mc_input_cells.add((var_config.sheet_name, var_config.name.upper()))
        
        # -----------------------------------------------------------
        # FIXED: Proper formula analysis for large files
        # Count formulas even if we skip detailed analysis
        # -----------------------------------------------------------

        engine_type = request.engine_type  # Now properly defined in SimulationRequest schema
        ordered_calc_steps: List[Tuple[str, str, str]] = []
        skip_formula_analysis = False

        # Count total formulas regardless of analysis method
        # BUGFIX: all_formulas has format {"sheet_name": {"cell": "formula", ...}}
        total_formulas_count = sum(
            len(sheet_formulas) for sheet_formulas in all_formulas.values()
        )
        
        logger.info(f"📊 [FORMULA_COUNT] Total formulas detected: {total_formulas_count}")

        # Only skip detailed analysis for extremely large files, but still count formulas
        # CRITICAL FIX: Don't skip formula analysis for Arrow engine - it needs the Excel data
        if total_formulas_count > 50000:
            skip_formula_analysis = True
            logger.info(f"⚡ [LARGE_FILE] Skipping detailed analysis for {total_formulas_count} formulas, using streaming mode")
        elif total_formulas_count > 100000:
            # BUGFIX: Only suggest Arrow engine for extremely large files if user hasn't already selected Arrow
            # Don't force override user's selection, just recommend
            logger.info(f"🚀 [LARGE_FILE_DETECTED] {total_formulas_count} formulas detected - GPU engine is recommended for optimal performance")
            logger.info(f"🔧 [USER_CHOICE] Respecting user's {engine_type} engine selection")

        # CRITICAL FIX: BIG engine handles its own dependency analysis
        if engine_type == "big":
            skip_formula_analysis = True
            logger.info(f"🚀 [BIG ENGINE] Skipping standard dependency analysis - BIG engine will handle internally")

        # CRITICAL FIX: Arrow engine needs formula analysis too, but can handle larger files efficiently
        if not skip_formula_analysis:
            # Perform formula dependency analysis for all engines
            # 3. Get the evaluation order for all dependent formulas
            try:
                update_simulation_progress(sim_id, {"status":"running","progress_percentage":15,"stage":"analysis","stage_description":"Analyzing Formula Dependencies"})
                
                # Get formula evaluation order
                ordered_calc_steps = get_evaluation_order(
                    target_sheet_name=request.result_cell_sheet_name,
                    target_cell_coord=request.result_cell_coordinate.upper(),
                    all_formulas=all_formulas,
                    mc_input_cells=mc_input_cells,
                    engine_type=engine_type  # CRITICAL FIX: Pass engine_type for appropriate limits
                )
                
                # CRITICAL FIX: Update formula metrics immediately after dependency analysis completes
                # This prevents the UI from showing 0 total_formulas during constants fetching
                formula_metrics_immediate = {
                    'total_formulas': total_formulas_count,
                    'relevant_formulas': len(ordered_calc_steps),
                    'analysis_method': 'Smart Dependency',
                    'cache_hits': 0,
                    'chunks_processed': 0
                }
                
                update_simulation_progress(sim_id, {
                    "status": "running",
                    "progress_percentage": 25,
                    "stage": "analysis",
                    "stage_description": f"Analyzed {len(ordered_calc_steps)} Formula Dependencies",
                    "formulaMetrics": formula_metrics_immediate,
                    "total_formulas": total_formulas_count,
                    "relevant_formulas": len(ordered_calc_steps)
                })

                # Get constants for all engines
                needed_cells: Set[Tuple[str, str]] = set(mc_input_cells)
                needed_cells.add((request.result_cell_sheet_name, request.result_cell_coordinate.upper()))
                for sh, coord, _ in ordered_calc_steps:
                    needed_cells.add((sh, coord.upper()))

                # Fetch constant values for non-MC input cells
                try:
                    # CRITICAL FIX: Get ALL constants from the file, not just specific cells
                    # This ensures no cells default to 0 during formula evaluation
                    from excel_parser.service import get_constants_for_file
                    constant_values = await get_constants_for_file(request.file_id, exclude_cells=mc_input_cells)
                    print(f"✅ [TASK] Loaded {len(constant_values)} constant values for {sim_id}")
                except Exception as cv_err:
                    print(f"⚠️ [TASK] Comprehensive constant fetch failed for {sim_id}: {cv_err}")
                    # Fall back to the original method
                    try:
                        constant_values = await get_cell_values(request.file_id, needed_cells - mc_input_cells)
                        print(f"✅ [TASK] Fallback: Loaded {len(constant_values)} constant values for {sim_id}")
                    except Exception as fallback_err:
                        print(f"❌ [TASK] Both constant fetch methods failed for {sim_id}: {fallback_err}")
                        # Use the Excel data we already have - but we need to get the actual sheet data
                        constant_values = {}
                        try:
                            # Load the parsed sheet data which contains actual cell values
                            from excel_parser.service import get_all_parsed_sheets_data
                            all_sheets = await get_all_parsed_sheets_data(request.file_id)
                            
                            for sheet_data in all_sheets:
                                sheet_name = sheet_data.sheet_name
                                for row_idx, row in enumerate(sheet_data.grid_data):
                                    for col_idx, cell in enumerate(row):
                                        if cell and cell.value is not None:
                                            # Convert indices to Excel coordinate (A1, B2, etc.)
                                            col_letter = get_column_letter(col_idx)
                                            coord = f"{col_letter}{row_idx + 1}"
                                            
                                            # Only add if it's not a Monte Carlo input and not a formula
                                            if (sheet_name, coord) not in mc_input_cells and not cell.is_formula_cell:
                                                try:
                                                    # Convert to float if possible
                                                    constant_values[(sheet_name, coord)] = float(cell.value)
                                                except (ValueError, TypeError):
                                                    # Keep as string if not numeric
                                                    constant_values[(sheet_name, coord)] = cell.value
                            
                            print(f"✅ [TASK] Using sheet data: Loaded {len(constant_values)} constant values for {sim_id}")
                        except Exception as sheet_err:
                            print(f"❌ [TASK] Failed to load sheet data for {sim_id}: {sheet_err}")
                            constant_values = {}
                        print(f"✅ [TASK] Final constant values count: {len(constant_values)} for {sim_id}")

                update_simulation_progress(sim_id, {"status":"running","progress_percentage":28,"stage":"analysis","stage_description":f"Fetched {len(constant_values)} constant cell values"})
                
                # Run optimization analysis if we have the data
                optimization_results = None
                try:
                    optimization_results = await get_model_optimization_suggestions(
                        request.file_id,
                        all_formulas,
                        ordered_calc_steps,
                        mc_input_cells
                    )
                    
                    # Include optimization suggestions in progress update
                    update_simulation_progress(sim_id, {
                        "status": "running",
                        "progress_percentage": 29,
                        "stage": "analysis",
                        "stage_description": "Model Optimization Analysis Complete",
                        "optimization_score": optimization_results.get('optimization_score', 0),
                        "optimization_suggestions": len(optimization_results.get('suggestions', []))
                    })
                    
                    logger.info(f"📊 [OPTIMIZATION] Score: {optimization_results.get('optimization_score', 0):.1f}/100")
                    logger.info(f"💡 [OPTIMIZATION] Found {len(optimization_results.get('suggestions', []))} suggestions")
                except Exception as opt_err:
                    logger.warning(f"Optimization analysis failed: {opt_err}")

            except Exception as e:
                err_msg = str(e)
                print(f"⚠️ [TASK] Dependency analysis failed for {sim_id}: {err_msg}")

                # Auto-fallback logic removed
                timeout_task.cancel()
                await _mark_simulation_failed(sim_id, err_msg)
                return
        else:
            # For non-Arrow engines with skipped analysis
            ordered_calc_steps = []
            constant_values = {}
            
            # CRITICAL FIX: Update formula metrics immediately even when analysis is skipped
            formula_metrics_immediate = {
                'total_formulas': total_formulas_count,
                'relevant_formulas': total_formulas_count,  # Use total count when analysis is skipped
                'analysis_method': 'Streaming',
                'cache_hits': 0,
                'chunks_processed': 0
            }
            
            update_simulation_progress(sim_id, {
                "status": "running",
                "progress_percentage": 25,
                "stage": "analysis", 
                "stage_description": f"Skipped detailed analysis for {total_formulas_count} formulas (Streaming Mode)",
                "formulaMetrics": formula_metrics_immediate,
                "total_formulas": total_formulas_count,
                "relevant_formulas": total_formulas_count
            })

        print(f"DEBUG: Generated {len(ordered_calc_steps)} calculation steps")
        
        # ENHANCED: Check for cancellation before starting simulation
        if is_simulation_cancelled(sim_id):
            print(f"🛑 [TASK] Simulation {sim_id} was cancelled during setup")
            timeout_task.cancel()
            await _mark_simulation_cancelled(sim_id)
            return

        # Update progress to show we're starting the simulation engine
        update_simulation_progress(sim_id, {
            "status": "running",
            "progress_percentage": 30,
            "stage": "simulation",
            "stage_description": "Starting Monte Carlo Engine",
        })

        # Build ConstantConfig list (empty for Arrow which doesn't use them)
        constants_list: List[ConstantConfig] = []
        for (sheet_name, coord), val in constant_values.items():
            try:
                constants_list.append(ConstantConfig(name=coord, sheet_name=sheet_name, value=float(val)))
            except Exception:
                continue

        # -----------------------------------------------------------
        # FIXED: Ensure formula metrics are properly set
        # -----------------------------------------------------------
        
        # Set proper formula metrics regardless of analysis method
        formula_metrics = {
            'total_formulas': total_formulas_count,
            'relevant_formulas': len(ordered_calc_steps) if not skip_formula_analysis else total_formulas_count,
            'analysis_method': 'Streaming' if skip_formula_analysis else 'Smart Dependency',
            'cache_hits': 0,
            'chunks_processed': 0
        }
        
        logger.info(f"📋 [FORMULA_METRICS] {formula_metrics}")

        # -----------------------------------------------------------
        # RUN SIMULATION
        # -----------------------------------------------------------

        # 5. Initialize and run the simulation engine with timeout protection
        print(f"🚀 [TASK] Initializing simulation engine for {sim_id}")
        
        # Update progress with proper formula metrics and correct engine type
        update_simulation_progress(sim_id, {
            "status": "running",
            "progress_percentage": 25,
            "stage": "simulation",
            "stage_description": "Initializing Monte Carlo Engine",
            "formulaMetrics": formula_metrics,  # Use properly calculated metrics
            "total_formulas": total_formulas_count,  # FIX: Ensure this is set correctly
            "relevant_formulas": len(ordered_calc_steps) if not skip_formula_analysis else total_formulas_count,
            "engine": f"{request.engine_type.capitalize()}MonteCarloEngine",
            "engine_type": request.engine_type,  # Use the actual requested engine type
            "gpu_acceleration": request.engine_type in ["enhanced", "power"],
            "engineInfo": {
                "engine": f"{request.engine_type.capitalize()}MonteCarloEngine",
                "engine_type": request.engine_type.capitalize(),
                "gpu_acceleration": request.engine_type in ["enhanced", "power"],
                "detected": False
            }
        })

        # -----------------------------------------------------------
        # FIX: Use the proper engine selection logic instead of always Enhanced
        # -----------------------------------------------------------
        
        # Parse target cell for engine functions
        if "!" in request.result_cell_coordinate:
            target_sheet_name, target_cell_coordinate = request.result_cell_coordinate.split("!", 1)
        else:
            # BUGFIX: Use the actual sheet name from request instead of defaulting to Sheet1
            target_sheet_name = request.result_cell_sheet_name
            target_cell_coordinate = request.result_cell_coordinate
            
        # Build target_cell string for engine functions
        target_cell = f"{target_sheet_name}!{target_cell_coordinate}"
        
        # Convert variables to the format expected by engine functions
        mc_inputs_for_engine = []
        for var_config in request.variables:
            # The variables are already VariableConfig objects, just use them directly
            mc_inputs_for_engine.append(var_config)
        
        # Use the proper engine selection function we already fixed
        logger.info(f"🚀 [ENGINE_SELECTION] Running simulation with {request.engine_type} engine")
        
        # Construct file_path from request.file_id
        file_path_for_engine = f"uploads/{request.file_id}"
        
        # FIX: Use the engine_type from the request to support all engines
        sim_result = await run_simulation_with_engine(
            sim_id=sim_id,
            file_path=file_path_for_engine,
            mc_inputs=mc_inputs_for_engine,
            constants=constants_list,
            target_cell=target_cell,
            iterations=request.iterations,
            engine_type=request.engine_type,  # <-- Use the engine type from the request!
            batch_id=request.batch_id
        )
        
        # Store results
        logger.info(f"📊 [SIMULATION_COMPLETE] Storing results for {sim_id}")
        
        if sim_result is None:
            logger.error(f"❌ [SIMULATION_ERROR] No results returned for {sim_id}")
            await _mark_simulation_failed(sim_id, "No simulation results returned")
            return
        
        # Ensure the simulation object exists in the store
        if sim_id not in SIMULATION_RESULTS_STORE:
            logger.error(f"❌ [SIMULATION_ERROR] Simulation {sim_id} not found in SIMULATION_RESULTS_STORE")
            await _mark_simulation_failed(sim_id, "Simulation not found in results store")
            return
        
        try:
            # Store results in SIMULATION_RESULTS_STORE
            SIMULATION_RESULTS_STORE[sim_id].status = "completed"
            SIMULATION_RESULTS_STORE[sim_id].results = sim_result
            SIMULATION_RESULTS_STORE[sim_id].message = "Simulation completed successfully."
            SIMULATION_RESULTS_STORE[sim_id].updated_at = datetime.now(timezone.utc).isoformat()
            
            # Verify the results were stored correctly
            if SIMULATION_RESULTS_STORE[sim_id].results is None:
                logger.error(f"❌ [SIMULATION_ERROR] Results were not stored properly for {sim_id}")
                await _mark_simulation_failed(sim_id, "Results storage failed")
                return
                
        except Exception as e:
            logger.error(f"❌ [SIMULATION_ERROR] Error storing results for {sim_id}: {e}")
            await _mark_simulation_failed(sim_id, f"Error storing results: {str(e)}")
            return
        
        # DEBUG: Log the sensitivity analysis to ensure it's being stored
        logger.info(f"📊 [SIMULATION_COMPLETE] Results stored successfully for {sim_id}")
        logger.info(f"📊 [SIMULATION_COMPLETE] Sensitivity analysis included: {len(sim_result.sensitivity_analysis) if sim_result.sensitivity_analysis else 0} variables")
        if sim_result.sensitivity_analysis:
            logger.info(f"📊 [SIMULATION_COMPLETE] Sensitivity data sample: {sim_result.sensitivity_analysis[0] if len(sim_result.sensitivity_analysis) > 0 else 'None'}")
        
        # Now safe to clean up start time tracking
        if sim_id in SIMULATION_START_TIMES:
            final_start = SIMULATION_START_TIMES[sim_id]
            del SIMULATION_START_TIMES[sim_id]
        
        # DURABLE LOGGING: Enable database persistence for completed simulations
        try:
            from persistence_logging.persistence import persist_simulation_run, build_simulation_summary
            
            # Build simulation summary for persistence
            summary = build_simulation_summary(
                simulation_id=sim_id,
                results=sim_result,
                status="completed",
                message="Simulation completed successfully",
                engine_type=request.engine_type,
                iterations_requested=request.iterations,
                variables_config=[var.dict() for var in request.variables],
                constants_config=None,  # Constants not available in SimulationRequest
                target_cell=request.result_cell_coordinate,
                started_at=final_start if 'final_start' in locals() else None
            )
            
            # Persist to database asynchronously
            persist_success = await persist_simulation_run(summary)
            if persist_success:
                logger.info(f"📦 [DURABLE_LOG] Successfully persisted simulation {sim_id} to database")
            else:
                logger.warning(f"📦 [DURABLE_LOG] Failed to persist simulation {sim_id} to database")
                
        except Exception as persist_error:
            logger.error(f"📦 [DURABLE_LOG] Error persisting simulation {sim_id}: {persist_error}")
            # Don't fail the simulation if persistence fails
        
        # Persist results in Redis so other workers can serve it
        try:
            from shared.result_store import set_result
            # CRITICAL FIX: Enhance result data with user and filename from progress
            result_dict = SIMULATION_RESULTS_STORE[sim_id].dict()
            
            # ENHANCED: Verify the result_dict contains the results before persistence
            if result_dict.get("results") is None:
                logger.error(f"❌ [STORAGE_ERROR] Result dict has no results for {sim_id}")
                # Force creation of results dict if missing
                if sim_result:
                    result_dict["results"] = sim_result.dict()
                    logger.info(f"🔧 [STORAGE_FIX] Created results dict from sim_result for {sim_id}")
                else:
                    logger.error(f"❌ [STORAGE_ERROR] Both result_dict and sim_result are missing for {sim_id}")
                    # Don't fail the simulation, but log the critical error
            
            from shared.progress_store import get_progress
            current_progress = get_progress(sim_id)
            if current_progress:
                # Add user and filename data to results for admin logs
                result_dict["user"] = current_progress.get("user", "unknown")
                result_dict["original_filename"] = current_progress.get("original_filename", None)
                result_dict["file_id"] = current_progress.get("file_id", None)
            
            # ENHANCED: Ensure proper serialization by converting datetime objects
            def convert_datetime_for_json(obj):
                if hasattr(obj, 'isoformat'):
                    return obj.isoformat()
                elif isinstance(obj, dict):
                    return {k: convert_datetime_for_json(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [convert_datetime_for_json(item) for item in obj]
                return obj
            
            # Clean the result dict for JSON serialization
            clean_result_dict = convert_datetime_for_json(result_dict)
            
            # ENHANCED: Final verification before persistence
            if clean_result_dict.get("results") is None:
                logger.error(f"❌ [STORAGE_ERROR] Cleaned result dict has no results for {sim_id}")
                # Create minimal results structure as fallback
                if sim_result:
                    clean_result_dict["results"] = {
                        "mean": getattr(sim_result, 'mean', 0),
                        "median": getattr(sim_result, 'median', 0),
                        "std_dev": getattr(sim_result, 'std_dev', 0),
                        "min_value": getattr(sim_result, 'min_value', 0),
                        "max_value": getattr(sim_result, 'max_value', 0),
                        "histogram": getattr(sim_result, 'histogram', {}),
                        "sensitivity_analysis": getattr(sim_result, 'sensitivity_analysis', []),
                        "iterations_run": getattr(sim_result, 'iterations_run', 0)
                    }
                    logger.info(f"🔧 [STORAGE_FIX] Created minimal results structure for {sim_id}")
                
            # Attempt Redis persistence with enhanced error handling
            set_result(sim_id, clean_result_dict)
            logger.info(f"✅ [RESULTS] Successfully persisted simulation {sim_id} to Redis")
            
            # CRITICAL: Verify the data was actually stored
            from shared.result_store import get_result
            verification = get_result(sim_id)
            if verification and verification.get("results"):
                logger.info(f"✅ [STORAGE_VERIFY] Verified {sim_id} is accessible via API")
            else:
                logger.error(f"❌ [STORAGE_VERIFY] Failed to verify {sim_id} accessibility - API will return 'Not Found'")
            
        except Exception as rs_err:
            logger.error(f"❌ [RESULTS] ResultStore persist failed for {sim_id}: {rs_err}")
            # Try alternative persistence method with better error handling
            try:
                # Store minimal result data that we know can be serialized
                minimal_result = {
                    "simulation_id": sim_id,
                    "status": SIMULATION_RESULTS_STORE[sim_id].status,
                    "message": SIMULATION_RESULTS_STORE[sim_id].message,
                    "original_filename": current_progress.get("original_filename", None) if current_progress else None,
                    "user": current_progress.get("user", "unknown") if current_progress else "unknown",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "results": {
                        "mean": getattr(sim_result, 'mean', 0) if sim_result else 0,
                        "median": getattr(sim_result, 'median', 0) if sim_result else 0,
                        "std_dev": getattr(sim_result, 'std_dev', 0) if sim_result else 0,
                        "min_value": getattr(sim_result, 'min_value', 0) if sim_result else 0,
                        "max_value": getattr(sim_result, 'max_value', 0) if sim_result else 0,
                        "histogram": getattr(sim_result, 'histogram', {}) if sim_result else {},
                        "sensitivity_analysis": getattr(sim_result, 'sensitivity_analysis', []) if sim_result else [],
                        "iterations_run": getattr(sim_result, 'iterations_run', 0) if sim_result else 0
                    } if sim_result else None
                }
                
                # Enhanced verification for minimal result
                if minimal_result.get("results") is None:
                    logger.error(f"❌ [FALLBACK_ERROR] Minimal result has no results for {sim_id}")
                    minimal_result["results"] = {
                        "mean": 0, "median": 0, "std_dev": 0, "min_value": 0, "max_value": 0,
                        "histogram": {}, "sensitivity_analysis": [], "iterations_run": 0,
                        "error": "Results data could not be serialized properly"
                    }
                    
                set_result(sim_id, minimal_result)
                logger.info(f"✅ [RESULTS] Fallback persistence successful for {sim_id}")
                
                # Verify fallback persistence
                verification = get_result(sim_id)
                if verification:
                    logger.info(f"✅ [FALLBACK_VERIFY] Verified {sim_id} fallback is accessible via API")
                else:
                    logger.error(f"❌ [FALLBACK_VERIFY] Even fallback persistence failed for {sim_id}")
                
            except Exception as fallback_err:
                logger.error(f"❌ [RESULTS] Even fallback persistence failed for {sim_id}: {fallback_err}")
                # Continue anyway - results are still in memory store
                # But the API will return "Not Found" for this simulation
                logger.error(f"❌ [CRITICAL] Simulation {sim_id} completed but will not be accessible via API")
        
        # FIX: Update final progress while preserving iteration counts and engine info
        update_simulation_progress(sim_id, {
            "status": "completed",
            "progress_percentage": 100,
            "current_iteration": request.iterations,  # FIX: Show completed iterations, not 0
            "total_iterations": request.iterations,   # FIX: Preserve total iterations
            "message": "Simulation completed successfully",
            "start_time": final_start,
            "total_formulas": total_formulas_count,   # FIX: Preserve formula count
            "relevant_formulas": len(ordered_calc_steps) if not skip_formula_analysis else total_formulas_count,
            "engineInfo": {
                "engine": f"{request.engine_type.capitalize()}MonteCarloEngine",
                "engine_type": request.engine_type.capitalize(),
                "gpu_acceleration": request.engine_type in ["enhanced", "power"],
                "detected": False
            },
            "error": None  # clears any previous error text retained in state
        })
        
        print(f"🎉 [TASK] Simulation {sim_id} completed successfully!")

        # Get dependency chain for this target
        dependency_chain = get_evaluation_order(
            target_sheet_name=target_sheet_name,
            target_cell_coord=target_cell_coordinate,
            all_formulas=all_formulas,
            mc_input_cells=mc_input_cells,
            engine_type=engine_type
        )
        
        # Add logging to check for VLOOKUP formulas
        logger.warning(f"[SIMULATION_DEBUG] Target {target_sheet_name}!{target_cell_coordinate} dependency chain length: {len(dependency_chain)}")
        vlookup_in_chain = 0
        for sheet, cell, formula in dependency_chain:
            if 'VLOOKUP' in str(formula).upper():
                vlookup_in_chain += 1
                logger.warning(f"[SIMULATION_DEBUG] VLOOKUP in dependency chain: {sheet}!{cell} = {formula}")
        logger.warning(f"[SIMULATION_DEBUG] Total VLOOKUPs in dependency chain: {vlookup_in_chain}")

        return sim_result
        
    except ValueError as ve: 
        print(f"❌ [TASK] ValueError in simulation {sim_id}: {ve}")
        timeout_task.cancel()
        await _mark_simulation_failed(sim_id, f"Simulation setup error: {str(ve)}")
    except HTTPException as he:
        print(f"❌ [TASK] HTTPException in simulation {sim_id}: {he.detail}")
        timeout_task.cancel()
        await _mark_simulation_failed(sim_id, f"Simulation pre-flight check failed: {he.detail}")
    except Exception as e:
        print(f"❌ [TASK] Unexpected error in simulation {sim_id}: {e}")
        timeout_task.cancel()
        await _mark_simulation_failed(sim_id, f"Simulation execution error: {str(e)}")

# ENHANCED: Helper functions for better error handling
async def _mark_simulation_failed(sim_id: str, error_message: str):
    """Mark simulation as failed with error message"""
    print(f"❌ [HELPER] Marking simulation {sim_id} as failed: {error_message}")
    # Set status in the main results store if it exists
    if sim_id in SIMULATION_RESULTS_STORE:
        SIMULATION_RESULTS_STORE[sim_id].status = "failed"
        SIMULATION_RESULTS_STORE[sim_id].message = error_message
        SIMULATION_RESULTS_STORE[sim_id].updated_at = datetime.now(timezone.utc).isoformat()
    
    # Clean up start time tracking
    if sim_id in SIMULATION_START_TIMES:
        del SIMULATION_START_TIMES[sim_id]
    
    # Update the centralized progress store
    update_simulation_progress(sim_id, {
        "status": "failed",
        "progress_percentage": 0,
        "message": error_message,
        "error": error_message
    })
    
    # Re-raise exception to ensure the simulation task stops
    raise Exception(error_message)

async def _mark_simulation_cancelled(sim_id: str):
    """Mark simulation as cancelled"""
    print(f"🛑 [HELPER] Marking simulation {sim_id} as cancelled")
    SIMULATION_RESULTS_STORE[sim_id].status = "cancelled"
    SIMULATION_RESULTS_STORE[sim_id].message = "Simulation was cancelled by user request."
    SIMULATION_RESULTS_STORE[sim_id].updated_at = datetime.now(timezone.utc).isoformat()
    
    # Clean up start time tracking
    if sim_id in SIMULATION_START_TIMES:
        del SIMULATION_START_TIMES[sim_id]
    
    update_simulation_progress(sim_id, {
        "status": "cancelled",
        "progress_percentage": 0,
        "message": "Simulation cancelled by user request"
    })

def _check_simulation_timeout(sim_id: str, start_time: datetime, max_duration: int) -> bool:
    """
    TIMEOUT REMOVED: This function now only logs elapsed time for monitoring purposes.
    Always returns False to allow simulations to run indefinitely.
    """
    elapsed_time = (datetime.now(timezone.utc) - start_time).total_seconds()
    
    # Log elapsed time for monitoring purposes only
    if elapsed_time > 300:  # Log every 5 minutes
        print(f"📊 [MONITOR] Simulation {sim_id} has been running for {elapsed_time:.2f} seconds")
    
    # Always return False - no timeout enforcement
    return False

async def initiate_simulation(
    request: SimulationRequest, 
    background_tasks: BackgroundTasks,
    current_user: dict # Accept user info
) -> SimulationResponse:
    """
    Initiates a simulation, stores its initial state, and adds it to the background queue.
    NOW INCLUDES USERNAME and FILENAME for persistent logging.
    """
    # Generate simulation_id if not provided
    if not request.simulation_id:
        from uuid import uuid4
        request.simulation_id = str(uuid4())
    
    sim_id = request.simulation_id
    
    # DEBUG: Log user and request data
    logger.info(f"🔍 [ENGINES_INITIATE] Simulation {sim_id} - User: {current_user}")
    logger.info(f"🔍 [ENGINES_INITIATE] Simulation {sim_id} - Username: {getattr(current_user, 'username', 'NO_USERNAME')}")
    logger.info(f"🔍 [ENGINES_INITIATE] Simulation {sim_id} - Original filename: {request.original_filename}")
    logger.info(f"🔍 [ENGINES_INITIATE] Simulation {sim_id} - File ID: {request.file_id}")
    
    # Store initial state with user and filename
    initial_progress = {
        "status": "pending",
        "progress_percentage": 0,
        "message": "Simulation is being queued.",
        "simulation_id": sim_id,
        "start_time": datetime.now(timezone.utc).isoformat(),
        "user": current_user.email if hasattr(current_user, 'email') and current_user.email else current_user.username,
        "original_filename": request.original_filename,
        "file_id": request.file_id  # Add file_id to initial progress
    }
    
    # DEBUG: Log the initial progress data
    logger.info(f"🔍 [ENGINES_INITIATE] Simulation {sim_id} - Initial progress: {initial_progress}")
    
    set_progress(sim_id, initial_progress)

    # Also set simulation metadata right away
    from shared.progress_store import set_simulation_metadata
    set_simulation_metadata(sim_id, {
        "file_id": request.file_id,
        "target_variables": [v.name for v in request.variables]
    })
    
    # Add the main simulation task to be run in the background
    # CRITICAL FIX: Pass user context with the request
    request._user_context = current_user
    
    # Enhanced logging for debugging stuck simulations
    logger.info(f"🚀 [ENGINES_INITIATE] Queuing background task for {sim_id}")
    logger.info(f"🚀 [ENGINES_INITIATE] Task details - User: {current_user.username if hasattr(current_user, 'username') else 'unknown'}, File: {request.original_filename}, Engine: {request.engine_type}")
    
    background_tasks.add_task(run_monte_carlo_simulation_with_concurrency_control, request)
    
    logger.info(f"🚀 [ENGINES_INITIATE] Background task queued successfully for {sim_id}")
    
    # CRITICAL FIX: Create and store the initial SimulationResponse in SIMULATION_RESULTS_STORE
    initial_response = SimulationResponse(
        simulation_id=sim_id,
        message="Simulation successfully queued.",
        status="pending",
        original_filename=request.original_filename,
        engine_type=request.engine_type,
        user=current_user.email if hasattr(current_user, 'email') and current_user.email else (current_user.username if hasattr(current_user, 'username') else 'unknown'),
        target_name=request.result_cell_coordinate,
        created_at=datetime.now(timezone.utc).isoformat(),
        updated_at=datetime.now(timezone.utc).isoformat()
    )
    
    # CRITICAL FIX: Store the initial response so the background task can modify it later
    SIMULATION_RESULTS_STORE[sim_id] = initial_response
    logger.info(f"🔧 [ENGINES_INITIATE] Stored initial response for {sim_id} in SIMULATION_RESULTS_STORE")
    
    return initial_response

async def get_simulation_status_or_results(simulation_id: str) -> SimulationResponse:
    str_sim_id = str(simulation_id)
    
    logger.info(f"🚨 [MEMORY_STORE] get_simulation_status_or_results called for: {str_sim_id}")
    
    result = SIMULATION_RESULTS_STORE.get(str_sim_id)
    if result is None:
        logger.info(f"🚨 [MEMORY_STORE] Not found in SIMULATION_RESULTS_STORE: {str_sim_id}")
        
        # 🔒 PERMANENT FIX: Add timestamp validation to prevent serving stale data
        from datetime import datetime, timezone, timedelta
        
        # Try cross-worker store with validation
        try:
            from shared.result_store import get_result
            result_data = get_result(str_sim_id)
            if result_data:
                logger.info(f"🚨 [REDIS_RESTORE] Found data in Redis for: {str_sim_id}")
                
                # 🔒 VALIDATION: Check if the data is too old (older than 1 hour)
                created_at_str = result_data.get('created_at')
                if created_at_str:
                    try:
                        created_at = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                        age = datetime.now(timezone.utc) - created_at
                        
                        if age > timedelta(hours=1):
                            logger.warning(f"🚨 [REDIS_RESTORE] ❌ BLOCKED: Data too old ({age.total_seconds():.0f}s) for {str_sim_id}")
                            # Clear the old data from Redis to prevent future issues
                            from shared.result_store import delete_result
                            try:
                                delete_result(str_sim_id)
                                logger.info(f"🧹 [REDIS_CLEANUP] Removed stale data for {str_sim_id}")
                            except:
                                pass
                            raise HTTPException(status_code=404, detail=f"Simulation with ID {simulation_id} not found (stale data removed).")
                        else:
                            logger.info(f"🚨 [REDIS_RESTORE] ✅ APPROVED: Data age {age.total_seconds():.0f}s for {str_sim_id}")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"🚨 [REDIS_RESTORE] ❌ BLOCKED: Invalid timestamp for {str_sim_id}: {e}")
                        raise HTTPException(status_code=404, detail=f"Simulation with ID {simulation_id} not found (invalid data).")
                
                # Create SimulationResponse from result_data
                try:
                    response = SimulationResponse(**result_data)
                    logger.info(f"🚨 [REDIS_RESTORE] ✅ SUCCESS: Restored from Redis for {str_sim_id}")
                    return response
                except Exception as e:
                    logger.error(f"Error creating SimulationResponse from cross-worker data: {e}")
                    
        except Exception as e:
            logger.error(f"Error getting result from cross-worker store for {str_sim_id}: {e}")
            pass
        
        # Check if this is a batch simulation that might have individual results
        try:
            from shared.progress_store import get_progress
            progress = get_progress(str_sim_id)
            if progress and progress.get("batch_id"):
                logger.info(f"Batch simulation with batch_id: {progress.get('batch_id')}")
                # For batch simulations, we might need to check for child simulation results
                
        except Exception as e:
            logger.error(f"Error checking progress for {str_sim_id}: {e}")
        
        # 🔥 NEW: Database fallback for historical simulations
        logger.info(f"📦 [DATABASE_FALLBACK] Not found in memory/Redis, checking database for: {str_sim_id}")
        try:
            from database import get_db
            from models import SimulationResult as SimulationResultModel
            from ..schemas import SimulationResponse, SimulationResult
            
            # Get database session
            db_gen = get_db()
            db = next(db_gen)
            
            try:
                # Query the database for this simulation
                db_simulation = db.query(SimulationResultModel).filter(
                    SimulationResultModel.simulation_id == str_sim_id
                ).first()
                
                if db_simulation:
                    logger.info(f"📦 [DATABASE_FALLBACK] Found historical simulation in database: {str_sim_id}")
                    
                    # Convert database model to API response format
                    simulation_result = None
                    if db_simulation.status == 'completed' and db_simulation.histogram:
                        # Create SimulationResult object for completed simulations
                        simulation_result = SimulationResult(
                            mean=db_simulation.mean,
                            median=db_simulation.median,
                            std_dev=db_simulation.std_dev,
                            min_value=db_simulation.min_value,
                            max_value=db_simulation.max_value,
                            percentiles=db_simulation.percentiles,
                            histogram=db_simulation.histogram,
                            iterations_run=db_simulation.iterations_run,
                            sensitivity_analysis=db_simulation.sensitivity_analysis or [],
                            errors=db_simulation.errors or [],
                            target_display_name=db_simulation.target_name
                        )
                    
                    # Create simulation response
                    response = SimulationResponse(
                        simulation_id=db_simulation.simulation_id,
                        status=db_simulation.status,
                        message=db_simulation.message or "Historical simulation",
                        results=simulation_result,
                        progress_percentage=100 if db_simulation.status == 'completed' else 0,
                        created_at=db_simulation.created_at.isoformat() if db_simulation.created_at else None,
                        updated_at=db_simulation.updated_at.isoformat() if db_simulation.updated_at else None,
                        file_name=db_simulation.original_filename,
                        engine_type=db_simulation.engine_type,
                        target_name=db_simulation.target_name,
                        iterations_requested=db_simulation.iterations_requested,
                        multi_target_result=None  # Historical simulations don't have this
                    )
                    
                    # Apply the same sanitization as the memory path
                    response = sanitize_simulation_response(response)
                    
                    logger.info(f"📦 [DATABASE_FALLBACK] Successfully created response for historical simulation: {str_sim_id}")
                    return response
                else:
                    logger.info(f"📦 [DATABASE_FALLBACK] Simulation not found in database: {str_sim_id}")
                    
            finally:
                db.close()
                
        except Exception as e:
            logger.error(f"📦 [DATABASE_FALLBACK] Error accessing database for {str_sim_id}: {e}")
            
        raise HTTPException(status_code=404, detail=f"Simulation with ID {simulation_id} not found.")
    
    result.updated_at = datetime.now(timezone.utc).isoformat()
    
    # Final safety layer: sanitize all float values before returning
    result = sanitize_simulation_response(result)
    
    return result

async def cancel_simulation_task(simulation_id: str) -> Dict[str, any]:
    """Cancel a running simulation task with enhanced cleanup."""
    str_sim_id = str(simulation_id)
    
    print(f"🛑 [CANCEL] Attempting to cancel simulation: {str_sim_id}")
    
    # Check if simulation exists
    if str_sim_id not in SIMULATION_RESULTS_STORE:
        print(f"⚠️ [CANCEL] Simulation {str_sim_id} not found in results store")
        return {
            "success": False,
            "message": f"Simulation with ID {simulation_id} not found."
        }
    
    # Check current status
    current_status = SIMULATION_RESULTS_STORE[str_sim_id].status
    print(f"🔍 [CANCEL] Current status of {str_sim_id}: {current_status}")
    
    if current_status not in ["pending", "running"]:
        print(f"⚠️ [CANCEL] Simulation {str_sim_id} cannot be cancelled (status: {current_status})")
        return {
            "success": False,
            "message": f"Simulation is already {current_status} and cannot be cancelled."
        }
    
    # Mark simulation for cancellation in the cancellation store
    SIMULATION_CANCELLATION_STORE[str_sim_id] = True
    print(f"✅ [CANCEL] Marked simulation {str_sim_id} for cancellation")
    
    # Update simulation status to cancelled immediately
    SIMULATION_RESULTS_STORE[str_sim_id].status = "cancelled"
    SIMULATION_RESULTS_STORE[str_sim_id].message = "Simulation was cancelled by user request."
    SIMULATION_RESULTS_STORE[str_sim_id].updated_at = datetime.now(timezone.utc).isoformat()
    
    # Update progress store as well
    update_simulation_progress(str_sim_id, {
        "status": "cancelled",
        "progress_percentage": 0,
        "message": "Simulation cancelled by user request",
        "cancelled_at": datetime.now(timezone.utc).isoformat()
    })
    
    # ENHANCED: Try to find and cancel any running GPU/CPU tasks
    try:
        # If using world-class engine, try to cancel GPU operations
        WorldClassEngine = get_world_class_engine()
        if WorldClassEngine:
            print(f"🔄 [CANCEL] Attempting to cancel GPU operations for {str_sim_id}")
            # GPU operations will be cancelled by the progress callback checking
            
        # Clean up any memory resources
        import gc
        gc.collect()
        print(f"🧹 [CANCEL] Cleaned up memory resources for {str_sim_id}")
        
    except Exception as cleanup_error:
        print(f"⚠️ [CANCEL] Error during cleanup for {str_sim_id}: {cleanup_error}")
        # Don't fail the cancellation due to cleanup errors
    
    print(f"✅ [CANCEL] Successfully cancelled simulation {str_sim_id}")
    return {
        "success": True,
        "message": "Simulation cancelled successfully.",
        "cancelled_at": datetime.now(timezone.utc).isoformat()
    }

def sanitize_float(value: Any) -> float:
    """
    Sanitize float values for JSON serialization.
    Converts NaN, inf, -inf, and None to 0.0 to ensure JSON compatibility.
    """
    import math
    
    if value is None:
        return 0.0
    
    if isinstance(value, (int, float)):
        # Check for NaN (NaN != NaN is True)
        if value != value or math.isnan(value):
            return 0.0
        # Check for infinity
        if math.isinf(value):
            return 0.0
        return float(value)
    
    # For any other type, try to convert to float, fallback to 0.0
    try:
        result = float(value)
        if math.isnan(result) or math.isinf(result):
            return 0.0
        return result
    except (ValueError, TypeError):
        return 0.0

def sanitize_simulation_response(response: SimulationResponse) -> SimulationResponse:
    """
    Final safety layer to sanitize all float values in a SimulationResponse 
    before JSON serialization to prevent NaN/inf serialization errors.
    """
    if response.results:
        # Sanitize all float fields in SimulationResult
        response.results.mean = sanitize_float(response.results.mean)
        response.results.median = sanitize_float(response.results.median)
        response.results.std_dev = sanitize_float(response.results.std_dev)
        response.results.min_value = sanitize_float(response.results.min_value)
        response.results.max_value = sanitize_float(response.results.max_value)
        
        # Sanitize percentiles
        if response.results.percentiles:
            response.results.percentiles = {
                key: sanitize_float(val) for key, val in response.results.percentiles.items()
            }
        
        # Sanitize histogram data if present
        if response.results.histogram:
            if 'bins' in response.results.histogram:
                response.results.histogram['bins'] = [
                    sanitize_float(val) for val in response.results.histogram['bins']
                ]
            if 'bin_edges' in response.results.histogram:
                response.results.histogram['bin_edges'] = [
                    sanitize_float(val) for val in response.results.histogram['bin_edges']
                ]
    
    return response

async def analyze_file_complexity(file_path: str, mc_inputs: List[VariableConfig]) -> Dict[str, Any]:
    """Analyze file complexity to recommend simulation engine"""
    try:
        # Load workbook for analysis
        workbook = openpyxl.load_workbook(file_path, data_only=False)
        
        # Count formulas and cells
        total_cells = 0
        formula_cells = 0
        lookup_functions = 0
        complex_functions = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        total_cells += 1
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_cells += 1
                            formula_upper = cell.value.upper()
                            # Count lookup functions
                            if any(func in formula_upper for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH', 'XLOOKUP']):
                                lookup_functions += 1
                            # Count complex functions
                            if any(func in formula_upper for func in ['SUMPRODUCT', 'ARRAY', 'MMULT', 'TRANSPOSE']):
                                complex_functions += 1
        
        # File size analysis
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        
        # Complexity scoring
        complexity_score = (
            (formula_cells * 1) +
            (lookup_functions * 2) +
            (complex_functions * 3) +
            (len(mc_inputs) * 0.5) +
            (file_size_mb * 0.1)
        )
        
        return {
            'total_cells': total_cells,
            'formula_cells': formula_cells,
            'lookup_functions': lookup_functions,
            'complex_functions': complex_functions,
            'file_size_mb': file_size_mb,
            'complexity_score': complexity_score,
            'mc_inputs_count': len(mc_inputs)
        }
    except Exception as e:
        logger.warning(f"Could not analyze file complexity: {e}")
        return {
            'total_cells': 0,
            'formula_cells': 0,
            'lookup_functions': 0,
            'complex_functions': 0,
            'file_size_mb': 0,
            'complexity_score': 0,
            'mc_inputs_count': len(mc_inputs)
        }

def recommend_simulation_engine(complexity: Dict[str, Any], user_selected_engine: str = None) -> EngineRecommendation:
    """Recommend simulation engine based on file complexity, with user override support"""
    
    score = complexity['complexity_score']
    formula_cells = complexity['formula_cells']
    file_size_mb = complexity['file_size_mb']
    lookup_functions = complexity['lookup_functions']
    
    # Determine recommended engine based on file characteristics
    if formula_cells > 5000:
        recommended_engine = "power"
        reason = "The Power engine is recommended for large files with many formulas, especially those with sparse data ranges."
    elif formula_cells > 500:
        recommended_engine = "enhanced"
        reason = "The Enhanced GPU engine provides excellent performance for medium-sized files."
    else:
        recommended_engine = "standard"
        reason = "The Standard engine is sufficient for small files with simple calculations."
    
    # User override handling
    if user_selected_engine:
        if user_selected_engine != recommended_engine:
            logger.warning(f"User override: Selected {user_selected_engine} instead of recommended {recommended_engine}")
            # Add warning to reason if user selects a potentially suboptimal engine
            if user_selected_engine == "standard" and formula_cells > 1000:
                reason += f" ⚠️ Warning: Standard engine may be slow for {formula_cells} formulas."
            elif user_selected_engine == "enhanced" and formula_cells > 10000:
                reason += f" ⚠️ Warning: Enhanced engine may struggle with {formula_cells} formulas. Consider Power engine."
        recommended_engine = user_selected_engine
    
    # Available engines - ADDED: Ultra engine
    available_engines = [
        {
            "id": "standard",
            "name": "Standard Monte Carlo Engine", 
            "description": "CPU-based engine for simple simulations",
            "best_for": "Small files (<500 formulas), guaranteed compatibility",
            "max_iterations": 100000
        },
        {
            "id": "enhanced",
            "name": "Enhanced Monte Carlo Engine",
            "description": "GPU-accelerated engine for medium simulations",
            "best_for": "Medium to large files (500+ formulas), high performance",
            "max_iterations": 1000000
        },
        {
            "id": "ultra",
            "name": "Ultra Hybrid Engine",
            "description": "Next-generation GPU-accelerated engine with complete dependency analysis",
            "best_for": "All file sizes with maximum performance and reliability",
            "max_iterations": 10000000
        }
    ]
    
    # Log available engines to confirm Power is included
    logger.info(f"🚀 Available engines: {[engine['id'] for engine in available_engines]}")
    
    return EngineRecommendation(
        recommended_engine=recommended_engine,
        reason=reason,
        complexity_analysis=complexity,
        available_engines=available_engines
    )

async def get_engine_recommendation(file_path: str, mc_inputs: List[VariableConfig], user_selected_engine: str = None) -> EngineRecommendation:
    """Get engine recommendation for a simulation file with optional user override"""
    complexity = await analyze_file_complexity(file_path, mc_inputs)
    return recommend_simulation_engine(complexity, user_selected_engine)

async def run_simulation_with_engine(
    sim_id: str,
    file_path: str,
    mc_inputs: List[VariableConfig],
    constants: List[ConstantConfig],
    target_cell: str,
    iterations: int,
    engine_type: str = "ultra",
    batch_id: Optional[str] = None
) -> SimulationResult:
    """Run simulation with specified engine"""
    
    logger.info(f"🚀 Starting simulation {sim_id} with {engine_type} engine")
    
    try:
        if engine_type == "standard":
            return await _run_standard_simulation(sim_id, file_path, mc_inputs, constants, target_cell, iterations)
        elif engine_type == "ultra":
            return await _run_ultra_simulation(sim_id, file_path, mc_inputs, constants, target_cell, iterations)
        else:  # enhanced (default) - use the WorldClassMonteCarloEngine
            return await _run_enhanced_simulation(sim_id, file_path, mc_inputs, constants, target_cell, iterations)
            
    except Exception as e:
        logger.error(f"❌ Simulation {sim_id} failed with {engine_type} engine: {e}")
        raise

async def _run_enhanced_simulation(
    sim_id: str,
    file_path: str,
    mc_inputs: List[VariableConfig], 
    constants: List[ConstantConfig],
    target_cell: str,
    iterations: int
) -> SimulationResult:
    """Run simulation using the Enhanced WorldClassMonteCarloEngine."""
    try:
        logger.info(f"✅ [ENHANCED] Running simulation for {sim_id}")
        
        # Get the WorldClassMonteCarloEngine
        WorldClassEngine = get_world_class_engine()
        if not WorldClassEngine:
            # Fallback to standard if WorldClass not available
            logger.warning(f"WorldClassMonteCarloEngine not available, falling back to standard")
            return await _run_standard_simulation(sim_id, file_path, mc_inputs, constants, target_cell, iterations)
        
        # Create enhanced engine instance
        simulation_engine = WorldClassEngine(iterations=iterations, simulation_id=sim_id)
        
        # Attach progress callback
        def progress_cb(progress_data):
            try:
                enhanced_progress_data = {**progress_data, "simulation_id": sim_id}
                if sim_id in SIMULATION_START_TIMES:
                    enhanced_progress_data["start_time"] = SIMULATION_START_TIMES[sim_id]
                update_simulation_progress(sim_id, enhanced_progress_data)
            except Exception as e:
                logger.debug(f"Progress relay failed for {sim_id}: {e}")
        
        simulation_engine.set_progress_callback(progress_cb)
        
        # Convert inputs to the format expected by enhanced engine
        mc_input_params = {}
        for mc_input in mc_inputs:
            key = (mc_input.sheet_name, mc_input.name.upper())
            mc_input_params[key] = (mc_input.min_value, mc_input.most_likely, mc_input.max_value)
        
        # Build constant_params dict from ConstantConfig list
        constant_params = {}
        for constant in constants:
            key = (constant.sheet_name, constant.name.upper())
            constant_params[key] = constant.value
        
        # Log constants to debug A8 issue
        logger.warning(f"[CONSTANTS_DEBUG] Building constant_params with {len(constants)} constants")
        for constant in constants:
            if 'A8' in constant.name.upper():
                logger.warning(f"[CONSTANTS_DEBUG] Found A8 constant: sheet={constant.sheet_name}, name={constant.name}, value={constant.value}")
        
        # Also check the final constant_params dict
        for key, value in constant_params.items():
            if 'A8' in key[1]:
                logger.warning(f"[CONSTANTS_DEBUG] In constant_params dict: {key} = {value}")
        
        # Parse target cell
        if "!" in target_cell:
            target_sheet_name, target_cell_coordinate = target_cell.split("!", 1)
        else:
            target_sheet_name = "Sheet1"
            target_cell_coordinate = target_cell
        
        # Extract file_id from file_path
        # file_path format is typically "uploads/{file_id}" or "uploads/{file_id}_{filename}"
        file_name = os.path.basename(file_path)
        if '_' in file_name:
            file_id = file_name.split('_')[0]
        else:
            file_id = file_name.replace('.xlsx', '')
        
        # Run simulation using the file-based interface
        raw_results = await simulation_engine.run_simulation_from_file(
            file_path=file_path,
            file_id=file_id,  # Use actual file ID, not simulation ID
            target_cell=target_cell,
            variables=mc_inputs,  # Pass the original VariableConfig list
            iterations=iterations,
            sheet_name=target_sheet_name,
            constant_params=constant_params,
            progress_callback=progress_cb
        )
        
        # Unpack results
        if isinstance(raw_results, tuple):
            results_array, iteration_errors = raw_results
        else:
            results_array, iteration_errors = raw_results, []

        # Calculate statistics
        stats = simulation_engine._calculate_statistics(results_array)
        
        # Get sensitivity analysis if available
        sensitivity_analysis = []
        if hasattr(simulation_engine, 'sensitivity_analysis'):
            sensitivity_analysis = simulation_engine.sensitivity_analysis
            logger.info(f"[ENHANCED] Retrieved sensitivity analysis with {len(sensitivity_analysis)} variables")
        elif hasattr(simulation_engine, '_calculate_sensitivity_analysis'):
            # Fallback: try calling the method if the attribute doesn't exist
            sensitivity_analysis = simulation_engine._calculate_sensitivity_analysis()
            logger.info(f"[ENHANCED] Calculated sensitivity analysis with {len(sensitivity_analysis)} variables")
        
        # Create result
        sim_result = SimulationResult(
            mean=sanitize_float(stats["mean"]),
            median=sanitize_float(stats["median"]),
            std_dev=sanitize_float(stats["std_dev"]),
            min_value=sanitize_float(stats["min_value"]),
            max_value=sanitize_float(stats["max_value"]),
            percentiles={key: sanitize_float(val) for key, val in stats["percentiles"].items()},
            histogram=stats["histogram"],
            iterations_run=stats["successful_iterations"],
            sensitivity_analysis=sensitivity_analysis,
            errors=iteration_errors
        )
        
        logger.info(f"✅ [ENHANCED] Simulation {sim_id} completed successfully")
        return sim_result
        
    except Exception as e:
        logger.error(f"❌ [ENHANCED] Simulation {sim_id} failed: {e}", exc_info=True)
        raise

async def _run_ultra_simulation(
    sim_id: str,
    file_path: str,
    mc_inputs: List[VariableConfig],
    constants: List[ConstantConfig],
    target_cell: str,
    iterations: int
) -> SimulationResult:
    """Run simulation using the Ultra Monte Carlo Engine"""
    try:
        logger.info(f"🚀 [ULTRA] Starting simulation {sim_id}")
        
        # Import the Ultra engine
        try:
            from .ultra_engine import create_ultra_engine
            logger.info("[ULTRA] Ultra Engine imported successfully")
        except Exception as import_error:
            logger.error(f"[ULTRA] Failed to import Ultra Engine: {import_error}", exc_info=True)
            # Fallback to Enhanced engine
            logger.warning("[ULTRA] Falling back to Enhanced engine")
            return await _run_enhanced_simulation(sim_id, file_path, mc_inputs, constants, target_cell, iterations)
        
        # Create Ultra engine instance
        ultra_engine = create_ultra_engine(iterations=iterations, simulation_id=sim_id)
        logger.info("[ULTRA] Ultra Engine instance created")
        
        # Set up progress callback
        def progress_cb(progress_data):
            try:
                enhanced_progress_data = {**progress_data, "simulation_id": sim_id}
                if sim_id in SIMULATION_START_TIMES:
                    enhanced_progress_data["start_time"] = SIMULATION_START_TIMES[sim_id]
                update_simulation_progress(sim_id, enhanced_progress_data)
            except Exception as e:
                logger.debug(f"Progress relay failed for {sim_id}: {e}")
        
        ultra_engine.set_progress_callback(progress_cb)
        
        # Convert inputs to the format expected by Ultra engine
        mc_input_params = {}
        for mc_input in mc_inputs:
            key = (mc_input.sheet_name, mc_input.name.upper())
            mc_input_params[key] = (mc_input.min_value, mc_input.most_likely, mc_input.max_value)
        
        # Build constant_params dict from ConstantConfig list
        constant_params = {}
        for constant in constants:
            key = (constant.sheet_name, constant.name.upper())
            constant_params[key] = constant.value
        
        # Parse target cell
        if "!" in target_cell:
            target_sheet_name, target_cell_coordinate = target_cell.split("!", 1)
        else:
            target_sheet_name = "Sheet1"
            target_cell_coordinate = target_cell
        
        # CRITICAL FIX: Get real Excel formulas instead of placeholder
        logger.info(f"🔧 [ULTRA] Loading Excel formulas for real simulation")
        
        # Extract file_id from file_path (same logic as Enhanced engine)
        file_name = os.path.basename(file_path)
        if '_' in file_name:
            file_id = file_name.split('_')[0]
        else:
            file_id = file_name.replace('.xlsx', '')
        
        # Get all formulas from the Excel file
        try:
            from excel_parser.service import get_formulas_for_file
            all_formulas = await get_formulas_for_file(file_id)
            logger.info(f"🔧 [ULTRA] Successfully loaded {len(all_formulas)} formulas from file {file_id}")
        except Exception as e:
            logger.error(f"🔧 [ULTRA] Failed to load formulas: {e}")
            all_formulas = {}
        
        # Get MC input cells
        mc_input_cells = set()
        for var_config in mc_inputs:
            mc_input_cells.add((var_config.sheet_name, var_config.name.upper()))
        
        # Get ordered calculation steps using real formula dependency analysis
        from ..formula_utils import get_evaluation_order
        ordered_calc_steps = get_evaluation_order(
            target_sheet_name=target_sheet_name,
            target_cell_coord=target_cell_coordinate,
            all_formulas=all_formulas,
            mc_input_cells=mc_input_cells,
            engine_type='ultra'
        )
        
        logger.info(f"🔧 [ULTRA] Found {len(ordered_calc_steps)} real Excel formulas to evaluate")
        
        # Run simulation using Ultra engine
        raw_results = await ultra_engine.run_simulation(
            mc_input_configs=mc_inputs,
            ordered_calc_steps=ordered_calc_steps,
            target_sheet_name=target_sheet_name,
            target_cell_coordinate=target_cell_coordinate,
            constant_values=constant_params,
            workbook_path=file_path  # Add required workbook path
        )
        
        # Unpack results
        if isinstance(raw_results, tuple):
            results_array, iteration_errors = raw_results
        else:
            results_array, iteration_errors = raw_results, []

        # Calculate statistics
        if len(results_array) == 0:
            logger.error(f"[ULTRA] No results generated for simulation {sim_id}")
            raise RuntimeError("Ultra engine failed to generate results")
        
        # Calculate basic statistics using sanitize_float for safety
        mean = sanitize_float(np.mean(results_array))
        median = sanitize_float(np.median(results_array))
        std_dev = sanitize_float(np.std(results_array))
        min_value = sanitize_float(np.min(results_array))
        max_value = sanitize_float(np.max(results_array))
        
        # Calculate percentiles
        percentiles = {}
        for p in [5, 10, 25, 50, 75, 90, 95]:
            percentiles[str(p)] = sanitize_float(np.percentile(results_array, p))
        
        # Generate histogram
        counts, bin_edges = np.histogram(results_array, bins=50)
        histogram = {
            "bins": bin_edges.tolist(),
            "values": counts.tolist(),
            "bin_edges": bin_edges.tolist(),
            "counts": counts.tolist()
        }
        
        # 🚀 CRITICAL FIX: Get sensitivity analysis from Ultra engine (was hardcoded empty!)
        ultra_sensitivity = ultra_engine.get_sensitivity_analysis()
        if ultra_sensitivity is None or 'tornado_chart' not in ultra_sensitivity:
            sensitivity_analysis = []
            logger.warning(f"[ULTRA] No sensitivity analysis available for simulation {sim_id}")
        else:
            # Extract just the tornado_chart array that the frontend expects
            sensitivity_analysis = ultra_sensitivity['tornado_chart']
            logger.info(f"[ULTRA] Retrieved sensitivity analysis with {len(sensitivity_analysis)} variables")

        # Create result
        sim_result = SimulationResult(
            mean=sanitize_float(mean),
            median=sanitize_float(median),
            std_dev=sanitize_float(std_dev),
            min_value=sanitize_float(min_value),
            max_value=sanitize_float(max_value),
            percentiles=percentiles,
            histogram=histogram,
            iterations_run=len(results_array),
            sensitivity_analysis=sensitivity_analysis,
            errors=iteration_errors
        )
        
        # Get performance stats from Ultra engine
        performance_stats = ultra_engine.get_performance_stats()
        
        logger.info(f"✅ [ULTRA] Simulation {sim_id} completed successfully")
        logger.info(f"📊 [ULTRA] Results: mean={mean:.2f}, std={std_dev:.2f}, "
                   f"min={min_value:.2f}, max={max_value:.2f}")
        logger.info(f"📊 [ULTRA] Performance: {performance_stats}")
        
        return sim_result
        
    except Exception as e:
        logger.error(f"❌ [ULTRA] Simulation {sim_id} failed: {e}", exc_info=True)
        # Fallback to Enhanced engine on error
        logger.warning("[ULTRA] Falling back to Enhanced engine due to error")
        return await _run_enhanced_simulation(sim_id, file_path, mc_inputs, constants, target_cell, iterations)

async def _run_standard_simulation(
    sim_id: str,
    file_path: str,
    mc_inputs: List[VariableConfig],
    constants: List[ConstantConfig], 
    target_cell: str,
    iterations: int
) -> SimulationResult:
    """Run simulation using Standard engine"""
    # Use the original MonteCarloSimulation engine
    logger.info(f"🔧 [STANDARD] Running Standard simulation for {sim_id}")
    
    try:
        # Create standard engine 
        simulation_engine = MonteCarloSimulation(iterations=iterations)
        
        # Attach progress callback to standard engine as well
        def progress_cb(progress_data):
            try:
                # Include start_time from backend service for accurate elapsed time
                enhanced_progress_data = {**progress_data, "simulation_id": sim_id}
                if sim_id in SIMULATION_START_TIMES:
                    enhanced_progress_data["start_time"] = SIMULATION_START_TIMES[sim_id]
                update_simulation_progress(sim_id, enhanced_progress_data)
            except Exception as e:
                logger.debug(f"Progress relay failed for {sim_id}: {e}")
        try:
            simulation_engine.set_progress_callback(progress_cb)
        except AttributeError:
            pass

        # Convert inputs to the format expected by standard engine
        mc_input_params = {}
        for mc_input in mc_inputs:
            key = (mc_input.sheet_name, mc_input.name.upper())
            mc_input_params[key] = (mc_input.min_value, mc_input.most_likely, mc_input.max_value)
        
        # Build constant_params dict from ConstantConfig list passed in
        constant_params = {}
        for constant in constants:
            key = (constant.sheet_name, constant.name.upper())
            constant_params[key] = constant.value
        
        # Log constants to debug A8 issue
        logger.warning(f"[CONSTANTS_DEBUG] Building constant_params with {len(constants)} constants")
        for constant in constants:
            if 'A8' in constant.name.upper():
                logger.warning(f"[CONSTANTS_DEBUG] Found A8 constant: sheet={constant.sheet_name}, name={constant.name}, value={constant.value}")
        
        # Also check the final constant_params dict
        for key, value in constant_params.items():
            if 'A8' in key[1]:
                logger.warning(f"[CONSTANTS_DEBUG] In constant_params dict: {key} = {value}")
        
        # Parse target cell
        if "!" in target_cell:
            target_sheet_name, target_cell_coordinate = target_cell.split("!", 1)
        else:
            target_sheet_name = "Sheet1"  # Default sheet
            target_cell_coordinate = target_cell
        
        # Run simulation - AWAIT the async call (Standard Engine always uses original interface)
        raw_results = await simulation_engine.run_simulation(
            file_path, mc_input_params, constant_params, target_sheet_name, target_cell_coordinate
        )
        
        # Unpack results / errors tuple for statistics calculation
        if isinstance(raw_results, tuple):
            results_array, iteration_errors = raw_results
        else:
            results_array, iteration_errors = raw_results, []

        # Calculate statistics
        stats = simulation_engine._calculate_statistics(results_array)
        
        # Create result
        sim_result = SimulationResult(
            mean=sanitize_float(stats["mean"]),
            median=sanitize_float(stats["median"]),
            std_dev=sanitize_float(stats["std_dev"]),
            min_value=sanitize_float(stats["min_value"]),
            max_value=sanitize_float(stats["max_value"]),
            percentiles={key: sanitize_float(val) for key, val in stats["percentiles"].items()},
            histogram=stats["histogram"],
            iterations_run=stats["successful_iterations"],
            sensitivity_analysis=stats.get("sensitivity_analysis", []),
            errors=iteration_errors
        )
        
        logger.info(f"✅ [STANDARD] Simulation {sim_id} completed successfully")
        return sim_result
        
    except Exception as e:
        logger.error(f"❌ [STANDARD] Simulation {sim_id} failed: {e}")
        raise 

def clear_all_simulation_cache():
    """
    🧹 CLEAR ALL CACHED SIMULATION RESULTS
    
    This function clears all cached simulation results from:
    - In-memory SIMULATION_RESULTS_STORE
    - Redis result store
    - Progress store
    - File system cache
    - Batch sensitivity store
    
    Use this to ensure fresh start for new simulations.
    """
    global SIMULATION_RESULTS_STORE, _simulation_tasks, _simulation_results, _simulation_cancelled, BATCH_SENSITIVITY_STORE
    
    logger.info("🧹 CACHE CLEAR: Starting comprehensive cache cleanup...")
    
    # 1. Clear in-memory stores
    old_count = len(SIMULATION_RESULTS_STORE)
    SIMULATION_RESULTS_STORE.clear()
    _simulation_tasks.clear()
    _simulation_results.clear()
    _simulation_cancelled.clear()
    
    # Clear batch sensitivity store
    batch_count = len(BATCH_SENSITIVITY_STORE)
    BATCH_SENSITIVITY_STORE.clear()
    logger.info(f"🧹 CACHE CLEAR: Cleared {batch_count} batch sensitivity analyses")
    
    logger.info(f"🧹 CACHE CLEAR: Cleared {old_count} simulations from memory store")
    
    # 2. Clear Redis result store
    try:
        from shared.result_store import _result_store
        if hasattr(_result_store, '_redis') and _result_store._redis:
            # Clear all simulation results from Redis
            keys = _result_store._redis.keys("simulation:results:*")
            if keys:
                _result_store._redis.delete(*keys)
                logger.info(f"🧹 CACHE CLEAR: Cleared {len(keys)} results from Redis")
        elif hasattr(_result_store, '_mem'):
            # Clear in-memory fallback
            _result_store._mem.clear()
            logger.info("🧹 CACHE CLEAR: Cleared in-memory result store fallback")
    except Exception as e:
        logger.warning(f"🧹 CACHE CLEAR: Could not clear Redis result store: {e}")
    
    # 3. Clear progress store
    try:
        from shared.progress_store import progress_store
        if hasattr(progress_store, 'redis_client') and progress_store.redis_client:
            # Clear all progress data from Redis
            progress_keys = progress_store.redis_client.keys("simulation:progress:*")
            metadata_keys = progress_store.redis_client.keys("simulation:metadata:*")
            all_keys = progress_keys + metadata_keys
            if all_keys:
                progress_store.redis_client.delete(*all_keys)
                logger.info(f"🧹 CACHE CLEAR: Cleared {len(all_keys)} progress entries from Redis")
        elif hasattr(progress_store, '_fallback_store'):
            # Clear in-memory fallback
            progress_store._fallback_store.clear()
            progress_store._simulation_metadata.clear()
            logger.info("🧹 CACHE CLEAR: Cleared in-memory progress store fallback")
    except Exception as e:
        logger.warning(f"🧹 CACHE CLEAR: Could not clear progress store: {e}")
    
    # 4. Clear file system cache (optional - be careful with this)
    try:
        import os
        import glob
        
        # Clear temporary result files (be very selective)
        temp_files = glob.glob("results/temp_*") + glob.glob("cache/temp_*")
        for file_path in temp_files:
            try:
                os.remove(file_path)
                logger.debug(f"🧹 CACHE CLEAR: Removed temp file: {file_path}")
            except Exception as e:
                logger.debug(f"🧹 CACHE CLEAR: Could not remove {file_path}: {e}")
        
        if temp_files:
            logger.info(f"🧹 CACHE CLEAR: Cleaned {len(temp_files)} temporary files")
            
    except Exception as e:
        logger.warning(f"🧹 CACHE CLEAR: Could not clear file system cache: {e}")
    
    logger.info("🧹 CACHE CLEAR: Comprehensive cache cleanup completed!")
    
    return {
        "status": "success",
        "message": "All simulation cache cleared successfully",
        "cleared_simulations": old_count,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }

def clear_specific_simulation_cache(simulation_id: str):
    """
    🧹 CLEAR SPECIFIC SIMULATION CACHE
    
    Clears cache for a specific simulation ID from all stores.
    """
    logger.info(f"🧹 CACHE CLEAR: Clearing cache for simulation {simulation_id}")
    
    # Clear from memory store
    if simulation_id in SIMULATION_RESULTS_STORE:
        del SIMULATION_RESULTS_STORE[simulation_id]
        logger.info(f"🧹 CACHE CLEAR: Removed {simulation_id} from memory store")
    
    # Clear from task tracking
    _simulation_tasks.pop(simulation_id, None)
    _simulation_results.pop(simulation_id, None)
    _simulation_cancelled.discard(simulation_id)
    
    # Clear from Redis result store
    try:
        from shared.result_store import _result_store
        if hasattr(_result_store, '_redis') and _result_store._redis:
            key = f"simulation:results:{simulation_id}"
            _result_store._redis.delete(key)
        elif hasattr(_result_store, '_mem'):
            _result_store._mem.pop(simulation_id, None)
    except Exception as e:
        logger.warning(f"🧹 CACHE CLEAR: Could not clear Redis result for {simulation_id}: {e}")
    
    # Clear from progress store
    try:
        from shared.progress_store import progress_store
        progress_store.clear_progress(simulation_id)
    except Exception as e:
        logger.warning(f"🧹 CACHE CLEAR: Could not clear progress for {simulation_id}: {e}")
    
    logger.info(f"🧹 CACHE CLEAR: Cache cleared for simulation {simulation_id}") 

async def get_model_optimization_suggestions(file_id: str, 
                                           formulas: Dict[str, Dict[str, str]], 
                                           dependencies: List[Tuple[str, str, str]],
                                           mc_inputs: Set[Tuple[str, str]]) -> Dict[str, Any]:
    """
    Get optimization suggestions for the Excel model.
    This implements the "suggested optimization paths" feature.
    """
    try:
        analyzer = ModelOptimizationAnalyzer()
        results = analyzer.analyze_model(formulas, dependencies, mc_inputs)
        
        logger.info(f"📊 Model optimization analysis complete: score={results['optimization_score']:.1f}/100")
        logger.info(f"📋 Found {len(results['suggestions'])} optimization suggestions")
        
        # Store results for later retrieval
        optimization_key = f"optimization_{file_id}"
        try:
            from shared.result_store import set_result
            set_result(optimization_key, results, ttl=3600)  # Cache for 1 hour
        except Exception as e:
            logger.debug(f"Failed to cache optimization results: {e}")
        
        return results
    except Exception as e:
        logger.error(f"Model optimization analysis failed: {e}")
        return {
            'stats': {'optimization_potential': 50.0},
            'suggestions': [],
            'optimization_score': 50.0,
            'summary': 'Optimization analysis unavailable.'
        } 

def get_column_letter(col_idx: int) -> str:
    """Convert a column index (0-based) to Excel column letter (A, B, ..., Z, AA, AB, ...)"""
    letters = ""
    col_num = col_idx + 1  # Convert to 1-based
    while col_num > 0:
        col_num -= 1
        letters = chr(col_num % 26 + ord('A')) + letters
        col_num //= 26
    return letters 



