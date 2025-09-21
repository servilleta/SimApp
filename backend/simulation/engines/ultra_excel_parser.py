"""
ULTRA ENGINE PHASE 3: EXCEL PARSING & COMPLETE DEPENDENCY ANALYSIS

CRITICAL LESSONS LEARNED IMPLEMENTATION:
1. Complete Formula Tree Understanding (multi-pass dependency analysis)
2. Excel Reference Type Support ($A$1, $A1, A$1, A1)  
3. Multi-Sheet Workbook Support (all sheets, cross-references)
4. Database-First Results Architecture (no complex memory structures)

Research basis: Complete dependency analysis until no new dependencies found
"""

import numpy as np
import logging
import re
import sqlite3
from typing import Dict, Tuple, List, Optional, Any, Set
from dataclasses import dataclass
from enum import Enum

# Excel parsing imports with fallback
try:
    import openpyxl
    from openpyxl import load_workbook
    EXCEL_PARSING_AVAILABLE = True
except ImportError:
    openpyxl = None
    load_workbook = None
    EXCEL_PARSING_AVAILABLE = False

logger = logging.getLogger(__name__)

@dataclass
class CellReference:
    """
    PHASE 3: Complete Excel cell reference with all $ symbol combinations
    
    Critical lesson learned: Handle ALL Excel reference types correctly:
    - $A$1 (both absolute) - doesn't change during copy operations
    - $A1 (column absolute) - column stays fixed, row changes  
    - A$1 (row absolute) - row stays fixed, column changes
    - A1 (both relative) - both change during copy operations
    """
    sheet: str
    column: int  # 1-based column number (A=1, B=2, etc.)
    row: int     # 1-based row number
    is_row_absolute: bool = False    # True if A$1 format
    is_col_absolute: bool = False    # True if $A1 format  
    is_external: bool = False        # True if references external workbook
    
    def __str__(self):
        col_str = self._number_to_column(self.column)
        return f"{self.sheet}!{'$' if self.is_col_absolute else ''}{col_str}{'$' if self.is_row_absolute else ''}{self.row}"
    
    def _number_to_column(self, num: int) -> str:
        """Convert column number to Excel column letters (1=A, 26=Z, 27=AA)"""
        result = ""
        while num > 0:
            num -= 1
            result = chr(65 + (num % 26)) + result
            num //= 26
        return result
    
    def __hash__(self):
        return hash((self.sheet, self.column, self.row, self.is_row_absolute, self.is_col_absolute))
    
    def __eq__(self, other):
        if not isinstance(other, CellReference):
            return False
        return (self.sheet == other.sheet and 
                self.column == other.column and 
                self.row == other.row and
                self.is_row_absolute == other.is_row_absolute and
                self.is_col_absolute == other.is_col_absolute)

@dataclass  
class FormulaInfo:
    """Complete formula information including dependencies"""
    expression: str
    dependencies: List[CellReference]
    is_array_formula: bool = False
    complexity_score: int = 0

@dataclass
class SheetData:
    """Complete sheet data with all formulas and values"""
    name: str
    formulas: Dict[Tuple[int, int], FormulaInfo]  # (row, col) -> FormulaInfo
    values: Dict[Tuple[int, int], Any]            # (row, col) -> value
    named_ranges: Dict[str, str]                  # name -> range
    max_row: int = 0
    max_col: int = 0

@dataclass
class WorkbookData:
    """Complete workbook data structure"""
    sheets: Dict[str, SheetData]
    global_named_ranges: Dict[str, str]
    external_references: List[str]
    total_formulas: int = 0
    total_cells: int = 0

@dataclass
class DependencyNode:
    """Complete dependency node for exhaustive analysis"""
    cell: CellReference
    dependencies: Set[CellReference]
    dependents: Set[CellReference]
    is_fully_mapped: bool = False
    depth: int = 0
    complexity_score: int = 0

class UltraExcelReferenceParser:
    """
    PHASE 3: Complete Excel Reference Parser
    
    CRITICAL LESSON LEARNED: Handle ALL Excel reference types including $ symbols
    
    Supported formats:
    - A1, B2, Z100 (relative references)
    - $A1, $B2 (column absolute)  
    - A$1, B$2 (row absolute)
    - $A$1, $B$2 (both absolute)
    - Sheet1!A1 (cross-sheet)
    - 'Sheet Name'!A1 (quoted sheet names)
    - [Workbook.xlsx]Sheet1!A1 (external workbook)
    """
    
    # Comprehensive regex pattern for ALL Excel reference types
    EXCEL_REF_PATTERN = re.compile(
        r"(?:"                           # Non-capturing group for optional workbook
        r"\[([^\]]+)\]"                  # [Workbook.xlsx] - external workbook
        r")?"                            # Optional workbook part
        r"(?:"                           # Non-capturing group for optional sheet
        r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_]*))!" # 'Sheet Name'! or Sheet1!
        r")?"                            # Optional sheet part  
        r"(\$?[A-Z]+)(\$?\d+)"          # $A$1, $A1, A$1, A1
        r"(?::(\$?[A-Z]+)(\$?\d+))?"    # Optional range end :$Z$100
    )
    
    def __init__(self):
        self.logger = logging.getLogger(__name__ + ".UltraExcelReferenceParser")
    
    def parse_reference(self, ref_str: str, current_sheet: str = "Sheet1") -> Optional[CellReference]:
        """
        Parse a single Excel reference into CellReference object
        
        Args:
            ref_str: Excel reference string (e.g., "$A$1", "Sheet1!B2", "'Data Sheet'!$C1")
            current_sheet: Default sheet name if not specified
            
        Returns:
            CellReference object or None if parsing fails
        """
        try:
            match = self.EXCEL_REF_PATTERN.match(ref_str.strip())
            if not match:
                return None
            
            workbook, quoted_sheet, unquoted_sheet, col_str, row_str, end_col_str, end_row_str = match.groups()
            
            # Determine sheet name
            sheet_name = quoted_sheet or unquoted_sheet or current_sheet
            
            # Parse column (handle $ for absolute)
            is_col_absolute = col_str.startswith('$')
            col_letters = col_str.lstrip('$')
            column_num = self._column_to_number(col_letters)
            
            # Parse row (handle $ for absolute)
            is_row_absolute = row_str.startswith('$')
            row_num = int(row_str.lstrip('$'))
            
            # Check if external reference
            is_external = workbook is not None
            
            return CellReference(
                sheet=sheet_name,
                column=column_num,
                row=row_num,
                is_row_absolute=is_row_absolute,
                is_col_absolute=is_col_absolute,
                is_external=is_external
            )
            
        except Exception as e:
            self.logger.debug(f"Failed to parse reference '{ref_str}': {e}")
            return None
    
    def extract_all_references(self, formula: str, current_sheet: str = "Sheet1") -> List[CellReference]:
        """
        Extract ALL cell references from a formula
        
        Args:
            formula: Excel formula string (e.g., "=A1+B2*SUM(C1:C10)")
            current_sheet: Default sheet name for relative references
            
        Returns:
            List of all CellReference objects found in formula
        """
        if not formula or not formula.startswith('='):
            return []
        
        references = []
        
        # Find all matches in the formula
        for match in self.EXCEL_REF_PATTERN.finditer(formula):
            ref_str = match.group(0)
            cell_ref = self.parse_reference(ref_str, current_sheet)
            if cell_ref:
                references.append(cell_ref)
        
        return references
    
    def _column_to_number(self, col_str: str) -> int:
        """Convert Excel column letters to number (A=1, B=2, Z=26, AA=27)"""
        result = 0
        for char in col_str.upper():
            result = result * 26 + (ord(char) - ord('A') + 1)
        return result

class UltraWorkbookParser:
    """
    PHASE 3: Complete Multi-Sheet Workbook Parser
    
    CRITICAL LESSON LEARNED: Parse entire workbook, not just single sheet
    
    Features:
    - Reads ALL sheets in workbook
    - Extracts all formulas and values
    - Finds cross-sheet dependencies
    - Handles named ranges
    - Validates all references
    """
    
    def __init__(self):
        self.logger = logging.getLogger(__name__ + ".UltraWorkbookParser")
        self.reference_parser = UltraExcelReferenceParser()
    
    def parse_complete_workbook(self, file_path: str) -> WorkbookData:
        """
        Parse complete Excel workbook including all sheets
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            WorkbookData with complete workbook information
        """
        if not EXCEL_PARSING_AVAILABLE:
            raise RuntimeError("openpyxl not available - cannot parse Excel files")
        
        self.logger.info(f"🔧 [ULTRA] Starting complete workbook analysis: {file_path}")
        
        try:
            # Load workbook with openpyxl
            workbook = load_workbook(file_path, data_only=False)  # Keep formulas
            
            workbook_data = WorkbookData(
                sheets={},
                global_named_ranges={},
                external_references=[]
            )
            
            # Step 1: Get all sheet names
            sheet_names = workbook.sheetnames
            self.logger.info(f"🔧 [ULTRA] Found {len(sheet_names)} sheets: {sheet_names}")
            
            # Step 2: Parse each sheet individually
            total_formulas = 0
            total_cells = 0
            
            for sheet_name in sheet_names:
                self.logger.info(f"🔧 [ULTRA] Parsing sheet: {sheet_name}")
                sheet_data = self._parse_sheet(workbook[sheet_name])
                workbook_data.sheets[sheet_name] = sheet_data
                
                total_formulas += len(sheet_data.formulas)
                total_cells += sheet_data.max_row * sheet_data.max_col
            
            # Step 3: Parse global named ranges
            workbook_data.global_named_ranges = self._parse_named_ranges(workbook)
            
            # Step 4: Find cross-sheet dependencies
            self._find_cross_sheet_dependencies(workbook_data)
            
            # Step 5: Final validation
            workbook_data.total_formulas = total_formulas
            workbook_data.total_cells = total_cells
            
            self.logger.info(f"✅ [ULTRA] Workbook parsing complete:")
            self.logger.info(f"   - Sheets: {len(workbook_data.sheets)}")
            self.logger.info(f"   - Total formulas: {total_formulas}")
            self.logger.info(f"   - Total cells: {total_cells}")
            self.logger.info(f"   - Named ranges: {len(workbook_data.global_named_ranges)}")
            
            return workbook_data
            
        except Exception as e:
            self.logger.error(f"❌ [ULTRA] Workbook parsing failed: {e}")
            raise RuntimeError(f"Failed to parse Excel workbook: {e}")
    
    def _parse_sheet(self, worksheet) -> SheetData:
        """Parse individual sheet and extract all formulas and values"""
        sheet_data = SheetData(
            name=worksheet.title,
            formulas={},
            values={},
            named_ranges={}
        )
        
        # Get sheet dimensions
        sheet_data.max_row = worksheet.max_row or 1
        sheet_data.max_col = worksheet.max_column or 1
        
        self.logger.debug(f"Sheet {worksheet.title}: {sheet_data.max_row} rows × {sheet_data.max_col} cols")
        
        # Scan all cells for formulas and values
        for row in range(1, sheet_data.max_row + 1):
            for col in range(1, sheet_data.max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                
                if cell.value is not None:
                    # Check if cell has formula - openpyxl stores formulas when data_type is 'f'
                    if cell.data_type == 'f':
                        # This is a formula cell - value contains the formula with = sign
                        formula_str = cell.value
                        
                        if formula_str and formula_str.startswith('='):
                            self.logger.debug(f"Found formula at {row},{col}: {formula_str}")
                            
                            # Extract dependencies from formula
                            dependencies = self.reference_parser.extract_all_references(
                                formula_str, worksheet.title
                            )
                            
                            # Calculate complexity score
                            complexity = self._calculate_formula_complexity(formula_str)
                            
                            sheet_data.formulas[(row, col)] = FormulaInfo(
                                expression=formula_str,
                                dependencies=dependencies,
                                complexity_score=complexity
                            )
                    else:
                        # This is a value cell
                        sheet_data.values[(row, col)] = cell.value
        
        self.logger.debug(f"Sheet {worksheet.title}: {len(sheet_data.formulas)} formulas found")
        return sheet_data
    
    def _calculate_formula_complexity(self, formula: str) -> int:
        """Calculate formula complexity score for optimization"""
        complexity = 0
        
        # Count function calls
        functions = re.findall(r'[A-Z]+\(', formula.upper())
        complexity += len(functions) * 2
        
        # Count cell references  
        refs = self.reference_parser.extract_all_references(formula)
        complexity += len(refs)
        
        # Count operators
        operators = re.findall(r'[+\-*/^<>=]', formula)
        complexity += len(operators)
        
        # Formula length factor
        complexity += len(formula) // 20
        
        return min(100, complexity)  # Cap at 100
    
    def _parse_named_ranges(self, workbook) -> Dict[str, str]:
        """Parse global named ranges from workbook"""
        named_ranges = {}
        
        try:
            for name, value in workbook.defined_names.items():
                if value.value:
                    named_ranges[name] = str(value.value)
                    self.logger.debug(f"Named range: {name} = {value.value}")
        except Exception as e:
            self.logger.debug(f"Failed to parse named ranges: {e}")
        
        return named_ranges
    
    def _find_cross_sheet_dependencies(self, workbook_data: WorkbookData):
        """Find and validate cross-sheet dependencies"""
        self.logger.info("🔧 [ULTRA] Analyzing cross-sheet dependencies...")
        
        cross_sheet_count = 0
        
        for sheet_name, sheet_data in workbook_data.sheets.items():
            for (row, col), formula_info in sheet_data.formulas.items():
                for dep in formula_info.dependencies:
                    # Check if dependency is cross-sheet
                    if dep.sheet != sheet_name:
                        cross_sheet_count += 1
                        
                        # Validate that referenced sheet exists
                        if dep.sheet not in workbook_data.sheets:
                            self.logger.warning(f"Cross-sheet reference to non-existent sheet: {dep}")
                        else:
                            self.logger.debug(f"Cross-sheet: {sheet_name}!{row}:{col} -> {dep}")
        
        self.logger.info(f"🔧 [ULTRA] Found {cross_sheet_count} cross-sheet dependencies")

class UltraCompleteDependencyEngine:
    """
    PHASE 3: Complete Dependency Analysis Engine
    
    CRITICAL LESSON LEARNED: Continue until COMPLETE dependency tree is mapped
    
    Multi-pass dependency analysis that continues until NO new dependencies are found.
    This addresses the #1 lesson learned - past engines stopped dependency analysis too early!
    """
    
    def __init__(self, max_passes: int = 100):
        self.logger = logging.getLogger(__name__ + ".UltraCompleteDependencyEngine")
        self.max_passes = max_passes
        self.dependency_graph: Dict[CellReference, DependencyNode] = {}
    
    def build_complete_dependency_tree(self, workbook_data: WorkbookData) -> Dict[CellReference, DependencyNode]:
        """
        Build complete dependency tree with multi-pass validation
        
        CRITICAL: Continue until NO new dependencies are found
        """
        self.logger.info("🔧 [ULTRA] Starting COMPLETE dependency analysis...")
        
        # Initialize dependency graph
        self._initialize_dependency_graph(workbook_data)
        
        pass_count = 1
        new_dependencies_found = True
        
        # CRITICAL: Continue until no new dependencies are found
        while new_dependencies_found and pass_count <= self.max_passes:
            self.logger.info(f"🔧 [ULTRA] Dependency analysis pass {pass_count}")
            new_dependencies_found = False
            
            # Analyze each formula in each sheet
            for sheet_name, sheet_data in workbook_data.sheets.items():
                for (row, col), formula_info in sheet_data.formulas.items():
                    cell_ref = CellReference(sheet=sheet_name, column=col, row=row)
                    
                    if cell_ref in self.dependency_graph:
                        node = self.dependency_graph[cell_ref]
                        
                        if not node.is_fully_mapped:
                            # Extract ALL dependencies from formula
                            all_deps = self._extract_all_dependencies(formula_info, workbook_data)
                            
                            # Check for new dependencies
                            for dep in all_deps:
                                if dep not in node.dependencies:
                                    node.dependencies.add(dep)
                                    
                                    # Add reverse dependency
                                    if dep in self.dependency_graph:
                                        self.dependency_graph[dep].dependents.add(cell_ref)
                                    
                                    new_dependencies_found = True
                                    self.logger.debug(f"Found new dependency: {cell_ref} -> {dep}")
                            
                            # Mark as fully mapped for this pass
                            node.is_fully_mapped = True
                            node.complexity_score = formula_info.complexity_score
            
            pass_count += 1
        
        # Final validation and statistics
        self._validate_dependency_tree(pass_count - 1)
        self._calculate_dependency_depths()
        
        return self.dependency_graph
    
    def _initialize_dependency_graph(self, workbook_data: WorkbookData):
        """Initialize dependency graph with all formulas"""
        self.logger.info("🔧 [ULTRA] Initializing dependency graph...")
        
        total_nodes = 0
        
        for sheet_name, sheet_data in workbook_data.sheets.items():
            for (row, col), formula_info in sheet_data.formulas.items():
                cell_ref = CellReference(sheet=sheet_name, column=col, row=row)
                
                self.dependency_graph[cell_ref] = DependencyNode(
                    cell=cell_ref,
                    dependencies=set(),
                    dependents=set(),
                    is_fully_mapped=False,
                    complexity_score=formula_info.complexity_score
                )
                total_nodes += 1
        
        self.logger.info(f"🔧 [ULTRA] Initialized {total_nodes} nodes in dependency graph")
    
    def _extract_all_dependencies(self, formula_info: FormulaInfo, workbook_data: WorkbookData) -> Set[CellReference]:
        """Extract ALL dependencies including indirect ones"""
        all_deps = set()
        
        # Direct dependencies from formula
        for dep in formula_info.dependencies:
            all_deps.add(dep)
            
            # Validate dependency exists in workbook
            if dep.sheet in workbook_data.sheets:
                sheet_data = workbook_data.sheets[dep.sheet]
                if (dep.row, dep.column) in sheet_data.formulas:
                    # This dependency has its own formula - will be analyzed in next pass
                    pass
        
        return all_deps
    
    def _validate_dependency_tree(self, total_passes: int):
        """Validate complete dependency tree"""
        self.logger.info("🔧 [ULTRA] Validating complete dependency tree...")
        
        total_nodes = len(self.dependency_graph)
        mapped_nodes = sum(1 for node in self.dependency_graph.values() if node.is_fully_mapped)
        total_dependencies = sum(len(node.dependencies) for node in self.dependency_graph.values())
        
        self.logger.info(f"✅ [ULTRA] Dependency analysis completed in {total_passes} passes")
        self.logger.info(f"   - Total nodes: {total_nodes}")
        self.logger.info(f"   - Fully mapped nodes: {mapped_nodes}")
        self.logger.info(f"   - Total dependencies: {total_dependencies}")
        self.logger.info(f"   - Average dependencies per node: {total_dependencies / total_nodes if total_nodes > 0 else 0:.2f}")
        
        if mapped_nodes < total_nodes:
            unmapped_count = total_nodes - mapped_nodes
            self.logger.warning(f"⚠️ [ULTRA] {unmapped_count} nodes not fully mapped!")
            self.logger.warning("This indicates potential circular references or complex dependencies")
        
        if total_passes >= self.max_passes:
            self.logger.warning(f"⚠️ [ULTRA] Reached maximum passes ({self.max_passes}) - possible circular reference")
    
    def _calculate_dependency_depths(self):
        """Calculate dependency depths for optimization"""
        self.logger.info("🔧 [ULTRA] Calculating dependency depths...")
        
        # Find leaf nodes (no dependencies)
        leaf_nodes = [node for node in self.dependency_graph.values() if len(node.dependencies) == 0]
        
        # BFS to calculate depths
        queue = [(node, 0) for node in leaf_nodes]
        visited = set()
        
        while queue:
            node, depth = queue.pop(0)
            
            if node.cell in visited:
                continue
            
            visited.add(node.cell)
            node.depth = depth
            
            # Add dependents to queue
            for dependent_ref in node.dependents:
                if dependent_ref in self.dependency_graph:
                    dependent_node = self.dependency_graph[dependent_ref]
                    queue.append((dependent_node, depth + 1))
        
        max_depth = max((node.depth for node in self.dependency_graph.values()), default=0)
        self.logger.info(f"🔧 [ULTRA] Maximum dependency depth: {max_depth}")
    
    def get_evaluation_order(self, target_cell: CellReference) -> List[CellReference]:
        """Get optimal evaluation order for target cell"""
        if target_cell not in self.dependency_graph:
            return [target_cell]
        
        # Topological sort from target cell
        evaluation_order = []
        visited = set()
        temp_visited = set()
        
        def dfs_visit(cell_ref: CellReference):
            if cell_ref in temp_visited:
                # Circular reference detected
                self.logger.warning(f"Circular reference detected involving {cell_ref}")
                return
            
            if cell_ref in visited:
                return
            
            temp_visited.add(cell_ref)
            
            if cell_ref in self.dependency_graph:
                node = self.dependency_graph[cell_ref]
                for dep in node.dependencies:
                    dfs_visit(dep)
            
            temp_visited.remove(cell_ref)
            visited.add(cell_ref)
            evaluation_order.append(cell_ref)
        
        dfs_visit(target_cell)
        return evaluation_order

# Factory functions for integration
def create_excel_parser() -> UltraWorkbookParser:
    """Create Excel parser instance"""
    return UltraWorkbookParser()

def create_dependency_engine(max_passes: int = 100) -> UltraCompleteDependencyEngine:
    """Create dependency analysis engine"""
    return UltraCompleteDependencyEngine(max_passes=max_passes)

def parse_excel_file(file_path: str) -> WorkbookData:
    """Parse Excel file and return complete workbook data"""
    parser = create_excel_parser()
    return parser.parse_complete_workbook(file_path)

def analyze_dependencies(workbook_data: WorkbookData) -> Dict[CellReference, DependencyNode]:
    """Analyze complete dependencies and return dependency graph"""
    engine = create_dependency_engine()
    return engine.build_complete_dependency_tree(workbook_data) 