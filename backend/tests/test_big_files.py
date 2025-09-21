"""
Big File Smoke Tests

Tests the system's ability to handle large Excel files with various sizes and complexities.
This validates our current limits and identifies bottlenecks.
"""

import pytest
import asyncio
import time
import psutil
import tempfile
import os
from pathlib import Path
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from fastapi.testclient import TestClient
from fastapi import UploadFile

from main import app
from config import settings
from shared.upload_middleware import upload_validator
from shared.file_cleanup import file_cleanup_service
from excel_parser.service import parse_excel_file
from simulation.enhanced_engine import WorldClassMonteCarloEngine

# Test client
client = TestClient(app)


class BigFileTestSuite:
    """Test suite for big file handling capabilities."""
    
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_files = {}
        self.memory_baseline = None
        
    def setup(self):
        """Set up test environment."""
        self.memory_baseline = psutil.Process().memory_info().rss / 1024 / 1024  # MB
        print(f"🔍 Memory baseline: {self.memory_baseline:.1f} MB")
        
    def teardown(self):
        """Clean up test files."""
        for file_path in self.test_files.values():
            if os.path.exists(file_path):
                os.remove(file_path)
        os.rmdir(self.temp_dir)
        
    def create_test_excel_file(self, size_category: str) -> str:
        """Create test Excel files of different sizes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"
        
        if size_category == "small":
            # ~5MB file with 1,000 formulas
            rows, cols = 50, 100
            formula_count = 1000
            filename = "small_test_5mb.xlsx"
        elif size_category == "medium":
            # ~25MB file with 10,000 formulas  
            rows, cols = 100, 200
            formula_count = 10000
            filename = "medium_test_25mb.xlsx"
        elif size_category == "large":
            # ~100MB file with 50,000 formulas
            rows, cols = 200, 500
            formula_count = 50000
            filename = "large_test_100mb.xlsx"
        elif size_category == "huge":
            # ~300MB file with 200,000 formulas
            rows, cols = 500, 1000
            formula_count = 200000
            filename = "huge_test_300mb.xlsx"
        else:
            raise ValueError(f"Unknown size category: {size_category}")
            
        print(f"🏗️ Creating {size_category} file: {rows}x{cols} = {rows*cols} cells, {formula_count} formulas")
        
        # Fill with data and formulas
        formula_count_actual = 0
        
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                
                # Mix of data types
                if (row + col) % 7 == 0 and formula_count_actual < formula_count:
                    # Formula cells
                    if col > 1:
                        cell.value = f"=A{row}*RAND()+{col}"
                    else:
                        cell.value = f"=RAND()*100"
                    formula_count_actual += 1
                elif (row + col) % 3 == 0:
                    # Random numbers
                    cell.value = (row * col * 0.1) % 1000
                else:
                    # Text/number mix
                    cell.value = f"Data_{row}_{col}"
        
        # Add extra formulas if needed
        extra_row = rows + 1
        for i in range(formula_count_actual, formula_count):
            col = (i % cols) + 1
            if col > cols:
                extra_row += 1
                col = 1
            cell = ws.cell(row=extra_row, column=col)
            cell.value = f"=RAND()*{i+1}"
            formula_count_actual += 1
            
        file_path = os.path.join(self.temp_dir, filename)
        wb.save(file_path)
        
        # Get actual file size
        file_size = os.path.getsize(file_path)
        file_size_mb = file_size / (1024 * 1024)
        
        print(f"✅ Created {filename}: {file_size_mb:.1f} MB, {formula_count_actual} formulas")
        
        self.test_files[size_category] = file_path
        return file_path
        
    def get_memory_usage(self) -> float:
        """Get current memory usage in MB."""
        return psutil.Process().memory_info().rss / 1024 / 1024
        
    def test_upload_validation(self, file_path: str) -> dict:
        """Test file upload validation."""
        print(f"\n📤 Testing upload validation for {os.path.basename(file_path)}")
        
        start_time = time.time()
        start_memory = self.get_memory_usage()
        
        try:
            # Create UploadFile-like object
            with open(file_path, "rb") as f:
                file_content = f.read()
                
            # Create mock UploadFile
            class MockUploadFile:
                def __init__(self, filename, content, content_type):
                    self.filename = filename
                    self.content_type = content_type
                    self.file = BytesIO(content)
                    
                def read(self):
                    return self.file.read()
                    
                def seek(self, pos):
                    return self.file.seek(pos)
            
            mock_file = MockUploadFile(
                filename=os.path.basename(file_path),
                content=file_content,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Test validation
            result = upload_validator.validate_file(mock_file)
            
            end_time = time.time()
            end_memory = self.get_memory_usage()
            
            validation_stats = {
                "success": True,
                "file_size_mb": result["size_mb"],
                "validation_time": end_time - start_time,
                "memory_used_mb": end_memory - start_memory,
                "result": result
            }
            
            print(f"✅ Validation passed: {result['size_mb']:.1f} MB in {validation_stats['validation_time']:.3f}s")
            return validation_stats
            
        except Exception as e:
            print(f"❌ Validation failed: {e}")
            return {
                "success": False,
                "error": str(e),
                "validation_time": time.time() - start_time,
                "memory_used_mb": self.get_memory_usage() - start_memory
            }
    
    def test_excel_parsing(self, file_path: str) -> dict:
        """Test Excel file parsing."""
        print(f"\n🔍 Testing Excel parsing for {os.path.basename(file_path)}")
        
        start_time = time.time()
        start_memory = self.get_memory_usage()
        
        try:
            # Create UploadFile for parsing
            with open(file_path, "rb") as f:
                file_content = f.read()
                
            # Create proper UploadFile
            upload_file = UploadFile(
                filename=os.path.basename(file_path),
                file=BytesIO(file_content),
                size=len(file_content)
            )
            
            # Parse file
            result = asyncio.run(parse_excel_file(upload_file))
            
            end_time = time.time()
            end_memory = self.get_memory_usage()
            
            parsing_stats = {
                "success": True,
                "parsing_time": end_time - start_time,
                "memory_used_mb": end_memory - start_memory,
                "sheet_count": len(result.sheets),
                "file_size_mb": result.file_size / (1024 * 1024)
            }
            
            # Calculate total cells and formulas
            total_cells = 0
            formula_cells = 0
            
            for sheet in result.sheets:
                sheet_rows = len(sheet.grid_data)
                if sheet_rows > 0:
                    sheet_cols = len(sheet.grid_data[0])
                    total_cells += sheet_rows * sheet_cols
                    
                    # Count formulas
                    for row in sheet.grid_data:
                        for cell in row:
                            if cell and hasattr(cell, 'formula') and cell.formula:
                                formula_cells += 1
            
            parsing_stats["total_cells"] = total_cells
            parsing_stats["formula_cells"] = formula_cells
            parsing_stats["formula_density"] = formula_cells / max(1, total_cells)
            
            print(f"✅ Parsing completed: {total_cells} cells, {formula_cells} formulas in {parsing_stats['parsing_time']:.3f}s")
            print(f"   Memory used: {parsing_stats['memory_used_mb']:.1f} MB")
            
            return parsing_stats
            
        except Exception as e:
            print(f"❌ Parsing failed: {e}")
            return {
                "success": False,
                "error": str(e),
                "parsing_time": time.time() - start_time,
                "memory_used_mb": self.get_memory_usage() - start_memory
            }
    
    def test_file_complexity_analysis(self, file_size_bytes: int) -> dict:
        """Test the file complexity analysis system."""
        print(f"\n🧮 Testing complexity analysis for {file_size_bytes / (1024*1024):.1f} MB file")
        
        start_time = time.time()
        
        try:
            # Test via HTTP endpoint
            response = client.get(f"/api/bigfiles/analysis/{file_size_bytes}")
            
            if response.status_code == 200:
                analysis = response.json()
                
                analysis_stats = {
                    "success": True,
                    "analysis_time": time.time() - start_time,
                    "category": analysis["file_analysis"]["category"],
                    "complexity_score": analysis["file_analysis"]["complexity_score"],
                    "processing_mode": analysis["file_analysis"]["processing_mode"],
                    "estimated_time": analysis["performance_estimate"]["estimated_time"],
                    "memory_usage": analysis["performance_estimate"]["memory_usage"],
                    "recommendations": analysis["recommendations"]
                }
                
                print(f"✅ Analysis completed: {analysis_stats['category']} file, {analysis_stats['complexity_score']:.1f}% complexity")
                print(f"   Processing mode: {analysis_stats['processing_mode']}")
                print(f"   Estimated time: {analysis_stats['estimated_time']}")
                
                return analysis_stats
            else:
                print(f"❌ Analysis failed: HTTP {response.status_code}")
                return {
                    "success": False,
                    "error": f"HTTP {response.status_code}",
                    "analysis_time": time.time() - start_time
                }
                
        except Exception as e:
            print(f"❌ Analysis failed: {e}")
            return {
                "success": False,
                "error": str(e),
                "analysis_time": time.time() - start_time
            }
    
    def test_memory_cleanup(self) -> dict:
        """Test file cleanup functionality."""
        print(f"\n🧹 Testing file cleanup system")
        
        start_time = time.time()
        
        try:
            # Get initial disk usage
            initial_usage = file_cleanup_service.get_disk_usage()
            
            # Run cleanup
            cleanup_result = file_cleanup_service.cleanup_old_files()
            
            # Get final disk usage  
            final_usage = file_cleanup_service.get_disk_usage()
            
            cleanup_stats = {
                "success": True,
                "cleanup_time": time.time() - start_time,
                "files_deleted": cleanup_result.get("upload_files_deleted", 0) + cleanup_result.get("temp_files_deleted", 0),
                "bytes_freed": cleanup_result.get("total_bytes_freed", 0),
                "initial_usage": initial_usage,
                "final_usage": final_usage,
                "cleanup_result": cleanup_result
            }
            
            print(f"✅ Cleanup completed: {cleanup_stats['files_deleted']} files deleted, {cleanup_stats['bytes_freed'] / (1024*1024):.1f} MB freed")
            
            return cleanup_stats
            
        except Exception as e:
            print(f"❌ Cleanup failed: {e}")
            return {
                "success": False,
                "error": str(e),
                "cleanup_time": time.time() - start_time
            }
    
    def run_comprehensive_test(self, size_categories: list = None) -> dict:
        """Run comprehensive big file smoke test."""
        if size_categories is None:
            size_categories = ["small", "medium"]  # Safe default
            
        print("🚀 STARTING BIG FILE SMOKE TEST")
        print("=" * 60)
        
        self.setup()
        
        results = {
            "test_start_time": time.time(),
            "memory_baseline_mb": self.memory_baseline,
            "categories_tested": size_categories,
            "results": {},
            "summary": {}
        }
        
        try:
            for category in size_categories:
                print(f"\n🎯 TESTING CATEGORY: {category.upper()}")
                print("-" * 40)
                
                category_results = {}
                
                # 1. Create test file
                file_path = self.create_test_excel_file(category)
                file_size = os.path.getsize(file_path)
                category_results["file_size_bytes"] = file_size
                category_results["file_size_mb"] = file_size / (1024 * 1024)
                
                # 2. Test upload validation
                category_results["validation"] = self.test_upload_validation(file_path)
                
                # 3. Test Excel parsing (if validation passed)
                if category_results["validation"]["success"]:
                    category_results["parsing"] = self.test_excel_parsing(file_path)
                else:
                    category_results["parsing"] = {"success": False, "skipped": True}
                
                # 4. Test complexity analysis
                category_results["complexity_analysis"] = self.test_file_complexity_analysis(file_size)
                
                # 5. Memory check
                current_memory = self.get_memory_usage()
                category_results["memory_usage_mb"] = current_memory
                category_results["memory_increase_mb"] = current_memory - self.memory_baseline
                
                results["results"][category] = category_results
                
                print(f"📊 {category.upper()} SUMMARY:")
                print(f"   File size: {category_results['file_size_mb']:.1f} MB")
                print(f"   Validation: {'✅' if category_results['validation']['success'] else '❌'}")
                print(f"   Parsing: {'✅' if category_results['parsing']['success'] else '❌'}")
                print(f"   Memory increase: {category_results['memory_increase_mb']:.1f} MB")
                
            # 6. Test cleanup system
            results["cleanup"] = self.test_memory_cleanup()
            
            # Generate summary
            results["test_end_time"] = time.time()
            results["total_test_time"] = results["test_end_time"] - results["test_start_time"]
            
            successful_tests = sum(1 for cat in results["results"].values() 
                                 if cat["validation"]["success"] and cat["parsing"]["success"])
            total_tests = len(size_categories)
            
            results["summary"] = {
                "success_rate": successful_tests / total_tests,
                "successful_tests": successful_tests,
                "total_tests": total_tests,
                "max_memory_usage_mb": max(cat["memory_usage_mb"] for cat in results["results"].values()),
                "total_files_processed_mb": sum(cat["file_size_mb"] for cat in results["results"].values()),
                "average_processing_time": sum(
                    cat["parsing"].get("parsing_time", 0) for cat in results["results"].values()
                    if cat["parsing"]["success"]
                ) / max(1, successful_tests)
            }
            
            print("\n🎉 BIG FILE SMOKE TEST COMPLETED")
            print("=" * 60)
            print(f"✅ Success rate: {results['summary']['success_rate']:.1%} ({successful_tests}/{total_tests})")
            print(f"📊 Total files processed: {results['summary']['total_files_processed_mb']:.1f} MB")
            print(f"💾 Max memory usage: {results['summary']['max_memory_usage_mb']:.1f} MB")
            print(f"⏱️ Average processing time: {results['summary']['average_processing_time']:.3f}s per file")
            print(f"🧹 Cleanup system: {'✅' if results['cleanup']['success'] else '❌'}")
            
            return results
            
        except Exception as e:
            print(f"❌ Test suite failed: {e}")
            results["error"] = str(e)
            return results
            
        finally:
            self.teardown()


# Pytest test functions
def test_small_file_processing():
    """Test small file processing (5MB)."""
    suite = BigFileTestSuite()
    results = suite.run_comprehensive_test(["small"])
    
    assert results["summary"]["success_rate"] == 1.0, "Small file processing should succeed"
    assert results["results"]["small"]["memory_increase_mb"] < 100, "Memory usage should be reasonable"

def test_medium_file_processing():
    """Test medium file processing (25MB)."""
    suite = BigFileTestSuite()
    results = suite.run_comprehensive_test(["medium"])
    
    assert results["summary"]["success_rate"] == 1.0, "Medium file processing should succeed"
    assert results["results"]["medium"]["memory_increase_mb"] < 500, "Memory usage should be manageable"

def test_large_file_processing():
    """Test large file processing (100MB)."""
    suite = BigFileTestSuite()
    results = suite.run_comprehensive_test(["large"])
    
    # Large files might be more challenging, so we allow for some tolerance
    assert results["summary"]["success_rate"] >= 0.5, "Large file processing should mostly succeed"

def test_upload_size_limits():
    """Test upload size limit enforcement."""
    suite = BigFileTestSuite()
    
    # This should be close to our 500MB limit but still within it
    large_file = suite.create_test_excel_file("large")
    file_size = os.path.getsize(large_file)
    
    # Should pass validation
    validation_result = suite.test_upload_validation(large_file)
    assert validation_result["success"], "Large file within limits should pass validation"
    
    # File size should be reported correctly
    assert abs(validation_result["file_size_mb"] - (file_size / (1024*1024))) < 1, "File size should be accurate"

def test_cleanup_system():
    """Test file cleanup system."""
    suite = BigFileTestSuite()
    cleanup_result = suite.test_memory_cleanup()
    
    assert cleanup_result["success"], "Cleanup system should work"
    assert cleanup_result["cleanup_time"] < 10, "Cleanup should be fast"


if __name__ == "__main__":
    # Run smoke test directly
    suite = BigFileTestSuite()
    
    # Test with small and medium files by default
    results = suite.run_comprehensive_test(["small", "medium"])
    
    print("\n📋 DETAILED RESULTS:")
    print("=" * 60)
    
    import json
    print(json.dumps({
        "summary": results["summary"],
        "test_time": results["total_test_time"],
        "categories": list(results["results"].keys())
    }, indent=2)) 