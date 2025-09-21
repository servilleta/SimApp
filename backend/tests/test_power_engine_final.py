import pytest
import asyncio
import numpy as np
import openpyxl
import os
from typing import List

from simulation.power_engine import PowerMonteCarloEngine
from simulation.schemas import VariableConfig as MonteCarloVariable

# Mark all tests in this file as asyncio
pytestmark = pytest.mark.asyncio

@pytest.fixture
def temp_excel_file_for_power_engine():
    """Create a temporary Excel file for testing the Power Engine."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Define constants and formulas
    ws["A1"] = 10  # Constant for IF
    ws["A2"] = 5   # Constant for Arithmetic
    # C1 and D1 are MC inputs
    ws["E1"] = "=IF(C1>D1, C1*10, D1+A1)"  # Test IF, arithmetic in false branch
    ws["F1"] = "=C1+A2"                    # Test simple arithmetic

    filepath = "temp_power_engine_final_test.xlsx"
    wb.save(filepath)

    yield filepath, ws.title

    os.remove(filepath)

@pytest.fixture
def power_engine():
    """Provides a configured PowerMonteCarloEngine instance."""
    return PowerMonteCarloEngine(iterations=1000)


async def test_power_engine_refactored_if_logic(power_engine, temp_excel_file_for_power_engine):
    """Tests the refactored engine's IF logic, including GPU path if available."""
    filepath, sheet_name = temp_excel_file_for_power_engine
    
    variables = [
        MonteCarloVariable(
            name="C1",
            sheet_name=sheet_name,
            min_value=70,
            most_likely=100,
            max_value=130
        ),
        MonteCarloVariable(
            name="D1", 
            sheet_name=sheet_name,
            min_value=80,
            most_likely=90,
            max_value=100
        )
    ]
    
    target_cell = f"{sheet_name}!E1"
    
    results, errors, _ = await power_engine.run_simulation(
        file_path=filepath,
        file_id="test_power_final_if",
        target_cell=target_cell,
        variables=variables,
        iterations=power_engine.iterations,
        sheet_name=sheet_name
    )
    
    assert not errors, f"Simulation returned errors: {errors}"
    assert results is not None, "Simulation for IF logic returned None"
    assert results.shape == (power_engine.iterations,), f"Expected {power_engine.iterations} results, got {results.shape}"
    assert np.any(results > 0), "All results for IF were zero, indicating a problem."

    # Since C1's mean > D1's mean, most results should be C1*10.
    # We check if the mean is in the ballpark of 100 * 10 = 1000.
    mean_result = np.mean(results)
    assert 800 < mean_result < 1200, f"Mean result {mean_result} is outside the expected range for the IF condition."
    print(f"\n✅ Power Engine test for IF logic passed. Mean result: {mean_result:.2f}")


async def test_power_engine_refactored_arithmetic(power_engine, temp_excel_file_for_power_engine):
    """Tests the refactored engine's simple arithmetic logic."""
    filepath, sheet_name = temp_excel_file_for_power_engine
    
    variables = [
        MonteCarloVariable(
            name="C1",
            sheet_name=sheet_name,
            min_value=70,
            most_likely=100,
            max_value=130
        ),
        MonteCarloVariable(
            name="D1",
            sheet_name=sheet_name, 
            min_value=80,
            most_likely=90,
            max_value=100
        )  # Not used but needed for file consistency
    ]
    
    target_cell = f"{sheet_name}!F1" # =C1+A2 where A2=5
    
    results, errors, _ = await power_engine.run_simulation(
        file_path=filepath,
        file_id="test_power_final_arithmetic",
        target_cell=target_cell,
        variables=variables,
        iterations=power_engine.iterations,
        sheet_name=sheet_name
    )
    
    assert not errors, f"Simulation returned errors: {errors}"
    assert results is not None, "Simulation for arithmetic logic returned None"
    assert results.shape == (power_engine.iterations,), f"Expected {power_engine.iterations} results, got {results.shape}"
    
    # Expected mean is mean of C1 + 5.
    expected_mean = 100 + 5
    mean_result = np.mean(results)
    assert np.isclose(mean_result, expected_mean, rtol=0.1), f"Mean result {mean_result} is not close to expected {expected_mean}."
    print(f"✅ Power Engine test for arithmetic logic passed. Mean result: {mean_result:.2f}") 