#!/usr/bin/env python3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Create a simple Monte Carlo test file
def create_test_excel():
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Monte Carlo Test"
    
    # Add headers
    headers = ['Variable', 'Value', 'Distribution', 'Parameters']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add test data
    test_data = [
        ['Revenue', 1000, 'Normal', 'mean=1000, std=100'],
        ['Cost', 600, 'Uniform', 'min=500, max=700'],
        ['Price', 50, 'Triangular', 'min=40, mode=50, max=60'],
        ['Quantity', 20, 'Poisson', 'lambda=20'],
        ['Discount', 0.1, 'Beta', 'alpha=2, beta=8']
    ]
    
    for row, data in enumerate(test_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Add formulas
    ws['A7'] = 'Profit'
    ws['B7'] = '=B2*(1-B5)-B3'  # Revenue*(1-Discount)-Cost
    
    ws['A8'] = 'ROI'
    ws['B8'] = '=B7/B3'  # Profit/Cost
    
    # Add some random data for testing
    for i in range(10, 20):
        ws[f'A{i}'] = f'Test{i-9}'
        ws[f'B{i}'] = np.random.randint(100, 1000)
    
    # Save the file
    wb.save('test-data/simple-monte-carlo.xlsx')
    print("Created test Excel file: test-data/simple-monte-carlo.xlsx")

if __name__ == "__main__":
    create_test_excel() 