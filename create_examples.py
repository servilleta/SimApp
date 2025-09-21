#!/usr/bin/env python3
"""
Create example Excel files for Monte Carlo simulation.
These files demonstrate proper structure for the simulation platform.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os

def create_business_model():
    """Create a business revenue model example."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Business Model"
    
    # Header styling
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    
    # Input Variables Section
    ws['A1'] = "BUSINESS REVENUE MODEL"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A3'] = "INPUT VARIABLES"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:D3')
    
    # Input variables
    ws['A5'] = "Monthly Customers"
    ws['B5'] = 1000
    ws['C5'] = "customers"
    
    ws['A6'] = "Average Order Value"
    ws['B6'] = 45.50
    ws['C6'] = "$"
    
    ws['A7'] = "Conversion Rate"
    ws['B7'] = 0.025
    ws['C7'] = "%"
    
    ws['A8'] = "Monthly Marketing Cost"
    ws['B8'] = 5000
    ws['C8'] = "$"
    
    ws['A9'] = "Cost of Goods Sold %"
    ws['B9'] = 0.35
    ws['C9'] = "%"
    
    # Calculations Section
    ws['A11'] = "CALCULATIONS"
    ws['A11'].font = header_font
    ws['A11'].fill = header_fill
    ws.merge_cells('A11:D11')
    
    ws['A13'] = "Monthly Revenue"
    ws['B13'] = "=B5*B6*B7"
    ws['C13'] = "$"
    
    ws['A14'] = "Monthly COGS"
    ws['B14'] = "=B13*B9"
    ws['C14'] = "$"
    
    ws['A15'] = "Gross Profit"
    ws['B15'] = "=B13-B14"
    ws['C15'] = "$"
    
    ws['A16'] = "Net Profit"
    ws['B16'] = "=B15-B8"
    ws['C16'] = "$"
    
    ws['A17'] = "Profit Margin"
    ws['B17'] = "=B16/B13"
    ws['C17'] = "%"
    
    # Annual projections
    ws['A19'] = "ANNUAL PROJECTIONS"
    ws['A19'].font = header_font
    ws['A19'].fill = header_fill
    ws.merge_cells('A19:D19')
    
    ws['A21'] = "Annual Revenue"
    ws['B21'] = "=B13*12"
    ws['C21'] = "$"
    
    ws['A22'] = "Annual Net Profit"
    ws['B22'] = "=B16*12"
    ws['C22'] = "$"
    
    # Instructions
    ws['F1'] = "INSTRUCTIONS FOR MONTE CARLO SIMULATION"
    ws['F1'].font = Font(bold=True, size=12)
    ws.merge_cells('F1:J1')
    
    instructions = [
        "1. Input Variables (B5:B9) can be varied in simulation",
        "2. Target cells for analysis:",
        "   - Monthly Revenue (B13)",
        "   - Net Profit (B16)",
        "   - Annual Revenue (B21)",
        "   - Annual Net Profit (B22)",
        "",
        "3. Suggested distributions:",
        "   - Customers: Normal (μ=1000, σ=100)",
        "   - Order Value: Normal (μ=45.50, σ=5)",
        "   - Conversion: Beta (α=2, β=80)",
        "   - Marketing Cost: Uniform (4000, 6000)",
        "   - COGS %: Triangular (0.30, 0.35, 0.40)"
    ]
    
    for i, instruction in enumerate(instructions, start=3):
        ws[f'F{i}'] = instruction
        if instruction.startswith(("1.", "2.", "3.")):
            ws[f'F{i}'].font = Font(bold=True)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['F'].width = 35
    
    return wb

def create_project_costs():
    """Create a project cost analysis example."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Costs"
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    
    ws['A1'] = "PROJECT COST ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Cost Components
    ws['A3'] = "COST COMPONENTS"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:D3')
    
    ws['A5'] = "Development Hours"
    ws['B5'] = 500
    ws['C5'] = "hours"
    
    ws['A6'] = "Hourly Rate"
    ws['B6'] = 150
    ws['C6'] = "$/hour"
    
    ws['A7'] = "Equipment Cost"
    ws['B7'] = 25000
    ws['C7'] = "$"
    
    ws['A8'] = "Software Licenses"
    ws['B8'] = 8000
    ws['C8'] = "$"
    
    ws['A9'] = "Marketing Budget"
    ws['B9'] = 15000
    ws['C9'] = "$"
    
    ws['A10'] = "Contingency %"
    ws['B10'] = 0.15
    ws['C10'] = "%"
    
    # Calculations
    ws['A12'] = "TOTAL COSTS"
    ws['A12'].font = header_font
    ws['A12'].fill = header_fill
    ws.merge_cells('A12:D12')
    
    ws['A14'] = "Development Cost"
    ws['B14'] = "=B5*B6"
    ws['C14'] = "$"
    
    ws['A15'] = "Base Project Cost"
    ws['B15'] = "=B14+B7+B8+B9"
    ws['C15'] = "$"
    
    ws['A16'] = "Contingency Amount"
    ws['B16'] = "=B15*B10"
    ws['C16'] = "$"
    
    ws['A17'] = "Total Project Cost"
    ws['B17'] = "=B15+B16"
    ws['C17'] = "$"
    
    ws['A18'] = "Cost per Hour"
    ws['B18'] = "=B17/B5"
    ws['C18'] = "$/hour"
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    
    return wb

def create_investment_returns():
    """Create an investment returns analysis example."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Investment"
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
    
    ws['A1'] = "INVESTMENT RETURNS ANALYSIS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    # Investment Parameters
    ws['A3'] = "INVESTMENT PARAMETERS"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    ws.merge_cells('A3:D3')
    
    ws['A5'] = "Initial Investment"
    ws['B5'] = 100000
    ws['C5'] = "$"
    
    ws['A6'] = "Annual Return Rate"
    ws['B6'] = 0.08
    ws['C6'] = "%"
    
    ws['A7'] = "Investment Period"
    ws['B7'] = 10
    ws['C7'] = "years"
    
    ws['A8'] = "Annual Fees %"
    ws['B8'] = 0.015
    ws['C8'] = "%"
    
    ws['A9'] = "Inflation Rate"
    ws['B9'] = 0.025
    ws['C9'] = "%"
    
    # Returns Calculation
    ws['A11'] = "RETURNS CALCULATION"
    ws['A11'].font = header_font
    ws['A11'].fill = header_fill
    ws.merge_cells('A11:D11')
    
    ws['A13'] = "Gross Return Rate"
    ws['B13'] = "=B6-B8"
    ws['C13'] = "%"
    
    ws['A14'] = "Real Return Rate"
    ws['B14'] = "=B13-B9"
    ws['C14'] = "%"
    
    ws['A15'] = "Future Value (Nominal)"
    ws['B15'] = "=B5*(1+B13)^B7"
    ws['C15'] = "$"
    
    ws['A16'] = "Future Value (Real)"
    ws['B16'] = "=B5*(1+B14)^B7"
    ws['C16'] = "$"
    
    ws['A17'] = "Total Nominal Gain"
    ws['B17'] = "=B15-B5"
    ws['C17'] = "$"
    
    ws['A18'] = "Total Real Gain"
    ws['B18'] = "=B16-B5"
    ws['C18'] = "$"
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    
    return wb

def main():
    """Create all example files."""
    output_dir = "frontend/public/examples"
    os.makedirs(output_dir, exist_ok=True)
    
    # Create example files
    examples = [
        ("business-model.xlsx", create_business_model()),
        ("project-costs.xlsx", create_project_costs()),
        ("investment-returns.xlsx", create_investment_returns())
    ]
    
    for filename, workbook in examples:
        filepath = os.path.join(output_dir, filename)
        workbook.save(filepath)
        print(f"Created: {filepath}")
    
    print("\nExample Excel files created successfully!")
    print("These files demonstrate proper structure for Monte Carlo simulation:")
    print("- Clear input variables")
    print("- Formula-based calculations") 
    print("- Target cells for analysis")

if __name__ == "__main__":
    main()
